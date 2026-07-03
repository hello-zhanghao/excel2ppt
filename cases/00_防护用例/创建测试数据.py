"""
防护用例数据生成脚本
生成「测试数据.xlsx」和「项目配置.xlsx」到同目录
每次代码修改后运行此脚本重建数据，再执行 run_test.py 验证
"""
import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


# ============================================================
# 1. 生成测试数据
# ============================================================
def create_data_file():
    """创建测试数据文件，包含2个Sheet，覆盖所有核心特性"""
    data_path = os.path.join(SCRIPT_DIR, "测试数据.xlsx")

    # --- Sheet1: 销售明细 ---
    # 设计原则：3个地区×2个产品×2个季度 = 12行，足够覆盖所有聚合但数据量极小
    sales_data = [
        # 地区, 产品, 季度, 销售额, 销量, 客户数, 销售额_A, 销售额_B
        ("华东", "产品A", "Q1", 1200, 100, 15, 1200, 1200),
        ("华东", "产品A", "Q2", 1800, 150, 20, 1800, 1800),
        ("华东", "产品B", "Q1", 800,  60, 10, 800,  800),
        ("华东", "产品B", "Q2", 1000, 80,  12, 1000, 1000),
        ("华北", "产品A", "Q1", 900,  75,  8, 900,  900),
        ("华北", "产品A", "Q2", 1100, 90,  10, 1100, 1100),
        ("华北", "产品B", "Q1", 600,  50,  6, 600,  600),
        ("华北", "产品B", "Q2", 700,  55,  7, 700,  700),
        ("华南", "产品A", "Q1", 1500, 120, 18, 1500, 1500),
        ("华南", "产品A", "Q2", 2000, 160, 25, 2000, 2000),
        ("华南", "产品B", "Q1", 500,  40,  5, 500,  500),
        ("华南", "产品B", "Q2", 650,  50,  8, 650,  650),
    ]
    df_sales = pd.DataFrame(sales_data, columns=[
        "地区", "产品", "季度", "销售额", "销量", "客户数", "销售额_A", "销售额_B"
    ])

    # --- Sheet2: 产品目录（用于JOIN演示）---
    catalog_data = [
        ("产品A", "高端系列", 3000),
        ("产品B", "经济系列", 2000),
    ]
    df_catalog = pd.DataFrame(catalog_data, columns=[
        "产品", "产品系列", "基准价"
    ])

    # --- Sheet3: 城市销售（用于长标签X轴旋转测试，8个城市>6触发旋转）---
    city_data = [
        ("广州市", 1500, 120),
        ("深圳市", 1800, 150),
        ("珠海市", 800, 60),
        ("佛山市", 1100, 90),
        ("东莞市", 1300, 100),
        ("中山市", 700, 55),
        ("惠州市", 950, 75),
        ("江门市", 650, 50),
    ]
    df_city = pd.DataFrame(city_data, columns=["城市", "销售额", "销量"])

    # 写入Excel
    with pd.ExcelWriter(data_path, engine="openpyxl") as writer:
        df_sales.to_excel(writer, sheet_name="销售明细", index=False)
        df_catalog.to_excel(writer, sheet_name="产品目录", index=False)
        df_city.to_excel(writer, sheet_name="城市销售", index=False)

    print(f"[OK] 数据文件已生成: {data_path}")
    return data_path


# ============================================================
# 2. 生成配置文件
# ============================================================
def create_config_file():
    """创建配置文件，包含透视分析和PPT配置两个Sheet"""
    config_path = os.path.join(SCRIPT_DIR, "项目配置.xlsx")
    wb = openpyxl.Workbook()

    # ---------- 样式定义 ----------
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_data(ws, nrows, ncols):
        for row in range(2, nrows + 2):
            for col in range(1, ncols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

    def auto_width(ws, ncols, min_w=8, max_w=30):
        for col in range(1, ncols + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val and len(str(val)) > max_len:
                    max_len = len(str(val))
            ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 4, min_w), max_w)

    # ============================================================
    # Sheet1: 透视分析
    # ============================================================
    ws_pivot = wb.active
    ws_pivot.title = "透视分析"

    pivot_headers = [
        "序号", "数据源", "Sheet", "行维度", "列维度", "值字段",
        "聚合方式", "结果Sheet", "行映射", "值映射",
        "分箱", "值计算", "是否计算", "过滤条件", "区块名"
    ]

    # 透视分析任务（12个任务，覆盖所有核心特性）
    pivot_rows = [
        # 任务1: 简单分组求和（基础场景）
        [1, "测试数据.xlsx", "销售明细", "地区", "", "销售额",
         "sum", "按地区汇总", "", "总销售额",
         "", "", "是", "", "简单分组求和"],

        # 任务2: 多字段多聚合（单字段多聚合 + 多字段）
        [2, "测试数据.xlsx", "销售明细", "地区", "", "销售额,销售额,客户数",
         "sum,avg,count", "地区多维统计", "", "总销售额,平均销售额,客户数",
         "", "", "是", "", "多字段多聚合"],

        # 任务3: 交叉表（行×列维度）
        [3, "测试数据.xlsx", "销售明细", "地区", "产品", "销售额",
         "sum", "地区产品交叉", "", "",
         "", "", "是", "", "交叉表"],

        # 任务4: 占比聚合（pct）
        [4, "测试数据.xlsx", "销售明细", "地区", "", "销售额",
         "pct", "地区占比", "", "销售额占比",
         "", "", "是", "", "占比聚合"],

        # 任务5: 计数占比（count_pct）
        [5, "测试数据.xlsx", "销售明细", "产品", "", "客户数",
         "count_pct", "产品计数占比", "", "客户占比",
         "", "", "是", "", "计数占比"],

        # 任务6: 分箱（自定义列名）
        [6, "测试数据.xlsx", "销售明细", "销售额", "", "客户数",
         "count", "销售额分箱", "", "客户数",
         "销售额=500,1000,1500,2000|销售额区间", "", "是", "", "分箱统计"],

        # 任务7: 过滤条件
        [7, "测试数据.xlsx", "销售明细", "地区", "", "销售额",
         "sum", "华东华北汇总", "", "总销售额",
         "", "", "是", "地区 = '华东' OR 地区 = '华北'", "过滤条件"],

        # 任务8: JOIN + 值计算（单列除常数）
        [8, "测试数据.xlsx@销售明细 JOIN 测试数据.xlsx@产品目录 ON 产品=产品", "销售明细", "产品", "", "销售额",
         "sum", "产品系列汇总", "", "总销售额",
         "", "/1000", "是", "", "JOIN+值计算(单列)"],

        # 任务9: 多列组合计算（销售额/销量=单价）
        # 注意：值计算表达式需要使用值映射后的列名
        [9, "测试数据.xlsx", "销售明细", "地区", "", "销售额,销量",
         "sum,sum", "地区单价分析", "", "总销售额,总销量",
         "", "总销售额/总销量=单价(万元/个)", "是", "", "多列组合计算"],

        # 任务10: 跳过任务（是否计算=否）
        [10, "测试数据.xlsx", "销售明细", "季度", "", "销售额",
         "sum", "季度汇总", "", "总销售额",
         "", "", "否", "", "跳过任务演示"],

        # 任务11-12: block合并测试（行维度不同、映射相同、分箱相同）
        [11, "测试数据.xlsx", "销售明细", "销售额_A", "", "客户数",
         "count", "block合并测试", "销售额_A=销售额区间", "客户数_A",
         "销售额_A=500,1000,1500,2000", "", "是", "",
         "block合并测试"],

        [12, "测试数据.xlsx", "销售明细", "销售额_B", "", "客户数",
         "count", "block合并测试", "销售额_B=销售额区间", "客户数_B",
         "销售额_B=500,1000,1500,2000", "", "是", "",
         "block合并测试"],

        # 任务13: 无行维度 → 横向一行输出（位置1:1对应）
        [13, "测试数据.xlsx", "销售明细", "", "", "销售额,客户数",
         "sum,avg", "无行维度汇总", "", "总销售额,平均客户数",
         "", "", "是", "", "无行维度_横向一行"],

        # 任务14: 同区块名合并测试（与任务1"简单分组求和"同名，但中间隔了其他任务）
        [14, "测试数据.xlsx", "销售明细", "地区", "", "销量",
         "sum", "按地区汇总", "", "总销量",
         "", "", "是", "", "简单分组求和"],

        # 任务15: 无行维度 → 生产标量"总销售额"，供后续任务公式引用
        [15, "测试数据.xlsx", "销售明细", "", "", "销售额",
         "sum", "历史标量", "", "总销售额",
         "", "", "是", "", "标量生产者"],

        # 任务16: 值计算引用历史标量 → 销量/总销售额（来自任务15）
        [16, "测试数据.xlsx", "销售明细", "地区", "", "销量",
         "sum", "按地区汇总", "", "地区销量",
         "", "销量/总销售额=销量占比", "是", "", "标量消费者"],
    ]

    # 写入表头
    for col, header in enumerate(pivot_headers, 1):
        ws_pivot.cell(row=1, column=col, value=header)
    # 写入数据
    for row_idx, row_data in enumerate(pivot_rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws_pivot.cell(row=row_idx, column=col_idx, value=val)

    style_header(ws_pivot, len(pivot_headers))
    style_data(ws_pivot, len(pivot_rows), len(pivot_headers))
    auto_width(ws_pivot, len(pivot_headers))
    ws_pivot.freeze_panes = "A2"

    # ============================================================
    # Sheet2: PPT配置
    # ============================================================
    ws_ppt = wb.create_sheet("PPT配置")

    ppt_headers = [
        "页码", "页面类型", "页面标题", "布局",
        "图表类型", "数据Sheet", "X轴", "Y轴",
        "图表标题", "区块名", "结论模板", "数据源",
        "是否生成", "文字内容"
    ]

    # PPT页面（12页配置，其中页12设为"否"跳过，实际生成11页）
    ppt_rows = [
        # 页1: 封面（副标题用 | 语法合并到页面标题）
        [1, "封面", "销售数据分析报告\n防护用例|基础防护测试 | 全特性覆盖", "",
         "", "", "", "",
         "", "", "", "",
         "是", ""],

        # 页2: 目录
        [2, "目录", "目录", "",
         "", "", "", "",
         "", "", "", "",
         "是", ""],

        # 页3: 章节分隔
        [3, "章节", "一、销售汇总分析|各地区销售数据多维分析", "",
         "", "", "", "",
         "", "", "", "",
         "是", ""],

        # 页4: 单图柱状图（引用透视结果）
        [4, "内容", "各地区销售汇总", "1图",
         "column", "按地区汇总", "地区", "总销售额",
         "各地区总销售额", "", "销售额最高地区: {max_cat} ({max_val})", "透视结果",
         "是", ""],

        # 页5: 2图左右（柱状图+饼图）
        [5, "内容", "多维统计分析", "2图左右",
         "bar", "地区多维统计", "地区", "总销售额",
         "各地区销售额对比", "", "", "透视结果",
         "是", ""],
        # 第二个图表（同页）
        ["", "", "", "",
         "pie", "地区占比", "地区", "销售额占比",
         "各地区销售额占比", "", "", "",
         "是", ""],

        # 页6: 折线图（引用透视结果）
        [6, "内容", "销售额分箱分布", "1图",
         "line", "销售额分箱", "销售额区间", "客户数",
         "销售额区间客户分布", "", "共 {count} 个区间", "透视结果",
         "是", ""],

        # 页7: 左图右文 + 文字内容
        [7, "内容", "分析结论与说明", "左图右文",
         "", "", "", "",
         "", "本报告由防护用例自动生成", "", "",
         "是", "本报告基于测试数据生成，覆盖以下特性：\n1. 透视分析（分组、交叉表、占比、分箱）\n2. 多种图表类型（柱状图、饼图、折线图）\n3. 配置化PPT生成\n4. 结论模板自动填充\n5. 目录/章节/结尾页\n6. X轴长标签自动旋转"],

        # 页8: block合并测试
        [8, "内容", "block合并测试", "1图",
         "column", "block合并测试", "销售额区间", "客户数_A",
         "block合并测试(列A)", "", "验证行维度不同+映射相同=合并区块", "透视结果",
         "是", ""],

        # 页9: 章节分隔
        [9, "章节", "二、扩展测试|长标签与布局验证", "",
         "", "", "", "",
         "", "", "", "",
         "是", ""],

        # 页10: 长标签测试（8个城市>6触发X轴旋转）
        [10, "内容", "城市销售长标签测试", "1图",
         "column", "城市销售", "城市", "销售额",
         "各城市销售额对比", "", "销售额最高城市: {max_cat} ({max_val})", "测试数据.xlsx",
         "是", ""],

        # 页11: 结尾页
        [11, "结尾", "谢谢|防护用例测试完成", "",
         "", "", "", "",
         "", "", "", "",
         "是", ""],

        # 页12: 是否生成=否（测试跳过功能，不应出现在PPT中）
        [12, "内容", "跳过测试页(不应生成)", "1图",
         "column", "按地区汇总", "地区", "总销售额",
         "不应出现的图表", "", "", "透视结果",
         "否", ""],
    ]

    # 写入表头
    for col, header in enumerate(ppt_headers, 1):
        ws_ppt.cell(row=1, column=col, value=header)
    # 写入数据
    for row_idx, row_data in enumerate(ppt_rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws_ppt.cell(row=row_idx, column=col_idx, value=val)

    style_header(ws_ppt, len(ppt_headers))
    style_data(ws_ppt, len(ppt_rows), len(ppt_headers))
    auto_width(ws_ppt, len(ppt_headers))
    ws_ppt.freeze_panes = "A2"

    # 保存
    wb.save(config_path)
    print(f"[OK] 配置文件已生成: {config_path}")
    return config_path


# ============================================================
# 主函数
# ============================================================
if __name__ == "__main__":
    print("=" * 60)
    print("  防护用例数据生成")
    print("=" * 60)
    create_data_file()
    create_config_file()
    print("\n生成完成！接下来运行: python run_test.py")
