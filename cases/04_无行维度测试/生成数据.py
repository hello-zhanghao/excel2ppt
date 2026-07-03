import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import os

sa = "销售明细"
sheet = "透视分析"
out_dir = os.path.dirname(__file__)

# ========== 生成数据 ==========
data = [
    ["华东", "产品A", 1200, 100, 4],
    ["华东", "产品B", 800,  60,  3],
    ["华北", "产品A", 1500, 120, 5],
    ["华北", "产品B", 900,  80,  4],
    ["华南", "产品A", 2000, 150, 7],
    ["华南", "产品B", 1200, 90,  6],
]
df = pd.DataFrame(data, columns=["地区", "产品", "销售额", "销量", "客户数"])
data_path = os.path.join(out_dir, "测试数据.xlsx")
with pd.ExcelWriter(data_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name=sa, index=False)

print(f"[OK] 数据: {data_path}")

# ========== 生成配置 ==========
wb = openpyxl.Workbook()
ws = wb.active
ws.title = sheet

hdr_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
dfont = Font(name="微软雅黑", size=10)
dalign = Alignment(vertical="center", wrap_text=True)

headers = ["序号", "数据源", "Sheet", "行维度", "列维度", "值字段",
           "聚合方式", "结果Sheet", "行映射", "值映射",
           "分箱", "值计算", "是否计算", "过滤条件", "区块名"]

tasks = [
    # 有行维度：销售额sum, 客户数avg → 按地区分组
    [1, "测试数据.xlsx", sa, "地区", "", "销售额,客户数",
     "sum,avg", "按地区汇总", "", "总销售额,平均客户数",
     "", "", "是", "", "按地区分组汇总"],

    # 无行维度：销售额sum, 客户数avg → 横向一行（1:1 对应）
    [2, "测试数据.xlsx", sa, "", "", "销售额,客户数",
     "sum,avg", "总体汇总", "", "总销售额,平均客户数",
     "", "", "是", "", "无行维度汇总"],

    # 无行维度 + pct + count（单值多聚合全组合）
    [3, "测试数据.xlsx", sa, "", "", "销售额",
     "sum,avg,count", "总体统计", "", "总销售额,平均销售额,记录数",
     "", "", "是", "", "单值多聚合"],
]

for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = hdr_align
    c.border = border

for ri, row in enumerate(tasks, 2):
    for ci, val in enumerate(row, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = dfont
        c.alignment = dalign
        c.border = border

for col in range(1, len(headers) + 1):
    max_len = 0
    for row in range(1, len(tasks) + 2):
        v = ws.cell(row=row, column=col).value
        if v:
            max_len = max(max_len, len(str(v)))
    ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 4, 8), 30)

ws.freeze_panes = "A2"

cfg_path = os.path.join(out_dir, "项目配置.xlsx")
wb.save(cfg_path)
print(f"[OK] 配置: {cfg_path}")
print("生成完成，运行: cd app && python main.py pivot ../cases/04_无行维度测试/项目配置.xlsx")
