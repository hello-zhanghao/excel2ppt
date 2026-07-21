import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

TOTAL_FONT = Font(name="Microsoft YaHei", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

BLOCK_TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
BLOCK_TITLE_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
BLOCK_TITLE_ALIGNMENT = Alignment(horizontal="center", vertical="center")

DATA_FONT = Font(name="Microsoft YaHei", size=10)
DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# 行维度列配色（区分分组维度列与数值列）
DIM_HEADER_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
DIM_DATA_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def write_results(tasks, results, errors, output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if len(tasks) != len(results):
        print(f"    [警告] 任务数({len(tasks)})与结果数({len(results)})不一致，取较小值对齐")

    groups = {}
    for task, result in zip(tasks, results):
        if result is None:
            continue
        # 识别布局模式：结果Sheet名后缀控制整 sheet 布局
        #   |混合 / |mixed      → 混合模式（区块名 |横 标记横向区块，其余纵向）
        #   |横向 / |横 / |horizontal → 纯横向模式（整 sheet 横向并列）
        #   无后缀              → 纯纵向模式（默认，整 sheet 向下追加）
        raw_sheet = str(task.get("结果Sheet", "分析结果"))
        layout_mode = "vertical"  # 默认纵向
        for suffix in ("|混合", "|mixed", "|Mixed"):
            if raw_sheet.endswith(suffix):
                layout_mode = "mixed"
                raw_sheet = raw_sheet[:-len(suffix)].strip()
                break
        if layout_mode == "vertical":
            for suffix in ("|横向", "|横", "|horizontal", "|Horizontal"):
                if raw_sheet.endswith(suffix):
                    layout_mode = "horizontal"
                    raw_sheet = raw_sheet[:-len(suffix)].strip()
                    break
        task["结果Sheet"] = raw_sheet  # 原地更新为剥离后缀的真实 sheet 名
        sheet_name = _safe_sheet_name(raw_sheet)
        if sheet_name not in groups:
            groups[sheet_name] = {"items": [], "mode": layout_mode}
        groups[sheet_name]["items"].append((task, result))
        # 同 sheet 多任务时，模式优先级：mixed > horizontal > vertical
        # 任一任务标记更高模式则整 sheet 升级（避免模式冲突）
        mode_priority = {"vertical": 0, "horizontal": 1, "mixed": 2}
        if mode_priority[layout_mode] > mode_priority[groups[sheet_name]["mode"]]:
            groups[sheet_name]["mode"] = layout_mode

    for sheet_name, group in groups.items():
        if group["mode"] == "mixed":
            _write_mixed_layout_sheet(wb, sheet_name, group["items"])
        elif group["mode"] == "horizontal":
            _write_horizontal_layout_sheet(wb, sheet_name, group["items"])
        else:
            _write_multi_result_sheet(wb, sheet_name, group["items"])

    if errors:
        ws = wb.create_sheet("错误信息")
        ws.append(["任务", "错误详情"])
        for err in errors:
            ws.append([err.get("序号", ""), err.get("错误", "")])
        _style_header_row(ws)

    wb.save(output_path)
    return output_path


def _write_multi_result_sheet(wb, sheet_name, group_items):
    ws = wb.create_sheet(sheet_name)

    # 按「区块名」分组（非连续的相同区块名自动合并），按首次出现顺序输出
    block_groups = []
    seen_blocks = set()
    for task, result in group_items:
        block_name = _get_block_title(task)
        if block_name not in seen_blocks:
            block_groups.append((block_name, []))
            seen_blocks.add(block_name)
        for name, items in block_groups:
            if name == block_name:
                items.append((task, result))
                break

    current_row = 1
    for block_name, items in block_groups:
        row_dims = _get_row_dims(items[0][0])

        if not isinstance(items[0][1], dict):
            for task, result in items:
                if current_row > 1:
                    current_row += 1
                _write_block_title(ws, block_name, current_row, 2)
                current_row += 1
                _write_scalar_block(ws, task, result, block_name, current_row,
                                    task.get("_pct_columns", []), task.get("_number_formats", {}))
                current_row = ws.max_row + 1
            continue

        if current_row > 1:
            current_row += 1

        if len(items) > 1:
            merged_df, merged_pct_cols, merged_number_formats = _merge_same_dim_results(items, row_dims)
            if merged_df is None:
                continue
            _write_block_title(ws, block_name, current_row, len(merged_df.columns))
            current_row += 1
            _write_df_block(ws, merged_df, current_row, merged_pct_cols, row_dims, merged_number_formats)
            current_row = ws.max_row + 1
        else:
            task, result = items[0]
            pct_cols = task.get("_pct_columns", [])
            num_fmts = task.get("_number_formats", {})
            for key, df in result.items():
                if current_row > 1:
                    current_row += 1
                _write_block_title(ws, block_name, current_row, len(df.columns))
                current_row += 1
                _write_df_block(ws, df, current_row, pct_cols, row_dims, num_fmts)
                current_row = ws.max_row + 1

    if ws.max_column and ws.max_row:
        _auto_fit_columns(ws)
        ws.freeze_panes = ws.cell(row=1, column=1)


def _write_horizontal_layout_sheet(wb, sheet_name, group_items):
    """横向布局：同 sheet 内多个区块从左到右并列，区块间空 1 列分隔。

    适用于区块名/行维度各不相同但希望避免行数过多的场景。
    每个区块独立写入自己的列区域：顶部区块标题行（合并该区块所有列）→ 表头行 → 数据行。
    """
    ws = wb.create_sheet(sheet_name)

    # 按「区块名」分组（与 _write_multi_result_sheet 一致），保持首次出现顺序
    block_groups = []
    seen_blocks = set()
    for task, result in group_items:
        block_name = _get_block_title(task)
        if block_name not in seen_blocks:
            block_groups.append((block_name, []))
            seen_blocks.add(block_name)
        for name, items in block_groups:
            if name == block_name:
                items.append((task, result))
                break

    current_col = 1
    for block_name, items in block_groups:
        # 同名区块多任务时仍按行维度横向合并（复用现有合并逻辑）
        row_dims = _get_row_dims(items[0][0])

        if not isinstance(items[0][1], dict):
            # 标量区块：逐个任务横向排列
            for task, result in items:
                if current_col > 1:
                    current_col += 1  # 区块间空 1 列
                block_cols = 2  # 标量区块固定 2 列（指标名+值）
                _write_block_title_at(ws, block_name, 1, current_col, block_cols)
                _write_scalar_block_at(ws, task, result, 2, current_col,
                                       task.get("_pct_columns", []), task.get("_number_formats", {}))
                current_col += block_cols
            continue

        if current_col > 1:
            current_col += 1  # 区块间空 1 列

        if len(items) > 1:
            merged_df, merged_pct_cols, merged_number_formats = _merge_same_dim_results(items, row_dims)
            if merged_df is None:
                continue
            block_cols = len(merged_df.columns)
            _write_block_title_at(ws, block_name, 1, current_col, block_cols)
            _write_df_block_at(ws, merged_df, 2, current_col, merged_pct_cols, row_dims, merged_number_formats)
            current_col += block_cols
        else:
            task, result = items[0]
            pct_cols = task.get("_pct_columns", [])
            num_fmts = task.get("_number_formats", {})
            # 取第一个 DataFrame（与行追加模式一致）
            df_to_write = None
            for key, df in result.items():
                df_to_write = df
                break
            if df_to_write is None:
                continue
            block_cols = len(df_to_write.columns)
            _write_block_title_at(ws, block_name, 1, current_col, block_cols)
            _write_df_block_at(ws, df_to_write, 2, current_col, pct_cols, row_dims, num_fmts)
            current_col += block_cols

    if ws.max_column and ws.max_row:
        _auto_fit_columns(ws)
        ws.freeze_panes = ws.cell(row=1, column=1)


def _parse_block_direction(block_name):
    """解析区块名的方向后缀。返回 (is_horizontal, clean_name)。

    后缀：|横 / |横向 / |horizontal / |Horizontal
    用于混合布局模式下标记单个区块为横向排列。
    """
    for suffix in ("|横向", "|横", "|horizontal", "|Horizontal"):
        if block_name.endswith(suffix):
            return True, block_name[:-len(suffix)].strip()
    return False, block_name


def _write_mixed_layout_sheet(wb, sheet_name, group_items):
    """混合布局：同一 sheet 内横向段（上方）+ 纵向段（下方）。

    区块名带 |横 / |横向 / |horizontal 后缀的区块归入横向段，从左到右并列；
    其余区块归入纵向段，在横向段下方空 1 行后向下追加。
    适用于部分区块需要并列对比、部分区块需要独立展开的场景。
    """
    ws = wb.create_sheet(sheet_name)

    # 按「区块名」分组（与 _write_multi_result_sheet 一致），保持首次出现顺序
    block_groups = []
    seen_blocks = set()
    for task, result in group_items:
        block_name = _get_block_title(task)
        if block_name not in seen_blocks:
            block_groups.append((block_name, []))
            seen_blocks.add(block_name)
        for name, items in block_groups:
            if name == block_name:
                items.append((task, result))
                break

    # 按方向拆分（保留原始顺序）
    horizontal_blocks = []  # [(clean_name, items)]
    vertical_blocks = []
    for block_name, items in block_groups:
        is_horiz, clean_name = _parse_block_direction(block_name)
        if is_horiz:
            horizontal_blocks.append((clean_name, items))
        else:
            vertical_blocks.append((clean_name, items))

    current_row = 1

    # 横向段（上方）：所有横向区块从左到右并列
    if horizontal_blocks:
        current_col = 1
        segment_max_row = current_row
        for block_name, items in horizontal_blocks:
            row_dims = _get_row_dims(items[0][0])

            if current_col > 1:
                current_col += 1  # 区块间空 1 列

            if not isinstance(items[0][1], dict):
                # 标量区块：逐个任务横向排列
                for task, result in items:
                    if current_col > 1:
                        current_col += 1  # 任务间空 1 列
                    block_cols = 2  # 标量区块固定 2 列
                    _write_block_title_at(ws, block_name, current_row, current_col, block_cols)
                    end_row = _write_scalar_block_at(ws, task, result, current_row + 1, current_col,
                                                     task.get("_pct_columns", []),
                                                     task.get("_number_formats", {}))
                    segment_max_row = max(segment_max_row, end_row)
                    current_col += block_cols
            else:
                if len(items) > 1:
                    merged_df, merged_pct_cols, merged_number_formats = _merge_same_dim_results(items, row_dims)
                    if merged_df is None:
                        continue
                    block_cols = len(merged_df.columns)
                    _write_block_title_at(ws, block_name, current_row, current_col, block_cols)
                    end_row = _write_df_block_at(ws, merged_df, current_row + 1, current_col,
                                                 merged_pct_cols, row_dims, merged_number_formats)
                    segment_max_row = max(segment_max_row, end_row)
                    current_col += block_cols
                else:
                    task, result = items[0]
                    pct_cols = task.get("_pct_columns", [])
                    num_fmts = task.get("_number_formats", {})
                    df_to_write = None
                    for key, df in result.items():
                        df_to_write = df
                        break
                    if df_to_write is None:
                        continue
                    block_cols = len(df_to_write.columns)
                    _write_block_title_at(ws, block_name, current_row, current_col, block_cols)
                    end_row = _write_df_block_at(ws, df_to_write, current_row + 1, current_col,
                                                 pct_cols, row_dims, num_fmts)
                    segment_max_row = max(segment_max_row, end_row)
                    current_col += block_cols

        current_row = segment_max_row + 1

    # 纵向段（下方）：在横向段下方空 1 行后向下追加
    if vertical_blocks:
        if current_row > 1:
            current_row += 1  # 段间空 1 行

        for block_name, items in vertical_blocks:
            row_dims = _get_row_dims(items[0][0])

            if current_row > 1:
                current_row += 1  # 区块间空 1 行

            if not isinstance(items[0][1], dict):
                # 标量区块：逐个任务纵向追加
                for task, result in items:
                    if current_row > 1:
                        current_row += 1
                    _write_block_title_at(ws, block_name, current_row, 1, 2)
                    current_row += 1
                    end_row = _write_scalar_block_at(ws, task, result, current_row, 1,
                                                     task.get("_pct_columns", []),
                                                     task.get("_number_formats", {}))
                    current_row = end_row + 1
                continue

            if len(items) > 1:
                merged_df, merged_pct_cols, merged_number_formats = _merge_same_dim_results(items, row_dims)
                if merged_df is None:
                    continue
                _write_block_title_at(ws, block_name, current_row, 1, len(merged_df.columns))
                current_row += 1
                end_row = _write_df_block_at(ws, merged_df, current_row, 1,
                                             merged_pct_cols, row_dims, merged_number_formats)
                current_row = end_row + 1
            else:
                task, result = items[0]
                pct_cols = task.get("_pct_columns", [])
                num_fmts = task.get("_number_formats", {})
                for key, df in result.items():
                    if current_row > 1:
                        current_row += 1
                    _write_block_title_at(ws, block_name, current_row, 1, len(df.columns))
                    current_row += 1
                    end_row = _write_df_block_at(ws, df, current_row, 1,
                                                 pct_cols, row_dims, num_fmts)
                    current_row = end_row + 1

    if ws.max_column and ws.max_row:
        _auto_fit_columns(ws)
        ws.freeze_panes = ws.cell(row=1, column=1)


def _write_block_title_at(ws, title, row, start_col, col_count):
    """在指定行、指定起始列写入区块标题（合并单元格，深蓝底白字）。"""
    end_col = start_col + col_count - 1
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.font = BLOCK_TITLE_FONT
    cell.fill = BLOCK_TITLE_FILL
    cell.alignment = BLOCK_TITLE_ALIGNMENT
    cell.border = THIN_BORDER
    if col_count > 1:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    for ci in range(start_col + 1, end_col + 1):
        c = ws.cell(row=row, column=ci)
        c.fill = BLOCK_TITLE_FILL
        c.border = THIN_BORDER


def _write_df_block_at(ws, df, start_row, start_col, pct_columns=None, row_dims=None, number_formats=None):
    """写入 DataFrame 区块到指定起始行、起始列（横向布局用，支持 start_col 偏移）。

    :return: 区块结束行号（含数据行最后一行）
    """
    pct_columns = set(pct_columns or [])
    row_dims = set(row_dims or [])
    number_formats = number_formats or {}
    headers = list(df.columns)
    # 表头行
    for ci, h in enumerate(headers):
        ws.cell(row=start_row, column=start_col + ci, value=str(h))
    _style_header_row_at(ws, start_row, start_col, len(headers), row_dims)
    # v2.54.18+ 预推导每列默认格式（依据原始数据小数位）
    # v2.54.19+ 百分比列不再固定 PCT_NUMBER_FORMAT，统一走"配置格式 > 推导格式"
    inferred_fmts = {}
    for col_name in headers:
        col_name_str = str(col_name) if col_name is not None else ""
        if col_name_str in number_formats:
            continue
        inferred_fmts[col_name_str] = _infer_column_number_format(df[col_name])
    # 数据行
    for ri, (idx, row_data) in enumerate(df.iterrows()):
        row_num = start_row + 1 + ri
        is_total = str(idx) == "合计" or str(idx) == "总计"
        for ci, val in enumerate(row_data):
            col_name = headers[ci] if ci < len(headers) else None
            col_name_str = str(col_name) if col_name is not None else ""
            cell = ws.cell(row=row_num, column=start_col + ci, value=_format_cell_value(val, col_name, False))
            excel_fmt = _ppt_fmt_to_excel_fmt(number_formats.get(col_name_str, ""))
            if excel_fmt:
                cell.number_format = excel_fmt
            elif isinstance(val, (int, float)):
                cell.number_format = inferred_fmts.get(col_name_str) or VALID_NUMBER_FORMAT
            cell.alignment = DATA_ALIGNMENT
            cell.font = TOTAL_FONT if is_total else DATA_FONT
            if is_total:
                cell.fill = TOTAL_FILL
            elif col_name_str in row_dims:
                cell.fill = DIM_DATA_FILL
            else:
                cell.fill = PatternFill()
            cell.border = THIN_BORDER
    # 结束行 = 表头 + 数据行数
    return start_row + len(df)


def _write_scalar_block_at(ws, task, result, start_row, start_col, pct_columns=None, number_formats=None):
    """写入标量区块到指定起始行、起始列（横向布局用）。

    :return: 区块结束行号（含最后一行数据）
    """
    pct_columns = set(pct_columns or [])
    number_formats = number_formats or {}
    if isinstance(result, dict):
        items = list(result.items())
    else:
        items = [("指标", result)]
    for ri, (key, val) in enumerate(items):
        row_num = start_row + ri
        # 指标名列
        cell_name = ws.cell(row=row_num, column=start_col, value=str(key))
        cell_name.font = DATA_FONT
        cell_name.alignment = DATA_ALIGNMENT
        cell_name.border = THIN_BORDER
        cell_name.fill = DIM_DATA_FILL
        # 值列
        cell_val = ws.cell(row=row_num, column=start_col + 1, value=_format_cell_value(val, key, False))
        excel_fmt = _ppt_fmt_to_excel_fmt(number_formats.get(str(key), ""))
        if excel_fmt:
            cell_val.number_format = excel_fmt
        elif isinstance(val, (int, float)):
            # v2.54.18+ 标量场景按单值推导
            cell_val.number_format = _infer_column_number_format([val]) or VALID_NUMBER_FORMAT
        cell_val.font = DATA_FONT
        cell_val.alignment = DATA_ALIGNMENT
        cell_val.border = THIN_BORDER
    return start_row + len(items) - 1 if items else start_row


def _style_header_row_at(ws, row, start_col, max_col, dim_cols=None):
    """为指定行、指定起始列的表头应用样式（横向布局用）。"""
    dim_cols = set(dim_cols or [])
    end_col = start_col + max_col - 1
    for ci in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=ci)
        cell.font = HEADER_FONT
        cell.fill = DIM_HEADER_FILL if (cell.value and str(cell.value) in dim_cols) else HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _get_block_title(task):
    """获取区块标题：优先用区块名，否则用任务序号"""
    title = task.get("区块名", task.get("备注"))
    if title is not None and str(title).strip():
        return str(title).strip()
    seq = task.get("序号", "?")
    return f"任务{seq}"


def _write_block_title(ws, title, row, col_count):
    """在指定行写入区块标题（合并单元格，深蓝底白字）"""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = BLOCK_TITLE_FONT
    cell.fill = BLOCK_TITLE_FILL
    cell.alignment = BLOCK_TITLE_ALIGNMENT
    cell.border = THIN_BORDER
    if col_count > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    for ci in range(2, col_count + 1):
        c = ws.cell(row=row, column=ci)
        c.fill = BLOCK_TITLE_FILL
        c.border = THIN_BORDER


def _get_row_dims(task):
    raw = str(task.get("行维度", "")).strip()
    if not raw:
        return []
    return [d.strip() for d in raw.split(",") if d.strip()]


def _merge_same_dim_results(items, row_dims):
    merged = None
    merged_pct_cols = set()
    merged_number_formats = {}  # {列名: Excel数字格式}

    for task, result in items:
        if not isinstance(result, dict):
            continue
        task_pct_cols = set(task.get("_pct_columns", []))
        task_number_formats = task.get("_number_formats", {}) or {}
        for key, df in result.items():
            if df is None:
                continue
            df = df.copy()

            if merged is None:
                merged = df
                # 记录本 df 中属于 pct 的列
                for c in df.columns:
                    col_str = str(c)
                    if col_str in task_pct_cols:
                        merged_pct_cols.add(col_str)
                    if col_str in task_number_formats:
                        merged_number_formats[col_str] = task_number_formats[col_str]
            else:
                dup = [c for c in df.columns if c in merged.columns]
                df_rest = df.drop(columns=dup, errors="ignore")
                for c in df_rest.columns:
                    col_str = str(c)
                    if col_str in task_pct_cols:
                        merged_pct_cols.add(col_str)
                    if col_str in task_number_formats:
                        merged_number_formats[col_str] = task_number_formats[col_str]
                merged = pd.concat([merged.reset_index(drop=True), df_rest.reset_index(drop=True)], axis=1)

    return merged, list(merged_pct_cols), merged_number_formats


def _write_df_block(ws, df, start_row, pct_columns=None, row_dims=None, number_formats=None):
    pct_columns = set(pct_columns or [])
    row_dims = set(row_dims or [])
    number_formats = number_formats or {}  # {列名: PPT格式串}，需转换为 Excel number_format
    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=ci, value=str(h))
    _style_header_row(ws, start_row, len(headers), row_dims)

    # v2.54.18+ 预推导每列默认格式（依据原始数据小数位），替代硬编码 0.00
    # v2.54.19+ 百分比列不再固定 PCT_NUMBER_FORMAT，统一走"配置格式 > 推导格式"
    inferred_fmts = {}
    for col_name in headers:
        col_name_str = str(col_name) if col_name is not None else ""
        if col_name_str in number_formats:
            continue  # 用户已配置格式，跳过推导
        inferred_fmts[col_name_str] = _infer_column_number_format(df[col_name])

    for ri, (idx, row_data) in enumerate(df.iterrows()):
        row_num = start_row + 1 + ri
        for ci, val in enumerate(row_data, 1):
            col_name = headers[ci - 1] if ci <= len(headers) else None
            col_name_str = str(col_name) if col_name is not None else ""
            cell = ws.cell(row=row_num, column=ci, value=_format_cell_value(val, col_name, False))
            # v2.54.19+ 优先级：自定义格式（聚合方式|PPT格式）> 推导格式 > 0.00 兜底
            # 百分比列不再强制 PCT_NUMBER_FORMAT，用户想百分比显示需配置 pct|.1% 等
            excel_fmt = _ppt_fmt_to_excel_fmt(number_formats.get(col_name_str, ""))
            if excel_fmt:
                cell.number_format = excel_fmt
            elif isinstance(val, (int, float)):
                cell.number_format = inferred_fmts.get(col_name_str) or VALID_NUMBER_FORMAT

        is_total = str(idx) == "合计" or str(idx) == "总计"
        for ci in range(1, len(headers) + 1):
            col_name = headers[ci - 1] if ci <= len(headers) else ""
            is_dim = col_name in row_dims
            cell = ws.cell(row=row_num, column=ci)
            cell.alignment = DATA_ALIGNMENT
            cell.font = TOTAL_FONT if is_total else DATA_FONT
            if is_total:
                cell.fill = TOTAL_FILL
            elif is_dim:
                cell.fill = DIM_DATA_FILL
            else:
                cell.fill = PatternFill()
            cell.border = THIN_BORDER


def _write_scalar_block(ws, task, result, remark, start_row, pct_columns=None, number_formats=None):
    pct_columns = set(pct_columns or [])
    number_formats = number_formats or {}  # {指标名: PPT格式串}，需转换为 Excel number_format
    ws.cell(row=start_row, column=1, value="指标")
    ws.cell(row=start_row, column=2, value="值")
    _style_header_row(ws, start_row, 2)

    row = start_row + 1
    for key, val in result.items():
        key_str = str(key)
        c1 = ws.cell(row=row, column=1, value=key)
        c2 = ws.cell(row=row, column=2, value=_format_cell_value(val, key, False))
        # v2.54.19+ 优先级：自定义格式 > 推导格式 > 0.00 兜底（百分比不再强制）
        excel_fmt = _ppt_fmt_to_excel_fmt(number_formats.get(key_str, ""))
        if excel_fmt:
            c2.number_format = excel_fmt
        elif isinstance(val, (int, float)):
            # v2.54.18+ 标量场景按单值推导（整数→'0'，浮点→按原值小数位）
            c2.number_format = _infer_column_number_format([val]) or VALID_NUMBER_FORMAT
        c1.alignment = DATA_ALIGNMENT
        c2.alignment = DATA_ALIGNMENT
        c1.border = THIN_BORDER
        c2.border = THIN_BORDER
        row += 1
def _safe_sheet_name(name):
    if not name or not str(name).strip():
        return "分析结果"
    name = str(name)
    illegal = r"[]:*?/\\"
    for ch in illegal:
        name = name.replace(ch, "_")
    return name[:31]


def _style_header_row(ws, row=1, max_col=None, dim_cols=None):
    dim_cols = set(dim_cols or [])
    if max_col is None:
        for cell in ws[row]:
            col_name = str(cell.value) if cell.value else ""
            cell.font = HEADER_FONT
            cell.fill = DIM_HEADER_FILL if col_name in dim_cols else HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
        return
    for ci in range(1, max_col + 1):
        cell = ws.cell(row=row, column=ci)
        col_name = str(cell.value) if cell.value else ""
        cell.font = HEADER_FONT
        cell.fill = DIM_HEADER_FILL if col_name in dim_cols else HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


# 百分比列名兜底关键词（仅当 task 未携带 _pct_columns 时使用）
# 注意："率"过于宽泛（功率/速率/频率等非百分比字段），已移除
PCT_KEYWORDS = ["占比", "pct", "百分比", "比例"]


def _is_pct_col(col_name, pct_columns=None):
    """判断列是否为百分比列。
    优先用 task 携带的 pct_columns 元信息（精确），列名关键词仅作兜底。
    """
    if col_name is None:
        return False
    name = str(col_name)
    if pct_columns and name in pct_columns:
        return True
    return any(kw in name for kw in PCT_KEYWORDS)


def _format_cell_value(val, col_name=None, is_pct=False):
    if val is None or (isinstance(val, float) and str(val) == "nan"):
        return ""
    if isinstance(val, float) and val == int(val):
        return int(val)
    return val

VALID_NUMBER_FORMAT = '0.00'
PCT_NUMBER_FORMAT = '0.0%'

# v2.54.18+ 默认格式推导上限：避免聚合产生的浮点噪声（如 avg 65.33333333333333）导致显示位数过长
_INFER_MAX_DECIMALS = 6


def _infer_column_number_format(series):
    """根据列的原始数值推导默认 Excel number_format。

    v2.54.18+ 替代原来硬编码的 VALID_NUMBER_FORMAT ('0.00')，让默认显示位数
    依据数据本身的小数位推导，避免整数也显示成 1200.00、1位小数也显示成 65.30。

    推导规则：
      - 所有非空值都是整数（含 1200.0 这种）→ '0'（整数格式）
      - 含浮点数 → 取最大小数位 N（上限 _INFER_MAX_DECIMALS=6）→ '0.' + '0'*N
      - 空列或全非数值 → None（不设置 number_format）

    注意：百分比列不在此推导（由 _is_pct_col + PCT_NUMBER_FORMAT 单独处理）。

    :param series: pandas Series 或可迭代的列数据
    :return: Excel number_format 字符串，或 None
    """
    numeric_vals = []
    for v in series:
        if v is None:
            continue
        if isinstance(v, bool):
            continue
        # pandas NaN/NA
        try:
            if pd.isna(v):
                continue
        except (TypeError, ValueError):
            pass
        if isinstance(v, (int, float)):
            numeric_vals.append(float(v))
        else:
            # 尝试转数值（字符串数字等）
            try:
                numeric_vals.append(float(v))
            except (ValueError, TypeError):
                continue
    if not numeric_vals:
        return None
    # 所有值都是整数（含 65.0 这种）→ 整数格式
    if all(v == int(v) for v in numeric_vals):
        return '0'
    # 取最大小数位（用 repr 避免 str 对某些浮点的四舍五入）
    max_decimals = 0
    for v in numeric_vals:
        if v == int(v):
            continue
        # repr(float) 在 Python 3 给出最短可往返表示，如 65.3 → '65.3' 而非 '65.29999...'
        s = repr(v)
        if 'e' in s.lower() or 'E' in s:
            # 科学计数法，跳过（罕见，通常不会出现在聚合结果中）
            continue
        if '.' in s:
            decimals = len(s.split('.')[1])
            max_decimals = max(max_decimals, decimals)
    # 上限保护，避免浮点噪声
    max_decimals = min(max_decimals, _INFER_MAX_DECIMALS)
    if max_decimals == 0:
        return '0'
    return '0.' + '0' * max_decimals


def _ppt_fmt_to_excel_fmt(ppt_fmt):
    """将 PPT 占位符的格式串（Python format spec 风格）转换为 Excel number_format。

    与 template_filler.py 的 _KNOWN_FMTS 语法一致，聚合方式的 |格式 后缀复用同一套语法。

    映射规则：
        .Nf       → 0.00（N 位小数，如 .2f → 0.00，.0f → 0）
        int / d   → 0（整数）
        .N%       → 0.00%（N 位小数百分比，如 .2% → 0.00%）
        ,.Nf      → #,##0.00（千分位 + N 位小数）
        ,.N%      → #,##0.00%（千分位 + 百分比）
        .Ne/.NE   → 0.00E+00（科学计数法）

    :param ppt_fmt: PPT 格式串（如 ".2f"、".1%"、",.0f"）
    :return: Excel number_format 字符串。无法识别时返回 None（用默认）。
    """
    if not ppt_fmt:
        return None
    fmt = ppt_fmt.strip()

    # 整数格式
    if fmt in ("int", "d"):
        return "0"

    # 千分位 + 百分比：,.N%
    if fmt.startswith(",.") and fmt.endswith("%"):
        try:
            digits = int(fmt[2:-1])
            return f"#,##0.{'0' * digits}%"
        except ValueError:
            return None

    # 千分位 + 小数：,.Nf
    if fmt.startswith(",.") and fmt.endswith("f"):
        try:
            digits = int(fmt[2:-1])
            return f"#,##0.{'0' * digits}" if digits > 0 else "#,##0"
        except ValueError:
            return None

    # 百分比：.N%
    if fmt.startswith(".") and fmt.endswith("%"):
        try:
            digits = int(fmt[1:-1])
            return f"0.{'0' * digits}%" if digits > 0 else "0%"
        except ValueError:
            return None

    # 小数：.Nf
    if fmt.startswith(".") and fmt.endswith("f"):
        try:
            digits = int(fmt[1:-1])
            return f"0.{'0' * digits}" if digits > 0 else "0"
        except ValueError:
            return None

    # 科学计数法：.Ne / .NE
    if fmt.startswith(".") and (fmt.endswith("e") or fmt.endswith("E")):
        try:
            digits = int(fmt[1:-1])
            zero_str = "0" * digits
            return f"0.{zero_str}E+00" if digits > 0 else "0E+00"
        except ValueError:
            return None

    return None


def _auto_fit_columns(ws):
    for col_cells in ws.columns:
        if not col_cells or len(col_cells) == 0:
            continue
        col_letter = get_column_letter(col_cells[0].column)
        max_length = 0
        for cell in col_cells:
            if cell.value:
                val_str = str(cell.value)
                length = 0
                for ch in val_str:
                    if "\u4e00" <= ch <= "\u9fff" or "\u3000" <= ch <= "\u303f" or "\uff00" <= ch <= "\uffef":
                        length += 2
                    else:
                        length += 1
                max_length = max(max_length, length)
        adjusted = min(max_length + 3, 40)
        ws.column_dimensions[col_letter].width = max(adjusted, 8)
