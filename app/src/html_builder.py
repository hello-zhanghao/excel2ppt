import os
import sys
import json
import base64
from datetime import datetime
from typing import List, Dict, Optional, Tuple

HUAWEI_PALETTE = [
    "#C8102E", "#182B49", "#5B9BD5", "#ED7D31", 
    "#70AD47", "#A5A5A5", "#FFC000", "#4472C4"
]

def read_excel_data(excel_path: str) -> List[Dict]:
    try:
        import openpyxl
    except ImportError:
        print("[警告] 缺少 openpyxl，无法读取 Excel 数据")
        return []

    if not os.path.exists(excel_path):
        return []

    sections = []
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        _parse_blocks(sections, rows, sheet_name)
    wb.close()
    return sections


def _is_block_title_row(row) -> bool:
    """判断是否为区块标题行：只有第一个单元格有值，其余为空（合并单元格效果）"""
    vals = [c for c in row if c is not None]
    return len(vals) == 1 and row[0] is not None


def _is_header_row(row) -> bool:
    """判断是否为表头行：至少2个非None值（且不是区块标题行）"""
    vals = [c for c in row if c is not None]
    return len(vals) >= 2


def _extract_data(rows, header_idx, col_indices, headers):
    """从指定行提取数据行"""
    data_rows = []
    for row in rows:
        vals = []
        for i in col_indices:
            v = row[i] if i < len(row) else None
            if v is None:
                vals.append("")
            elif isinstance(v, float):
                vals.append(f"{v:.4f}" if v != int(v) else str(int(v)))
            else:
                vals.append(str(v))
        if any(v.strip() for v in vals):
            data_rows.append(vals)
    return data_rows


def _parse_blocks(sections, rows, sheet_name):
    """解析一个 sheet 内的所有区块，每个区块一个 section。"""
    n = len(rows)
    i = 0

    while i < n:
        # 跳过空行和非区块标题行/非表头行
        if i < n - 1:
            row = rows[i]
            next_row = rows[i + 1]
            if _is_block_title_row(row) and _is_header_row(next_row):
                # 区块标题行 + 表头行 → 新区块
                block_name = str(row[0]) if row[0] else sheet_name

                headers = [str(c) if c is not None else "" for c in next_row]
                col_indices = [ci for ci, h in enumerate(headers) if h.strip()]
                headers = [headers[ci] for ci in col_indices]

                data_start = i + 2
                data_end = data_start
                while data_end < n:
                    if data_end < n - 1 and _is_block_title_row(rows[data_end]) and _is_header_row(rows[data_end + 1]):
                        break
                    data_end += 1

                data_rows = _extract_data(rows[data_start:data_end], i + 1, col_indices, headers)

                sections.append({
                    "name": block_name,
                    "sheet": sheet_name,
                    "headers": headers,
                    "rows": data_rows,
                })
                i = data_end
                continue
            elif _is_header_row(row) and not _is_block_title_row(row):
                # 无区块标题，直接是表头 → 一个区块，区块名用 sheet 名
                headers = [str(c) if c is not None else "" for c in row]
                col_indices = [ci for ci, h in enumerate(headers) if h.strip()]
                headers = [headers[ci] for ci in col_indices]

                data_start = i + 1
                data_end = data_start
                while data_end < n:
                    if data_end < n - 1 and _is_block_title_row(rows[data_end]) and _is_header_row(rows[data_end + 1]):
                        break
                    data_end += 1

                data_rows = _extract_data(rows[data_start:data_end], i, col_indices, headers)

                sections.append({
                    "name": sheet_name,
                    "sheet": sheet_name,
                    "headers": headers,
                    "rows": data_rows,
                })
                i = data_end
                continue
        i += 1

    # 兜底：原逻辑，找第一个表头
    if not sections:
        header_idx = 0
        for ri, row in enumerate(rows[:3]):
            if _is_header_row(row):
                header_idx = ri
                break
        headers = [str(c) if c is not None else "" for c in rows[header_idx]]
        col_indices = [ci for ci, h in enumerate(headers) if h.strip()]
        headers = [headers[ci] for ci in col_indices]
        data_rows = _extract_data(rows[header_idx + 1:], header_idx, col_indices, headers)
        sections.append({
            "name": sheet_name,
            "headers": headers,
            "rows": data_rows,
        })


def _detect_geo_columns(headers: List[str]) -> Optional[Tuple[int, int]]:
    """检测经纬度列。返回 (lat_col_idx, lon_col_idx) 或 None。"""
    lat_keys = ["纬度", "lat", "latitude"]
    lon_keys = ["经度", "lon", "lng", "longitude"]
    lat_idx = None
    lon_idx = None
    for i, h in enumerate(headers):
        h_lower = h.lower().strip()
        if lat_idx is None and any(k in h_lower for k in lat_keys):
            lat_idx = i
        if lon_idx is None and any(k in h_lower for k in lon_keys):
            lon_idx = i
    if lat_idx is not None and lon_idx is not None:
        return (lat_idx, lon_idx)
    return None


def _infer_chart_type(headers: List[str], rows: List[List[str]]) -> str:
    # 优先：含经纬度列 → 地图
    if _detect_geo_columns(headers) is not None:
        return "map"

    # 标量数据：只有1行数据（无行维度，横向一行）→ 不画图，用指标卡片展示
    if len(rows) == 1:
        return "scalar"

    numeric_cols = 0
    category_cols = 0
    for h in headers:
        h_lower = h.lower()
        if any(key in h_lower for key in ["金额", "数量", "销售额", "利润", "成本", "值", "计数", "总计", "合计"]):
            numeric_cols += 1
        else:
            category_cols += 1

    total_rows = len(rows)

    for h in headers:
        h_lower = h.lower()
        if any(key in h_lower for key in ["占比", "百分比", "pct"]):
            return "pie"
        if any(key in h_lower for key in ["日期", "月份", "时间", "季度", "年"]):
            return "line"

    if total_rows > 20:
        return "table"

    if numeric_cols >= 2 and category_cols == 1:
        return "bar"

    if numeric_cols == 1 and category_cols == 1:
        return "bar"

    if numeric_cols == 1 and category_cols >= 2:
        return "bar"

    return "table"


def _compute_summary(headers: List[str], rows: List[List[str]]) -> Dict:
    summary = {}
    numeric_indices = []
    
    for i, h in enumerate(headers):
        h_lower = h.lower()
        if any(key in h_lower for key in ["金额", "数量", "销售额", "利润", "成本", "值", "占比", "百分比", "pct", "计数"]):
            numeric_indices.append(i)
    
    for idx in numeric_indices:
        col_name = headers[idx]
        values = []
        for row in rows:
            if idx < len(row) and row[idx] != "":
                try:
                    values.append(float(row[idx]))
                except ValueError:
                    pass
        
        if values:
            summary[col_name] = {
                "sum": round(sum(values), 2),
                "avg": round(sum(values) / len(values), 2),
                "max": round(max(values), 2),
                "min": round(min(values), 2),
                "count": len(values),
            }
    
    for row in rows:
        if any("合计" in str(v) or "总计" in str(v) for v in row):
            for i, v in enumerate(row):
                if i < len(headers):
                    try:
                        val = float(v) if v and str(v) != "" else None
                        if val is not None:
                            summary[f"合计_{headers[i]}"] = val
                    except ValueError:
                        pass
    
    return summary


def _escape_html(text: str) -> str:
    if text is None:
        return ""
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _format_num(val, key: str = "") -> str:
    """格式化数值用于 summary 显示。
    - 列名/key 含 占比/百分比/pct 且值在 0-1 之间 → 百分比 8.00%
    - 绝对值 >= 100000000 → 亿（如 1.50亿）
    - 绝对值 >= 10000 → 万（如 1.28万）
    - 否则千分位分隔，保留2位小数（.00 时显示整数）
    """
    if val is None or val == "":
        return str(val) if val == "" else ""
    try:
        num = float(val)
    except (ValueError, TypeError):
        return str(val)

    key_lower = str(key).lower()
    is_pct = any(k in key_lower for k in ["占比", "百分比", "pct"])
    if is_pct and 0 <= num <= 1:
        return f"{num*100:.2f}%"

    abs_num = abs(num)
    if abs_num >= 100000000:
        return f"{num/100000000:.2f}亿"
    if abs_num >= 10000:
        return f"{num/10000:.2f}万"

    s = f"{num:,.2f}"
    if s.endswith(".00"):
        s = s[:-3]
    return s


def _is_numeric_cell(v) -> bool:
    """判断单元格内容是否为数字（用于表格数字列右对齐）。"""
    if v is None or v == "":
        return False
    try:
        float(str(v).replace(",", "").replace("%", "").replace("亿", "").replace("万", ""))
        return True
    except (ValueError, TypeError):
        return False


def _summary_icon(col_name: str) -> str:
    """根据列名智能匹配 emoji 图标。"""
    n = str(col_name).lower()
    if "金额" in n:
        return "💰"
    if "数量" in n:
        return "📦"
    if "销售额" in n:
        return "📈"
    if "利润" in n:
        return "💹"
    if "占比" in n or "百分比" in n or "pct" in n:
        return "📊"
    if "计数" in n:
        return "#"
    return "📌"


def _detect_chart_columns(headers: List[str]) -> Tuple[List[int], List[int]]:
    """从表头自动检测分类列和数值列。返回 (category_cols: List[int], value_cols: List[int])。
    支持多列分类（多维度透视），category_cols = 非数值列（排除经纬度列）。
    """
    geo = _detect_geo_columns(headers)
    geo_set = set(geo) if geo else set()

    category_cols = []
    value_cols = []

    for i, h in enumerate(headers):
        if i in geo_set:
            continue
        h_lower = h.lower()
        if any(key in h_lower for key in ["金额", "数量", "销售额", "利润", "成本", "值", "占比", "百分比", "pct", "计数", "总计", "合计"]):
            value_cols.append(i)
        else:
            category_cols.append(i)

    # 没有数值列时，第一列作为分类，其余作为数值
    if not value_cols:
        if headers:
            category_cols = [0]
            value_cols = [i for i in range(1, len(headers)) if i not in geo_set]
        else:
            category_cols = []
            value_cols = []

    return category_cols, value_cols


def _generate_chart_options(chart_id: str, data: Dict, chart_type: str) -> str:
    headers = data["headers"]
    rows = data["rows"]
    title = data.get("title", "")

    if not rows or not headers:
        return ""

    category_cols, value_cols = _detect_chart_columns(headers)
    if not value_cols:
        return ""

    categories = []
    series_data = []

    for row in rows:
        # 多列分类拼接：["华东", "产品A"] → "华东-产品A"
        if category_cols:
            parts = [str(row[cc]) for cc in category_cols if cc < len(row) and row[cc] != ""]
            categories.append("-".join(parts) if parts else "")
        else:
            categories.append("")

    for vc in value_cols:
        values = []
        for row in rows:
            if vc < len(row) and row[vc] != "":
                try:
                    values.append(float(row[vc]))
                except ValueError:
                    values.append(0)
            else:
                values.append(0)
        series_data.append({
            "name": headers[vc],
            "data": values,
        })
    
    # 通用 tooltip 样式（适配主题由 JS 端覆盖，此处为静态回退）
    tooltip_style = {
        "backgroundColor": "#ffffff",
        "borderColor": "#e9ecef",
        "borderWidth": 1,
        "textStyle": {"color": "#1a1a2e"},
        "extraCssText": "box-shadow: 0 2px 10px rgba(0,0,0,0.1); border-radius: 8px;"
    }
    base_grid = {"left": "3%", "right": "4%", "bottom": "15%", "top": "15%", "containLabel": True}
    anim = {"animationDuration": 1000, "animationEasing": "cubicOut"}

    def _bar_gradient(color):
        return {
            "type": "linear", "x": 0, "y": 0, "x2": 0, "y2": 1,
            "colorStops": [
                {"offset": 0, "color": color},
                {"offset": 1, "color": color + "55"}
            ]
        }

    def _area_gradient(color):
        return {
            "type": "linear", "x": 0, "y": 0, "x2": 0, "y2": 1,
            "colorStops": [
                {"offset": 0, "color": color + "99"},
                {"offset": 1, "color": color + "0D"}
            ]
        }

    if chart_type == "pie":
        if series_data:
            pie_data = []
            for i, cat in enumerate(categories):
                pie_data.append({
                    "name": cat,
                    "value": series_data[0]["data"][i] if i < len(series_data[0]["data"]) else 0,
                })
            total = sum(d["value"] for d in pie_data)
            return json.dumps({
                "title": {"text": title, "left": "center", "textStyle": {"fontSize": 14}},
                "tooltip": {"trigger": "item", "formatter": "{b}: {c} ({d}%)", **tooltip_style},
                "legend": {"orient": "vertical", "left": "left"},
                "graphic": {
                    "type": "text",
                    "left": "center",
                    "top": "center",
                    "style": {
                        "text": "总计\n" + _format_num(total, title),
                        "textAlign": "center",
                        "textVerticalAlign": "middle",
                        "fontSize": 13,
                        "fill": "#6c757d",
                    }
                },
                "series": [{
                    "name": title,
                    "type": "pie",
                    "radius": ["45%", "72%"],
                    "center": ["50%", "50%"],
                    "avoidLabelOverlap": True,
                    "itemStyle": {"borderRadius": 6, "borderColor": "#fff", "borderWidth": 2},
                    "label": {"show": True},
                    "data": pie_data,
                    "color": HUAWEI_PALETTE,
                }],
                **anim,
            })

    if chart_type == "line":
        series = []
        for i, sd in enumerate(series_data):
            c = HUAWEI_PALETTE[i % len(HUAWEI_PALETTE)]
            series.append({
                "name": sd["name"],
                "type": "line",
                "data": sd["data"],
                "smooth": True,
                "symbol": "circle",
                "symbolSize": 8,
                "lineStyle": {"width": 3, "color": c},
                "itemStyle": {"color": c, "borderColor": "#fff", "borderWidth": 2},
                "areaStyle": {"color": _area_gradient(c)},
            })
        return json.dumps({
            "title": {"text": title, "left": "center", "textStyle": {"fontSize": 14}},
            "tooltip": {"trigger": "axis", **tooltip_style},
            "legend": {"data": [sd["name"] for sd in series_data], "bottom": 0},
            "grid": base_grid,
            "xAxis": {"type": "category", "data": categories, "axisLabel": {"rotate": 30}},
            "yAxis": {"type": "value"},
            "series": series,
            **anim,
        })

    if chart_type == "bar":
        series = []
        for i, sd in enumerate(series_data):
            c = HUAWEI_PALETTE[i % len(HUAWEI_PALETTE)]
            series.append({
                "name": sd["name"],
                "type": "bar",
                "data": sd["data"],
                "barWidth": "50%",
                "itemStyle": {"color": _bar_gradient(c), "borderRadius": [6, 6, 0, 0]},
            })
        return json.dumps({
            "title": {"text": title, "left": "center", "textStyle": {"fontSize": 14}},
            "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}, **tooltip_style},
            "legend": {"data": [sd["name"] for sd in series_data], "bottom": 0},
            "grid": base_grid,
            "xAxis": {"type": "category", "data": categories, "axisLabel": {"rotate": 30}},
            "yAxis": {"type": "value"},
            "series": series,
            **anim,
        })

    return ""


def _generate_geo_chart_options(chart_id: str, data: Dict, chart_type: str) -> str:
    """生成地图类型 chartOptions（ECharts geo + scatter）。

    chart_type: 'map'（散点地图）或 'heatmap'（热力地图）。
    数据需含经纬度列（lat/纬度、lon/经度），其余数值列作为指标。
    """
    headers = data["headers"]
    rows = data["rows"]
    title = data.get("title", "")

    if not rows or not headers:
        return ""

    geo = _detect_geo_columns(headers)
    if not geo:
        return ""

    lat_idx, lon_idx = geo
    # 指标列：除经纬度外的数值列
    metric_indices = []
    for i, h in enumerate(headers):
        if i in (lat_idx, lon_idx):
            continue
        h_lower = h.lower()
        if any(k in h_lower for k in ["金额", "数量", "销售额", "利润", "成本", "值", "占比", "百分比", "pct", "计数", "总计", "合计"]):
            metric_indices.append(i)
    # 没有识别到指标列时，取经纬度外的第一个数值列
    if not metric_indices:
        for i, h in enumerate(headers):
            if i in (lat_idx, lon_idx):
                continue
            try:
                if rows:
                    float(rows[0][i])
                    metric_indices.append(i)
                    break
            except (ValueError, IndexError):
                continue

    # 名称列（站点名等）：经纬度和指标外的第一列
    name_idx = None
    for i, h in enumerate(headers):
        if i in (lat_idx, lon_idx) or i in metric_indices:
            continue
        name_idx = i
        break

    # 收集数据点
    points = []
    for row in rows:
        try:
            lat = float(row[lat_idx])
            lon = float(row[lon_idx])
        except (ValueError, IndexError):
            continue
        if lat == 0 and lon == 0:
            continue
        name = str(row[name_idx]) if name_idx is not None and name_idx < len(row) else ""
        values = [lon, lat]
        for mi in metric_indices:
            try:
                values.append(float(row[mi]))
            except (ValueError, IndexError):
                values.append(0)
        points.append({"name": name, "value": values})

    if not points:
        return ""

    # 计算经纬度范围用于 geo 中心点和缩放
    lats = [p["value"][1] for p in points]
    lons = [p["value"][0] for p in points]
    center_lon = (min(lons) + max(lons)) / 2
    center_lat = (min(lats) + max(lats)) / 2
    span = max(max(lons) - min(lons), max(lats) - min(lats), 0.1)
    import math
    zoom = round(max(1, min(15, 8 - math.log10(span + 1))), 2)

    metric_names = [headers[mi] for mi in metric_indices] or ["指标"]
    metric_idx_in_value = 2  # value 数组中指标的起始位置

    # 散点大小归一化（用第一个指标列）
    if metric_indices:
        m_vals = [p["value"][metric_idx_in_value] for p in points]
        m_min, m_max = min(m_vals), max(m_vals)
        if m_max == m_min:
            sizes = [20] * len(points)
        else:
            sizes = [int(10 + (v - m_min) / (m_max - m_min) * 40) for v in m_vals]
    else:
        sizes = [20] * len(points)

    # tooltip 的 formatter 是 JS 函数，不能放进 json，改用模板字符串拼装整个选项
    tooltip_fn_parts = ['function(p){var s=p.name+"<br/>";var lon=p.value[0],lat=p.value[1];',
                        's+="经度:"+lon.toFixed(4)+", 纬度:"+lat.toFixed(4)+"<br/>";']
    for i, mn in enumerate(metric_names):
        tooltip_fn_parts.append(f's+="{mn}:"+p.value[{i+2}]+"<br/>";')
    tooltip_fn_parts.append('return s;}')
    tooltip_fn = "".join(tooltip_fn_parts)

    series_type = "scatter" if chart_type == "map" else "effectScatter"
    item_style = {"color": HUAWEI_PALETTE[0], "opacity": 0.85,
                  "borderColor": "#333", "borderWidth": 0.5} if chart_type == "map" else \
                 {"color": HUAWEI_PALETTE[3], "shadowBlur": 10,
                  "shadowColor": "rgba(237, 125, 49, 0.5)"}
    extra_series = {}
    if chart_type == "heatmap":
        extra_series = {"showEffectOn": "render", "rippleEffect": {"brushType": "stroke"}}

    # 拼装 option JSON，formatter 用 JS 函数字面量（不能 json.dumps）
    opt = {
        "title": {"text": title, "left": "center", "textStyle": {"fontSize": 14}},
        "tooltip": {"trigger": "item"},  # formatter 稍后手动替换
        "geo": {
            "map": "china",
            "roam": True,
            "zoom": zoom,
            "center": [center_lon, center_lat],
            "itemStyle": {"areaColor": "#F2F0EB", "borderColor": "#999"},
            "emphasis": {"itemStyle": {"areaColor": "#DCE6F0"}, "label": {"show": False}},
        },
        "series": [{
            "name": metric_names[0],
            "type": series_type,
            "coordinateSystem": "geo",
            "data": points,
            "symbolSize": sizes,
            "itemStyle": item_style,
            "label": {"show": True, "formatter": "{b}", "position": "right",
                      "fontSize": 10, "color": "#182B49"},
            **extra_series,
        }],
    }
    opt_json = json.dumps(opt, ensure_ascii=False)
    # 在 tooltip 对象内插入 formatter 函数
    opt_json = opt_json.replace('"tooltip": {"trigger": "item"}',
                                f'"tooltip": {{"trigger": "item", "formatter": {tooltip_fn}}}')
    return opt_json


def _build_chart_data_js(sections: List[Dict]) -> str:
    """序列化所有非地图 section 的完整数据到 JS chartData 对象，供 JS 端动态构建图表（支持列选择开关）。"""
    parts = []
    for sec in sections:
        sid = sec["id"]
        headers = sec["headers"]
        rows = sec["rows"]
        if not rows or not headers:
            continue
        # 跳过 PPT 幻灯片 section 和地图 section
        if sec.get("ppt_images"):
            continue
        geo = _detect_geo_columns(headers)
        if geo:
            continue
        category_cols, value_cols = _detect_chart_columns(headers)
        if not value_cols:
            continue

        categories = []
        for row in rows:
            # 多列分类拼接
            if category_cols:
                cat_parts = [str(row[cc]) for cc in category_cols if cc < len(row) and row[cc] != ""]
                categories.append("-".join(cat_parts) if cat_parts else "")
            else:
                categories.append("")

        series_list = []
        for vc in value_cols:
            values = []
            for row in rows:
                if vc < len(row) and row[vc] != "":
                    try:
                        values.append(float(row[vc]))
                    except ValueError:
                        values.append(0)
                else:
                    values.append(0)
            series_list.append({"name": headers[vc], "data": values, "selected": True})

        col_labels = [headers[vc] for vc in value_cols]
        parts.append(f"chartData['{sid}'] = {{\"title\": {json.dumps(sec.get('title', ''))}, "
                     f"\"categories\": {json.dumps(categories)}, "
                     f"\"series\": {json.dumps(series_list, ensure_ascii=False)}, "
                     f"\"colLabels\": {json.dumps(col_labels, ensure_ascii=False)}}};")
    return "\n".join(parts)


def _build_geo_data_js(sec: Dict) -> str:
    """序列化地图 section 的完整数据，供 JS 端动态构建图表（支持指标列切换+筛选）。

    返回 JS 字符串：geoData['sectionId'] = {...}
    """
    headers = sec["headers"]
    rows = sec["rows"]
    geo = _detect_geo_columns(headers)
    if not geo:
        return ""

    lat_idx, lon_idx = geo
    # 指标列候选：除经纬度外的数值列
    metric_indices = []
    for i, h in enumerate(headers):
        if i in (lat_idx, lon_idx):
            continue
        h_lower = h.lower()
        if any(k in h_lower for k in ["金额", "数量", "销售额", "利润", "成本", "值", "占比", "百分比", "pct", "计数", "总计", "合计"]):
            metric_indices.append(i)
    if not metric_indices:
        for i, h in enumerate(headers):
            if i in (lat_idx, lon_idx):
                continue
            try:
                if rows:
                    float(rows[0][i])
                    metric_indices.append(i)
            except (ValueError, IndexError):
                continue

    # 筛选列候选：非经纬度、非指标、非名称的列（通常是分类列）
    name_idx = None
    for i, h in enumerate(headers):
        if i in (lat_idx, lon_idx) or i in metric_indices:
            continue
        if name_idx is None:
            name_idx = i
    # 筛选列：所有非数值列（排除经纬度）
    filter_indices = []
    for i, h in enumerate(headers):
        if i in (lat_idx, lon_idx) or i in metric_indices:
            continue
        filter_indices.append(i)

    # 收集每个筛选列的唯一值
    filter_values = {}
    for fi in filter_indices:
        vals = set()
        for row in rows:
            if fi < len(row) and row[fi] != "":
                vals.add(str(row[fi]))
        # 唯一值超过 50 个时不作为筛选列（太多无意义）
        if len(vals) <= 50:
            filter_values[fi] = sorted(vals)

    data_obj = {
        "headers": headers,
        "rows": rows,
        "latIdx": lat_idx,
        "lonIdx": lon_idx,
        "nameIdx": name_idx,
        "metricIndices": metric_indices,
        "filterIndices": list(filter_values.keys()),
        "filterValues": {str(k): v for k, v in filter_values.items()},
    }
    return f"  geoData['{sec['id']}'] = {json.dumps(data_obj, ensure_ascii=False)};\n"


def _build_col_selector_html(sec: Dict) -> str:
    """构建列选择复选框面板 HTML。用于普通图表（非地图），允许用户勾选要展示的数值列。"""
    headers = sec["headers"]
    rows = sec["rows"]
    geo = _detect_geo_columns(headers)
    if geo or not rows:
        return ""
    _, value_cols = _detect_chart_columns(headers)
    if len(value_cols) <= 1:
        return ""  # 只有1列时不显示选择器
    sid = sec["id"]
    checkboxes = ""
    for vi, vc in enumerate(value_cols):
        col_name = headers[vc]
        checkboxes += f'<label class="col-cb-item"><input type="checkbox" checked onchange="onColChange(\'{sid}\')" data-series="{vi}">{_escape_html(col_name)}</label>'
    select_all_btn = (f'<button class="col-sel-toggle" type="button" '
                      f'onclick="onColSelectAll(\'{sid}\', true)">全选</button>'
                      f'<button class="col-sel-toggle" type="button" '
                      f'onclick="onColSelectAll(\'{sid}\', false)">全不选</button>')
    return f'<div class="col-selector" id="col_sel_{sid}"><span class="col-sel-label">列选择:</span>{select_all_btn}{checkboxes}</div>'


def _build_geo_controls_html(sec: Dict) -> str:
    """构建地图 section 的控件面板 HTML（指标列下拉 + 筛选列下拉 + 筛选值）。"""
    headers = sec["headers"]
    geo = _detect_geo_columns(headers)
    if not geo:
        return ""
    lat_idx, lon_idx = geo

    # 所有可用列（排除经纬度），指标列和筛选列共用同一套选项
    all_col_indices = [i for i in range(len(headers)) if i not in (lat_idx, lon_idx)]

    metric_options = "".join(
        f'<option value="{mi}">{_escape_html(headers[mi])}</option>' for mi in all_col_indices
    )

    # 筛选列选项（所有非经纬度列都可筛选，唯一值不超过50个）
    filter_indices = []
    for i in all_col_indices:
        vals = set()
        for row in sec["rows"]:
            if i < len(row) and row[i] != "":
                vals.add(str(row[i]))
        if len(vals) <= 50:
            filter_indices.append(i)
    filter_options = '<option value="">-- 不筛选 --</option>' + "".join(
        f'<option value="{fi}">{_escape_html(headers[fi])}</option>' for fi in filter_indices
    )

    sid = sec["id"]
    return f'''<div class="geo-controls" id="geo_ctrl_{sid}" style="display:none">
  <div class="geo-ctrl-item"><label>指标列:</label>
    <select id="geo_metric_{sid}" onchange="onGeoMetricChange('{sid}')">{metric_options}</select>
  </div>
  <div class="geo-ctrl-item"><label>筛选列:</label>
    <select id="geo_filter_col_{sid}" onchange="onGeoFilterColChange('{sid}')">{filter_options}</select>
  </div>
  <div class="geo-ctrl-item" id="geo_filter_vals_wrap_{sid}" style="display:none">
    <label>筛选值:</label><div class="geo-filter-vals" id="geo_filter_vals_{sid}"></div>
  </div>
</div>'''


def generate_html_report(
    excel_path: Optional[str] = None,
    ppt_path: Optional[str] = None,
    ppt_config: Optional[List[Dict]] = None,
    output_dir: Optional[str] = None,
    report_title: str = "数据分析报告",
    report_subtitle: str = "透视分析结果",
) -> Optional[str]:
    if not output_dir:
        output_dir = os.path.dirname(excel_path) if excel_path else "."
    os.makedirs(output_dir, exist_ok=True)

    excel_sections = []
    if excel_path and os.path.exists(excel_path):
        print("[HTML] 读取 Excel 数据...")
        excel_sections = read_excel_data(excel_path)

    sections = []

    if excel_sections:
        print(f"[HTML] 从透视数据生成报告结构（共 {len(excel_sections)} 个区块）...")
        for sec in excel_sections:
            if sec["name"] in ("错误信息",):
                continue

            chart_type = _infer_chart_type(sec["headers"], sec["rows"])
            summary = _compute_summary(sec["headers"], sec["rows"])

            has_geo = _detect_geo_columns(sec["headers"]) is not None
            # 数值列数量决定可用的扩展图表类型（radar/scatter 需>=2列）
            _, _vcols = _detect_chart_columns(sec["headers"])
            n_vcols = len(_vcols)
            if has_geo:
                available = ["map", "heatmap", "table"]
            elif chart_type == "scalar":
                available = ["table"]
            else:
                available = ["bar", "line", "area", "pie", "table"]
                if n_vcols >= 2:
                    available = ["bar", "line", "area", "pie", "radar", "scatter", "table"]

            sections.append({
                "id": f"section_{len(sections)}",
                "title": sec["name"],
                "subtitle": "",
                "sheet": sec.get("sheet", sec["name"]),
                "chart_type": chart_type,
                "available_charts": available,
                "headers": sec["headers"],
                "rows": sec["rows"],
                "summary": summary,
            })

    if ppt_path and os.path.exists(ppt_path):
        print("[HTML] 正在将 PPT 转为图片...")
        try:
            import win32com.client
            import pythoncom
            
            pythoncom.CoInitialize()
            ppt_app = win32com.client.Dispatch("PowerPoint.Application")
            ppt_app.Visible = 0
            pres = ppt_app.Presentations.Open(os.path.abspath(ppt_path), WithWindow=False)
            
            ppt_images = []
            for i, slide in enumerate(pres.Slides, 1):
                img_path = os.path.join(output_dir, f"slide_{i:02d}.png")
                slide.Export(img_path, "PNG", 1280, 720)
                with open(img_path, "rb") as f:
                    img_b64 = base64.b64encode(f.read()).decode("utf-8")
                ppt_images.append(img_b64)
            
            pres.Close()
            ppt_app.Quit()
            pythoncom.CoUninitialize()
            
            if ppt_images:
                ppt_section = {
                    "id": "section_ppt",
                    "title": "PPT 预览",
                    "subtitle": "点击图片可缩放查看",
                    "chart_type": "ppt",
                    "available_charts": [],
                    "headers": [],
                    "rows": [],
                    "summary": {},
                    "ppt_images": ppt_images,
                }
                sections.insert(0, ppt_section)
        except Exception as e:
            print(f"[警告] PPT转图片失败: {e}")

    if not sections:
        print("[错误] 没有可用数据")
        return None

    all_summary = {}
    for sec in sections:
        if sec.get("summary"):
            for key, val in sec["summary"].items():
                if isinstance(val, dict):
                    all_summary[key] = val
                else:
                    all_summary[key] = val

    section_html_parts = []
    toc_html_parts = []
    all_chart_options_js = ""
    all_geo_data_js = ""
    all_chart_data_js = ""

    # 按 Sheet 分组生成二级目录
    sheet_groups = {}
    sheet_order = []
    ppt_sections = []
    for sec in sections:
        if sec.get("ppt_images"):
            ppt_sections.append(sec)
            continue
        sh = sec.get("sheet", sec.get("title", ""))
        if sh not in sheet_groups:
            sheet_groups[sh] = []
            sheet_order.append(sh)
        sheet_groups[sh].append(sec)

    # PPT 幻灯片作为第一个一级目录
    if ppt_sections:
        toc_html_parts.append(f'<li class="toc-group open"><div class="toc-group-hd" onclick="toggleTocGroup(this)"><span class="toc-num">0</span>PPT 幻灯片<span class="toc-arrow">▸</span></div><ul class="toc-sub">')
        for psec in ppt_sections:
            toc_html_parts.append(f'<li><a href="#{psec["id"]}">{_escape_html(psec.get("title", "幻灯片"))}</a></li>')
        toc_html_parts.append('</ul></li>')

    block_idx = 1
    for sh in sheet_order:
        items = sheet_groups[sh]
        # 所有 sheet 都作为一级目录，区块作为二级
        toc_html_parts.append(f'<li class="toc-group"><div class="toc-group-hd" onclick="toggleTocGroup(this)"><span class="toc-num">{block_idx}</span>{_escape_html(sh)}<span class="toc-arrow">▸</span></div><ul class="toc-sub">')
        block_idx += 1
        for sub in items:
            toc_html_parts.append(f'<li><a href="#{sub["id"]}">{_escape_html(sub.get("title", ""))}</a></li>')
        toc_html_parts.append('</ul></li>')

    # section HTML 循环
    sec_idx = 0
    for sec in sections:
        accent = HUAWEI_PALETTE[sec_idx % len(HUAWEI_PALETTE)]

        if sec.get("ppt_images"):
            slide_html = ""
            for idx, img_b64 in enumerate(sec["ppt_images"], 1):
                slide_html += f'''
    <div class="slide-card">
      <div class="slide-num">第 {idx} 页</div>
      <img src="data:image/png;base64,{img_b64}" alt="Slide {idx}" onclick="toggleZoom(this)">
    </div>'''
            section_html_parts.append(f'''
<div id="{sec["id"]}" class="section" style="--accent: {accent}">
  <h2>{_escape_html(sec["title"])}</h2>
  <p class="subtitle">{_escape_html(sec["subtitle"])}</p>
  {slide_html}
</div>''')
            sec_idx += 1
            continue

        chart_options_js = ""
        is_geo_section = "map" in sec["available_charts"] or "heatmap" in sec["available_charts"]
        for ct in sec["available_charts"]:
            if ct in ("map", "heatmap"):
                opts = _generate_geo_chart_options(sec["id"], sec, ct)
            elif ct in ("area", "radar", "scatter"):
                # 这三种类型由 JS 端 buildChartOption 动态构建，无需 Python 预生成 option
                continue
            else:
                opts = _generate_chart_options(sec["id"], sec, ct)
            if opts:
                chart_options_js += f"  chartOptions['{sec['id']}_{ct}'] = {opts};\n"
        all_chart_options_js += chart_options_js

        geo_controls_html = ""
        if is_geo_section:
            geo_js = _build_geo_data_js(sec)
            if geo_js:
                all_geo_data_js += geo_js
                geo_controls_html = _build_geo_controls_html(sec)

        # 数据摘要（用 _format_num 格式化）
        summary_html = ""
        if sec.get("summary"):
            summary_items = []
            for key, val in sec["summary"].items():
                if isinstance(val, dict):
                    for stat, num in val.items():
                        stat_labels = {"sum": "合计", "avg": "平均", "max": "最大", "min": "最小", "count": "数量"}
                        summary_items.append(f"{stat_labels.get(stat, stat)}: {_format_num(num, key)}")
                else:
                    summary_items.append(f"{key}: {_format_num(val, key)}")
            if summary_items:
                summary_html = f'<div class="summary-bar"><span>数据摘要:</span> {" | ".join(summary_items)}</div>'

        # 表格构建（含排序表头 + 工具栏 + 维度列高亮）
        # 计算维度列集合（分类列），用于颜色区分
        _geo = _detect_geo_columns(sec["headers"])
        _geo_set = set(_geo) if _geo else set()
        _cat_cols, _ = _detect_chart_columns(sec["headers"])
        _dim_set = set(_cat_cols) - _geo_set

        _dim_class = ' class="dim-th"'
        header_cells = "".join(
            f'<th onclick="sortTable(\'{sec["id"]}\', {i})" data-sort="" data-col="{i}"{_dim_class if i in _dim_set else ""}>{_escape_html(h)}<span class="sort-arrow"></span></th>'
            for i, h in enumerate(sec["headers"])
        )
        body_rows = ""
        for row in sec["rows"]:
            is_total = any("合计" in str(v) or "总计" in str(v) for v in row)
            row_class = "total-row" if is_total else ""
            cells = ""
            for ci, v in enumerate(row):
                if ci in _dim_set:
                    cells += f'<td class="dim-cell">{_escape_html(v)}</td>'
                elif _is_numeric_cell(v):
                    cells += f'<td class="num-cell">{_escape_html(v)}</td>'
                else:
                    cells += f'<td>{_escape_html(v)}</td>'
            body_rows += f'<tr class="{row_class}">{cells}</tr>'

        row_count = len(sec["rows"])
        row_count_html = f'<span class="table-row-count">共 {row_count} 行</span>'
        table_toolbar_html = (
            f'<div class="table-toolbar">'
            f'<input type="text" class="table-search" placeholder="搜索表格内容..." oninput="filterTable(\'{sec["id"]}\', this.value)">'
            f'<button class="table-tool-btn" onclick="exportTableCSV(\'{sec["id"]}\')">导出CSV</button>'
            f'</div>'
        )
        table_block_html = (
            f'<div class="table-container">'
            f'<div class="table-header"><span>数据详情</span>{row_count_html}</div>'
            f'{table_toolbar_html}'
            f'<div class="table-wrap">'
            f'<table id="table_{sec["id"]}"><thead><tr>{header_cells}</tr></thead><tbody>{body_rows}</tbody></table>'
            f'</div></div>'
        )

        # 标量数据：用指标卡片网格展示，不画图
        if sec["chart_type"] == "scalar":
            _, value_cols = _detect_chart_columns(sec["headers"])
            row = sec["rows"][0] if sec["rows"] else []
            metric_cards = ""
            for vc in value_cols:
                col_name = sec["headers"][vc] if vc < len(sec["headers"]) else ""
                val = row[vc] if vc < len(row) else ""
                formatted = _format_num(val, col_name)
                metric_cards += f'<div class="metric-card"><div class="metric-label">{_escape_html(col_name)}</div><div class="metric-value">{_escape_html(formatted)}</div></div>'
            section_html_parts.append(f'''
<div id="{sec["id"]}" class="section" style="--accent: {accent}">
  <h2>{_escape_html(sec["title"])}</h2>
  <p class="subtitle">{_escape_html(sec["subtitle"])}</p>
  {summary_html}
  <div class="metric-grid">{metric_cards}</div>
  {table_block_html}
</div>''')
            sec_idx += 1
            continue

        # 普通图表 section
        chart_buttons_html = ""
        labels = {"bar": "柱状图", "line": "折线图", "area": "面积图", "pie": "饼图",
                  "radar": "雷达图", "scatter": "散点图", "table": "表格",
                  "map": "地图", "heatmap": "热力图"}
        for ct in sec["available_charts"]:
            active = "active" if ct == sec["chart_type"] else ""
            chart_buttons_html += f'<button class="chart-btn {active}" onclick="switchChart(\'{sec["id"]}\', \'{ct}\')">{labels.get(ct, ct)}</button>'

        col_selector_html = _build_col_selector_html(sec)

        # 图表工具栏：全屏 / 导出PNG / 堆叠切换（仅含柱状图时显示堆叠按钮）
        stack_btn = f'<button class="chart-tool-btn" onclick="toggleStack(\'{sec["id"]}\')">堆叠/并排</button>' if "bar" in sec["available_charts"] else ""
        chart_toolbar_html = (
            f'<div class="chart-toolbar">'
            f'<button class="chart-tool-btn" onclick="toggleChartFullscreen(\'{sec["id"]}\')">全屏</button>'
            f'<button class="chart-tool-btn" onclick="exportChartPNG(\'{sec["id"]}\')">导出PNG</button>'
            f'{stack_btn}'
            f'</div>'
        )

        section_html_parts.append(f'''
<div id="{sec["id"]}" class="section" style="--accent: {accent}">
  <h2>{_escape_html(sec["title"])}</h2>
  <p class="subtitle">{_escape_html(sec["subtitle"])}</p>
  {summary_html}
  <div class="chart-container" id="chart_ctr_{sec["id"]}">
    <div class="chart-buttons">{chart_buttons_html}</div>
    {col_selector_html}
    {geo_controls_html}
    {chart_toolbar_html}
    <div id="chart_{sec["id"]}" class="chart-canvas"></div>
  </div>
  {table_block_html}
</div>''')
        sec_idx += 1

    all_chart_data_js = _build_chart_data_js(sections)

    summary_cards_html = ""
    top_summary = list(all_summary.items())[:6]
    if top_summary:
        for i, (key, val) in enumerate(top_summary):
            accent = HUAWEI_PALETTE[i % len(HUAWEI_PALETTE)]
            icon = _summary_icon(key)
            display_key = key.replace("合计_", "") if key.startswith("合计_") else key
            if isinstance(val, dict):
                val_str = _format_num(val.get("sum", ""), key)
                stat_label = "合计"
            else:
                val_str = _format_num(val, key)
                stat_label = ""
            trend_html = f'<div class="card-trend">{stat_label}</div>' if stat_label else ""
            summary_cards_html += (
                f'<div class="summary-card" style="--accent: {accent}">'
                f'<span class="card-icon">{icon}</span>'
                f'<div class="card-body">'
                f'<div class="card-label">{_escape_html(display_key)}</div>'
                f'<div class="card-value">{_escape_html(val_str)}</div>'
                f'{trend_html}'
                f'</div></div>'
            )

    html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=5.0, user-scalable=yes">
<title>{_escape_html(report_title)}</title>
<style>
  :root {{
    --bg: #f0f2f5;
    --surface: #ffffff;
    --surface-2: #f8f9fa;
    --surface-3: #f5f7fa;
    --text: #1a1a2e;
    --text-secondary: #6c757d;
    --primary: #2E75B6;
    --primary-dark: #1a4a7a;
    --primary-light: #e8f0fe;
    --border: #e9ecef;
    --shadow: 0 4px 20px rgba(0,0,0,0.06);
    --shadow-hover: 0 8px 30px rgba(0,0,0,0.1);
    --radius: 12px;
    --radius-sm: 8px;
    --header-grad: linear-gradient(135deg, #2E75B6, #1a4a7a);
    --accent: #2E75B6;
    --zebra: #fafbfc;
    --hover-bg: #e9f5ff;
    --th-bg: #f8f9fa;
    --th-color: #495057;
    --total-bg: #D9E2F3;
    --dim-bg: #e8f0fe;
    --dim-color: #1a4a7a;
    --tooltip-bg: #ffffff;
    --tooltip-border: #e9ecef;
    --tooltip-text: #1a1a2e;
    --backdrop: rgba(255,255,255,0.85);
  }}
  [data-theme="dark"] {{
    --bg: #1a1a2e;
    --surface: #16213e;
    --surface-2: #1f2a4a;
    --surface-3: #1a2440;
    --text: #e0e0e0;
    --text-secondary: #a0a0a0;
    --primary: #5B9BD5;
    --primary-dark: #4A8BC2;
    --primary-light: #1a2a4a;
    --border: #2a3a5a;
    --shadow: 0 4px 20px rgba(0,0,0,0.3);
    --shadow-hover: 0 8px 30px rgba(0,0,0,0.4);
    --header-grad: linear-gradient(135deg, #16213e, #0f1626);
    --zebra: #1f2a4a;
    --hover-bg: #1a2a4a;
    --th-bg: #1f2a4a;
    --th-color: #c0c8d8;
    --total-bg: #243456;
    --dim-bg: #1a2a4a;
    --dim-color: #8ab8f0;
    --tooltip-bg: #16213e;
    --tooltip-border: #2a3a5a;
    --tooltip-text: #e0e0e0;
    --backdrop: rgba(22,33,62,0.85);
  }}
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: -apple-system, "Microsoft YaHei", "PingFang SC", sans-serif; background: var(--bg); color: var(--text); font-size: 14px; line-height: 1.7; overflow-x: hidden; transition: background 0.3s, color 0.3s; }}
  .header {{ background: var(--header-grad); color: #fff; padding: 22px 16px; text-align: center; position: sticky; top: 0; z-index: 100; box-shadow: 0 4px 20px rgba(0,0,0,0.15); backdrop-filter: blur(10px); -webkit-backdrop-filter: blur(10px); }}
  .header h1 {{ font-size: 22px; margin-bottom: 4px; font-weight: 700; letter-spacing: 0.5px; }}
  .header p {{ font-size: 13px; opacity: 0.85; font-weight: 400; }}
  .nav-bar {{ display: flex; justify-content: center; align-items: center; gap: 12px; padding: 12px 16px; background: var(--backdrop); border-bottom: 1px solid var(--border); flex-wrap: wrap; backdrop-filter: blur(10px); -webkit-backdrop-filter: blur(10px); position: sticky; top: 0; z-index: 99; }}
  .nav-btn {{ padding: 7px 18px; background: var(--surface-3); border: 1px solid var(--border); border-radius: 999px; font-size: 13px; color: var(--text-secondary); cursor: pointer; transition: all 0.25s; font-weight: 500; }}
  .nav-btn:hover {{ background: var(--primary); color: #fff; border-color: var(--primary); transform: translateY(-1px); box-shadow: 0 4px 12px rgba(0,0,0,0.15); }}
  .nav-btn.theme-toggle {{ background: var(--primary-light); color: var(--primary); border-color: var(--primary); }}
  .nav-search {{ padding: 7px 14px; border: 1px solid var(--border); border-radius: 999px; font-size: 13px; color: var(--text); background: var(--surface); min-width: 240px; outline: none; transition: all 0.2s; }}
  .nav-search:focus {{ border-color: var(--primary); box-shadow: 0 0 0 3px var(--primary-light); }}
  .main-layout {{ display: flex; width: 100%; }}
  .sidebar {{ width: 250px; min-width: 250px; padding: 20px 16px; background: var(--surface); position: fixed; left: 0; top: 110px; bottom: 0; overflow-y: auto; z-index: 10; box-shadow: 2px 0 12px rgba(0,0,0,0.06); border-right: 1px solid var(--border); }}
  .sidebar h3 {{ font-size: 14px; color: var(--primary); margin-bottom: 14px; padding-left: 10px; border-left: 3px solid var(--primary); font-weight: 600; }}
  .sidebar ul {{ list-style: none; padding-left: 0; }}
  .sidebar li {{ margin-bottom: 4px; }}
  .sidebar a {{ text-decoration: none; color: var(--text-secondary); font-size: 13px; transition: all 0.2s; display: flex; align-items: center; gap: 6px; padding: 7px 10px; border-radius: var(--radius-sm); line-height: 1.4; border-left: 3px solid transparent; }}
  .sidebar a:hover {{ color: var(--primary); background: var(--primary-light); padding-left: 13px; }}
  .toc-num {{ display: inline-flex; align-items: center; justify-content: center; min-width: 20px; height: 20px; padding: 0 4px; background: var(--primary-light); color: var(--primary); border-radius: 4px; font-size: 11px; font-weight: 600; flex-shrink: 0; }}
  .toc-group {{ margin-bottom: 2px; }}
  .toc-group-hd {{ display: flex; align-items: center; gap: 6px; padding: 7px 10px; border-radius: var(--radius-sm); cursor: pointer; font-size: 13px; color: var(--text); user-select: none; transition: all 0.2s; }}
  .toc-group-hd:hover {{ color: var(--primary); background: var(--primary-light); }}
  .toc-arrow {{ font-size: 10px; color: var(--text-secondary); transition: transform 0.2s; margin-left: auto; }}
  .toc-group.open .toc-arrow {{ transform: rotate(90deg); }}
  .toc-sub {{ display: none; list-style: none; padding-left: 18px; }}
  .toc-group.open .toc-sub {{ display: block; }}
  .toc-hd {{ display: flex; align-items: center; gap: 6px; margin-bottom: 14px; }}
  .toc-hd h3 {{ margin-bottom: 0; flex: 1; }}
  .toc-toggle-btn {{ padding: 4px 10px; background: var(--surface-3); border: 1px solid var(--border); border-radius: 999px; font-size: 11px; color: var(--text-secondary); cursor: pointer; transition: all 0.2s; }}
  .toc-toggle-btn:hover {{ background: var(--primary); color: #fff; border-color: var(--primary); }}
  .sidebar a.active {{ color: var(--primary); background: var(--primary-light); font-weight: 600; border-left-color: var(--primary); }}
  .sidebar a.search-hit {{ color: #C8102E; font-weight: 600; }}
  .summary-cards {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(180px, 1fr)); gap: 10px; margin-top: 16px; }}
  .summary-card {{ background: var(--accent); color: #fff; padding: 14px; border-radius: var(--radius-sm); display: flex; align-items: center; gap: 10px; position: relative; overflow: hidden; box-shadow: 0 2px 10px rgba(0,0,0,0.12); transition: transform 0.2s; }}
  .summary-card:hover {{ transform: translateY(-2px); }}
  .summary-card::after {{ content: ''; position: absolute; top: 0; left: 0; right: 0; bottom: 50%; background: linear-gradient(180deg, rgba(255,255,255,0.18), transparent); pointer-events: none; }}
  [data-theme="dark"] .summary-card {{ background: color-mix(in srgb, var(--accent) 38%, var(--surface)); border: 1px solid color-mix(in srgb, var(--accent) 50%, transparent); }}
  .summary-card .card-icon {{ font-size: 24px; z-index: 1; }}
  .summary-card .card-body {{ flex: 1; min-width: 0; z-index: 1; }}
  .summary-card .card-label {{ font-size: 11px; opacity: 0.85; margin-bottom: 2px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
  .summary-card .card-value {{ font-size: 18px; font-weight: bold; }}
  .summary-card .card-trend {{ font-size: 10px; opacity: 0.75; margin-top: 2px; }}
  .metric-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(170px, 1fr)); gap: 14px; margin-bottom: 24px; }}
  .metric-card {{ background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 18px; text-align: center; box-shadow: var(--shadow); transition: all 0.25s; }}
  .metric-card:hover {{ transform: translateY(-3px); box-shadow: var(--shadow-hover); border-color: var(--primary); }}
  .metric-card .metric-label {{ font-size: 12px; color: var(--text-secondary); margin-bottom: 8px; }}
  .metric-card .metric-value {{ font-size: 24px; font-weight: bold; color: var(--primary); }}
  .content-area {{ max-width: 1280px; width: 100%; padding: 24px; }}
  .content-wrapper {{ margin-left: 250px; flex: 1; display: flex; justify-content: center; min-width: 0; }}
  .section {{ background: var(--surface); border-radius: var(--radius); box-shadow: var(--shadow); padding: 28px; margin-bottom: 24px; border-left: 4px solid var(--accent); transition: box-shadow 0.3s, transform 0.3s; }}
  .section:hover {{ box-shadow: var(--shadow-hover); }}
  .section h2 {{ font-size: 20px; color: var(--text); margin-bottom: 4px; font-weight: 600; }}
  .section .subtitle {{ font-size: 13px; color: var(--text-secondary); margin-bottom: 16px; font-weight: 400; }}
  .summary-bar {{ background: var(--surface-3); padding: 12px 16px; border-radius: var(--radius-sm); margin-bottom: 18px; font-size: 13px; color: var(--text-secondary); }}
  .summary-bar span {{ font-weight: 600; color: var(--primary); }}
  .chart-container {{ margin-bottom: 24px; }}
  .chart-buttons {{ display: flex; gap: 8px; margin-bottom: 14px; flex-wrap: wrap; }}
  .chart-btn {{ padding: 7px 16px; background: var(--surface-3); border: 1px solid var(--border); border-radius: 999px; font-size: 12px; color: var(--text-secondary); cursor: pointer; transition: all 0.2s; font-weight: 500; }}
  .chart-btn:hover {{ background: var(--primary-light); border-color: var(--primary); color: var(--primary); }}
  .chart-btn.active {{ background: var(--primary); color: #fff; border-color: var(--primary); box-shadow: 0 2px 8px rgba(0,0,0,0.15); }}
  .chart-toolbar {{ display: flex; gap: 8px; margin-bottom: 10px; }}
  .chart-tool-btn {{ padding: 5px 12px; background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius-sm); font-size: 11px; color: var(--text-secondary); cursor: pointer; transition: all 0.2s; }}
  .chart-tool-btn:hover {{ background: var(--primary); color: #fff; border-color: var(--primary); }}
  .chart-container.fullscreen {{ position: fixed; top: 0; left: 0; width: 100vw; height: 100vh; background: var(--surface); z-index: 9999; padding: 24px; overflow: auto; }}
  .chart-container.fullscreen .chart-canvas {{ height: calc(100vh - 140px) !important; }}
  .col-selector {{ display: flex; flex-wrap: wrap; gap: 10px; align-items: center; padding: 8px 12px; margin-bottom: 10px; background: var(--surface-2); border: 1px solid var(--border); border-radius: var(--radius-sm); font-size: 12px; }}
  .col-sel-label {{ font-weight: 600; color: var(--text-secondary); margin-right: 2px; }}
  .col-cb-item {{ display: inline-flex; align-items: center; gap: 4px; cursor: pointer; color: var(--text); user-select: none; }}
  .col-cb-item input {{ margin: 0; accent-color: var(--primary); }}
  .col-cb-item:hover {{ color: var(--primary); }}
  .col-sel-toggle {{ padding: 4px 10px; background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius-sm); font-size: 11px; color: var(--text-secondary); cursor: pointer; transition: all 0.2s; }}
  .col-sel-toggle:hover {{ background: var(--primary); color: #fff; border-color: var(--primary); }}
  .geo-controls {{ display: none; flex-wrap: wrap; gap: 12px; align-items: flex-start; padding: 12px 14px; margin-bottom: 12px; background: var(--surface-3); border: 1px solid var(--border); border-radius: var(--radius-sm); font-size: 12px; }}
  .geo-ctrl-item {{ display: flex; align-items: center; gap: 6px; }}
  .geo-ctrl-item label {{ color: var(--text-secondary); font-weight: 600; white-space: nowrap; }}
  .geo-ctrl-item select {{ padding: 5px 10px; border: 1px solid var(--border); border-radius: var(--radius-sm); font-size: 12px; max-width: 160px; background: var(--surface); color: var(--text); }}
  .geo-filter-vals {{ display: flex; flex-wrap: wrap; gap: 6px 12px; max-width: 520px; }}
  .geo-fv-item {{ display: inline-flex; align-items: center; gap: 3px; cursor: pointer; color: var(--text); }}
  .geo-fv-item input {{ margin: 0; accent-color: var(--primary); }}
  .chart-canvas {{ height: 400px; }}
  .table-container {{ border-top: 1px solid var(--border); padding-top: 18px; }}
  .table-header {{ font-size: 15px; font-weight: 600; color: var(--primary); margin-bottom: 12px; display: flex; justify-content: space-between; align-items: center; }}
  .table-row-count {{ font-size: 12px; font-weight: normal; color: var(--text-secondary); }}
  .table-toolbar {{ display: flex; gap: 8px; align-items: center; margin-bottom: 12px; }}
  .table-search {{ padding: 6px 12px; border: 1px solid var(--border); border-radius: var(--radius-sm); font-size: 12px; min-width: 220px; outline: none; background: var(--surface); color: var(--text); transition: all 0.2s; }}
  .table-search:focus {{ border-color: var(--primary); box-shadow: 0 0 0 3px var(--primary-light); }}
  .table-tool-btn {{ padding: 6px 14px; background: var(--surface-3); border: 1px solid var(--border); border-radius: var(--radius-sm); font-size: 12px; color: var(--text-secondary); cursor: pointer; transition: all 0.2s; }}
  .table-tool-btn:hover {{ background: var(--primary); color: #fff; border-color: var(--primary); }}
  .table-wrap {{ overflow: auto; max-height: 380px; -webkit-overflow-scrolling: touch; border: 1px solid var(--border); border-radius: var(--radius-sm); }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ background: var(--th-bg); color: var(--th-color); font-weight: 700; padding: 11px 14px; text-align: left; border-bottom: 2px solid var(--border); position: sticky; top: 0; z-index: 2; cursor: pointer; user-select: none; white-space: nowrap; }}
  th:hover {{ background: var(--primary-light); color: var(--primary); }}
  th.dim-th {{ background: var(--dim-bg); color: var(--dim-color); }}
  .sort-arrow {{ font-size: 10px; color: var(--primary); margin-left: 4px; }}
  tbody tr.hidden {{ display: none; }}
  td {{ padding: 9px 14px; border-bottom: 1px solid var(--border); color: var(--text); text-align: center; }}
  td.num-cell {{ font-variant-numeric: tabular-nums; }}
  td.dim-cell {{ background: var(--dim-bg); color: var(--dim-color); font-weight: 600; text-align: center; white-space: nowrap; }}
  tbody tr:nth-child(even) {{ background: var(--zebra); }}
  tbody tr:hover {{ background: var(--hover-bg); }}
  tbody tr.total-row {{ background: var(--total-bg); font-weight: bold; }}
  .slide-card {{ background: var(--surface); border-radius: var(--radius); box-shadow: var(--shadow); margin-bottom: 14px; overflow: hidden; }}
  .slide-num {{ background: var(--primary); color: #fff; font-size: 12px; padding: 7px 14px; font-weight: 600; }}
  .slide-card img {{ width: 100%; display: block; cursor: pointer; }}
  .slide-card img.zoomed {{ position: fixed; top: 0; left: 0; width: 100vw; height: 100vh; object-fit: contain; background: rgba(0,0,0,0.92); z-index: 1000; }}
  .footer {{ text-align: center; padding: 28px; color: var(--text-secondary); font-size: 12px; }}
  .back-to-top {{ position: fixed; right: 24px; bottom: 24px; width: 46px; height: 46px; border-radius: 50%; background: var(--primary); color: #fff; border: none; font-size: 22px; cursor: pointer; box-shadow: 0 4px 14px rgba(0,0,0,0.25); opacity: 0; visibility: hidden; transition: all 0.3s; z-index: 90; }}
  .back-to-top.show {{ opacity: 1; visibility: visible; }}
  .back-to-top:hover {{ background: var(--primary-dark); transform: translateY(-2px); }}
  .section.search-hit {{ outline: 2px solid #C8102E; outline-offset: 2px; }}
  @media (max-width: 768px) {{
    .main-layout {{ flex-direction: column; }}
    .sidebar {{ position: static; width: 100%; min-width: 100%; height: auto; box-shadow: none; border-bottom: 1px solid var(--border); top: auto; }}
    .content-wrapper {{ margin-left: 0; }}
    .content-area {{ max-width: 100%; padding: 16px; }}
    .summary-cards {{ grid-template-columns: repeat(2, 1fr); }}
    .chart-canvas {{ height: 300px; }}
    .section {{ padding: 20px; }}
  }}
</style>
</head>
<body>
<div class="header">
  <h1>{_escape_html(report_title)}</h1>
  <p>{_escape_html(report_subtitle)}</p>
</div>
<div class="nav-bar">
  <button class="nav-btn" onclick="scrollToTop()">🏠 回到顶部</button>
  <button class="nav-btn" onclick="window.print()">🖨 打印报告</button>
  <button class="nav-btn theme-toggle" id="themeToggleBtn" onclick="toggleTheme()">🌙 暗色模式</button>
  <input type="text" id="globalSearchInput" class="nav-search" placeholder="🔍 全局搜索区块标题..." oninput="globalSearch(this.value)">
</div>
<div class="main-layout">
  <div class="sidebar">
    <div class="toc-hd">
      <h3>报告目录</h3>
      <button class="toc-toggle-btn" onclick="expandAllToc()">展开全部</button>
      <button class="toc-toggle-btn" onclick="collapseAllToc()">收起全部</button>
    </div>
    <ul>{''.join(toc_html_parts)}</ul>
    <h3>数据摘要</h3>
    <div class="summary-cards">{summary_cards_html}</div>
  </div>
  <div class="content-wrapper">
    <div class="content-area">
      {''.join(section_html_parts)}
      <div class="footer">由 Excel 统一分析工具自动生成 | {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</div>
    </div>
  </div>
</div>
<button id="backToTopBtn" class="back-to-top" onclick="scrollToTop()" title="返回顶部">↑</button>
<script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
<script>
  // 地图数据加载状态：'pending' | 'ready' | 'failed'
  var _chinaMapStatus = 'pending';
  function _loadChinaMap(cb) {{
    if (_chinaMapStatus === 'ready') {{ cb(); return; }}
    if (_chinaMapStatus === 'pending') {{
      _chinaMapStatus = 'loading';
      var s = document.createElement('script');
      s.src = 'https://cdn.jsdelivr.net/npm/echarts@4.9.0/map/js/china.js';
      s.onload = function() {{
        // echarts 4.x 的 china.js 会注册到 echarts.registerMap
        _chinaMapStatus = 'ready';
        cb();
      }};
      s.onerror = function() {{
        // 回退：尝试用 fetch 加载 GeoJSON
        fetch('https://geo.datav.aliyun.com/areas_v3/bound/100000_full.json')
          .then(function(r) {{ return r.json(); }})
          .then(function(data) {{ echarts.registerMap('china', data); _chinaMapStatus='ready'; cb(); }})
          .catch(function() {{ _chinaMapStatus='failed'; cb(); }});
      }};
      document.head.appendChild(s);
    }} else {{
      // loading 中：轮询
      var t = setInterval(function() {{
        if (_chinaMapStatus === 'ready' || _chinaMapStatus === 'failed') {{
          clearInterval(t); cb();
        }}
      }}, 100);
    }}
  }}
</script>
<script>
  var chartInstances = {{}};
  var chartOptions = {{}};
  var geoData = {{}};  // 地图 section 的完整数据，供动态构建
  var chartData = {{}};  // 普通 section 的图表数据，供动态构建+列选择
  {all_chart_options_js}
  {all_geo_data_js}
  {all_chart_data_js}
  
  function initCharts() {{
    Object.keys(chartOptions).forEach(function(key) {{
      // key 格式: section_0_bar / section_1_pie，sectionId 含下划线，从末尾分割
      var lastUnderscore = key.lastIndexOf('_');
      var sectionId = key.substring(0, lastUnderscore);
      var chartType = key.substring(lastUnderscore + 1);
      var container = document.getElementById('chart_' + sectionId);
      if (container && chartType === getDefaultChart(sectionId)) {{
        renderChart(sectionId, chartType);
      }}
    }});
  }}
  
  function getDefaultChart(sectionId) {{
    var btns = document.querySelectorAll('#' + sectionId + ' .chart-btn');
    for (var i = 0; i < btns.length; i++) {{
      if (btns[i].classList.contains('active')) {{
        return btns[i].getAttribute('onclick').match(/'([^']+)'/g)[1].replace(/'/g, '');
      }}
    }}
    return 'bar';
  }}
  
  function renderChart(sectionId, chartType) {{
    var container = document.getElementById('chart_' + sectionId);
    if (!container) return;

    // 显示/隐藏地图控件面板
    var geoCtrl = document.getElementById('geo_ctrl_' + sectionId);
    if (geoCtrl) {{
      geoCtrl.style.display = (chartType === 'map' || chartType === 'heatmap') ? 'flex' : 'none';
    }}

    // 显示/隐藏列选择面板
    var colSel = document.getElementById('col_sel_' + sectionId);
    if (colSel) {{
      colSel.style.display = (chartType !== 'table' && chartType !== 'map' && chartType !== 'heatmap') ? 'flex' : 'none';
    }}

    if (chartType === 'table') {{
      container.style.display = 'none';
      return;
    }}

    container.style.display = 'block';

    if (chartInstances[sectionId]) {{
      chartInstances[sectionId].dispose();
    }}

    var doRender = function() {{
      var chart = echarts.init(container);
      // 地图类型且有 geoData → 动态构建（支持指标列切换+筛选）
      var options;
      if ((chartType === 'map' || chartType === 'heatmap') && geoData[sectionId]) {{
        options = buildGeoOptionFromData(sectionId, chartType);
      }} else if (chartData[sectionId]) {{
        options = buildChartOption(sectionId, chartType);
      }} else {{
        options = chartOptions[sectionId + '_' + chartType];
      }}
      if (options) {{
        chart.setOption(options);
        chartInstances[sectionId] = chart;
      }}
    }};

    // 地图类型需先加载 china 地图数据
    if (chartType === 'map' || chartType === 'heatmap') {{
      _loadChinaMap(doRender);
    }} else {{
      doRender();
    }}
  }}

  // 根据当前控件状态动态构建地图 ECharts option
  function buildGeoOptionFromData(sectionId, chartType) {{
    var d = geoData[sectionId];
    if (!d) return null;

    // 读取控件状态
    var metricSel = document.getElementById('geo_metric_' + sectionId);
    var metricIdx = metricSel ? parseInt(metricSel.value) : (d.metricIndices[0] || -1);

    var filterColSel = document.getElementById('geo_filter_col_' + sectionId);
    var filterColIdx = filterColSel && filterColSel.value ? parseInt(filterColSel.value) : -1;

    // 收集选中的筛选值
    var selectedFilterVals = {{}};
    if (filterColIdx >= 0) {{
      var checkboxes = document.querySelectorAll('input[name="geo_fv_' + sectionId + '"]:checked');
      for (var i = 0; i < checkboxes.length; i++) {{
        selectedFilterVals[checkboxes[i].value] = true;
      }}
    }}

    // 过滤+收集数据点
    var points = [];
    var metricVals = [];
    for (var r = 0; r < d.rows.length; r++) {{
      var row = d.rows[r];
      var lat = parseFloat(row[d.latIdx]);
      var lon = parseFloat(row[d.lonIdx]);
      if (isNaN(lat) || isNaN(lon)) continue;
      if (lat === 0 && lon === 0) continue;
      // 筛选
      if (filterColIdx >= 0) {{
        var fv = String(row[filterColIdx] || '');
        if (Object.keys(selectedFilterVals).length > 0 && !selectedFilterVals[fv]) continue;
      }}
      var name = d.nameIdx !== null && d.nameIdx < row.length ? String(row[d.nameIdx]) : '';
      var values = [lon, lat];
      for (var mi = 0; mi < d.metricIndices.length; mi++) {{
        var v = parseFloat(row[d.metricIndices[mi]]) || 0;
        values.push(v);
      }}
      points.push({{name: name, value: values}});
      var mv = metricIdx >= 0 ? (parseFloat(row[metricIdx]) || 0) : 0;
      metricVals.push(mv);
    }}

    if (points.length === 0) {{
      return {{title: {{text: '无符合条件的数据', left: 'center'}}}};
    }}

    // 经纬度范围
    var lats = points.map(function(p) {{ return p.value[1]; }});
    var lons = points.map(function(p) {{ return p.value[0]; }});
    var centerLon = (Math.min.apply(null, lons) + Math.max.apply(null, lons)) / 2;
    var centerLat = (Math.min.apply(null, lats) + Math.max.apply(null, lats)) / 2;
    var span = Math.max(Math.max.apply(null, lons) - Math.min.apply(null, lons),
                        Math.max.apply(null, lats) - Math.min.apply(null, lats), 0.1);
    var zoom = Math.max(1, Math.min(15, 8 - Math.log10(span + 1)));

    // 指标名
    var metricName = metricIdx >= 0 ? d.headers[metricIdx] : '指标';
    // 指标在 value 数组中的位置
    var metricPos = d.metricIndices.indexOf(metricIdx);
    if (metricPos < 0) metricPos = 0;
    var metricValuePos = metricPos + 2;  // +2 因为前两位是 lon, lat

    // 散点大小归一化
    var mMin = Math.min.apply(null, metricVals);
    var mMax = Math.max.apply(null, metricVals);
    var sizes = [];
    for (var i = 0; i < metricVals.length; i++) {{
      if (mMax === mMin) sizes.push(20);
      else sizes.push(Math.round(10 + (metricVals[i] - mMin) / (mMax - mMin) * 40));
    }}

    // tooltip formatter 稍后以真实函数注入（避免 eval）
    var metricHeaders = d.metricIndices.map(function(mi) {{ return d.headers[mi]; }});

    var seriesType = chartType === 'map' ? 'scatter' : 'effectScatter';
    var _labelColor = isDarkTheme() ? '#e0e0e0' : '#182B49';
    var _areaColor = isDarkTheme() ? '#1f2a4a' : '#F2F0EB';
    var _areaBorder = isDarkTheme() ? '#3a4a6a' : '#999';
    var _emphArea = isDarkTheme() ? '#2a3a5a' : '#DCE6F0';
    var itemStyle = chartType === 'map' ?
      {{color: '#C8102E', opacity: 0.85, borderColor: _labelColor, borderWidth: 0.5}} :
      {{color: '#ED7D31', shadowBlur: 10, shadowColor: 'rgba(237,125,49,0.5)'}};
    var extra = chartType === 'heatmap' ?
      {{showEffectOn: 'render', rippleEffect: {{brushType: 'stroke'}}}} : {{}};

    var opt = {{
      title: {{text: metricName + ' 分布', left: 'center', textStyle: {{fontSize: 14, color: getChartTextColor()}}}},
      tooltip: {{trigger: 'item'}},
      geo: {{
        map: 'china', roam: true, zoom: zoom,
        center: [centerLon, centerLat],
        itemStyle: {{areaColor: _areaColor, borderColor: _areaBorder}},
        emphasis: {{itemStyle: {{areaColor: _emphArea}}, label: {{show: false}}}}
      }},
      series: [{{
        name: metricName, type: seriesType, coordinateSystem: 'geo',
        data: points, symbolSize: sizes, itemStyle: itemStyle,
        label: {{show: true, formatter: '{{b}}', position: 'right', fontSize: 10, color: _labelColor}},
      }}].concat([extra])
    }};
    // 合并 extra 到 series[0]
    for (var k in extra) {{ opt.series[0][k] = extra[k]; }}
    // 移除多余的空对象
    opt.series = [opt.series[0]];
    // 直接注入 tooltip formatter 为真实函数（不使用 eval）
    opt.tooltip.formatter = function(p) {{
      var s = p.name + '<br/>';
      var lon = p.value[0], lat = p.value[1];
      s += '经度:' + lon.toFixed(4) + ', 纬度:' + lat.toFixed(4) + '<br/>';
      for (var i = 0; i < metricHeaders.length; i++) {{
        s += metricHeaders[i] + ':' + p.value[i + 2] + '<br/>';
      }}
      return s;
    }};
    return opt;
  }}

  // 指标列切换
  function onGeoMetricChange(sectionId) {{
    var activeBtn = document.querySelector('#' + sectionId + ' .chart-btn.active');
    var chartType = activeBtn ? (activeBtn.getAttribute('onclick').match(/'([^']+)'$/) || [])[1] : 'map';
    renderChart(sectionId, chartType);
  }}

  // 筛选列切换 → 填充筛选值复选框
  function onGeoFilterColChange(sectionId) {{
    var sel = document.getElementById('geo_filter_col_' + sectionId);
    var valsWrap = document.getElementById('geo_filter_vals_wrap_' + sectionId);
    var valsBox = document.getElementById('geo_filter_vals_' + sectionId);
    if (!sel || !valsBox) return;
    if (!sel.value) {{
      valsWrap.style.display = 'none';
      onGeoMetricChange(sectionId);
      return;
    }}
    var d = geoData[sectionId];
    var vals = d.filterValues[sel.value] || [];
    var html = '';
    // 全选按钮
    html += '<label class="geo-fv-item"><input type="checkbox" name="geo_fv_' + sectionId + '" value="__all__" checked onchange="onGeoFilterAllChange(\\'' + sectionId + '\\')"> 全选</label>';
    for (var i = 0; i < vals.length; i++) {{
      html += '<label class="geo-fv-item"><input type="checkbox" name="geo_fv_' + sectionId + '" value="' + vals[i].replace(/"/g, '&quot;') + '" checked onchange="onGeoFilterValChange(\\'' + sectionId + '\\')"> ' + vals[i] + '</label>';
    }}
    valsBox.innerHTML = html;
    valsWrap.style.display = 'block';
    onGeoMetricChange(sectionId);
  }}

  // 单个筛选值变化
  function onGeoFilterValChange(sectionId) {{
    onGeoMetricChange(sectionId);
  }}

  // 数值格式化（JS 端，与 Python _format_num 对应）：千分位/万/亿/百分比
  function formatNumber(val, isPct) {{
    if (val === null || val === undefined || val === '') return '';
    var num = Number(val);
    if (isNaN(num)) return String(val);
    if (isPct && num >= 0 && num <= 1) return (num * 100).toFixed(2) + '%';
    var abs = Math.abs(num);
    if (abs >= 100000000) return (num / 100000000).toFixed(2) + '亿';
    if (abs >= 10000) return (num / 10000).toFixed(2) + '万';
    var s = num.toFixed(2).replace(/\B(?=(\d{{3}})+(?!\d))/g, ',');
    if (s.indexOf('.00') === s.length - 3) s = s.slice(0, -3);
    return s;
  }}
  function isPctName(name) {{ return /占比|百分比|pct/i.test(name || ''); }}

  // 轴 tooltip formatter（统一格式化数值，百分比系列显示百分比）
  function axisTipFormatter(params) {{
    var s = (params[0].axisValueLabel || params[0].name) + '<br/>';
    for (var i = 0; i < params.length; i++) {{
      var p = params[i];
      s += p.marker + p.seriesName + ': ' + formatNumber(p.value, isPctName(p.seriesName)) + '<br/>';
    }}
    return s;
  }}
  var yAxisLabelFormatter = function(v) {{ return formatNumber(v, false); }};

  // 主题辅助：根据当前主题返回图表配色
  function isDarkTheme() {{ return document.documentElement.getAttribute('data-theme') === 'dark'; }}
  function getChartTextColor() {{ return isDarkTheme() ? '#e0e0e0' : '#1a1a2e'; }}
  function getChartAxisColor() {{ return isDarkTheme() ? '#a0a0a0' : '#6c757d'; }}
  function getChartSplitColor() {{ return isDarkTheme() ? '#2a3a5a' : '#e9ecef'; }}
  // ECharts 线性渐变（柱状图填充：顶部不透明→底部半透明）
  function barGradient(color) {{
    return {{type: 'linear', x: 0, y: 0, x2: 0, y2: 1, colorStops: [
      {{offset: 0, color: color}},
      {{offset: 1, color: color + '55'}}
    ]}};
  }}
  // ECharts 面积渐变（折线/面积图填充）
  function areaGradient(color) {{
    return {{type: 'linear', x: 0, y: 0, x2: 0, y2: 1, colorStops: [
      {{offset: 0, color: color + '99'}},
      {{offset: 1, color: color + '0D'}}
    ]}};
  }}

  // 根据当前列选择状态动态构建普通图表 ECharts option
  function buildChartOption(sectionId, chartType) {{
    var d = chartData[sectionId];
    if (!d) return chartOptions[sectionId + '_' + chartType];

    // 读取列选择状态
    var selected = {{}};
    var checkboxes = document.querySelectorAll('#col_sel_' + sectionId + ' input[type="checkbox"]');
    for (var i = 0; i < checkboxes.length; i++) {{
      selected[parseInt(checkboxes[i].getAttribute('data-series'))] = checkboxes[i].checked;
    }}

    // 筛选已选系列
    var activeSeries = [];
    for (var i = 0; i < d.series.length; i++) {{
      if (selected[i] !== false) {{
        activeSeries.push(d.series[i]);
      }}
    }}
    if (activeSeries.length === 0) {{
      activeSeries = [d.series[0]];
    }}

    var palette = ['#C8102E', '#182B49', '#5B9BD5', '#ED7D31', '#70AD47', '#A5A5A5', '#FFC000', '#4472C4'];
    var legendData = activeSeries.map(function(s){{return s.name;}});

    // 主题相关公共配置
    var tipBg = isDarkTheme() ? '#16213e' : '#ffffff';
    var tipBorder = isDarkTheme() ? '#2a3a5a' : '#e9ecef';
    var tipText = isDarkTheme() ? '#e0e0e0' : '#1a1a2e';
    var tipCss = 'box-shadow: 0 2px 10px rgba(0,0,0,0.15); border-radius: 8px;';
    var axisColor = getChartAxisColor();
    var splitColor = getChartSplitColor();
    var titleColor = getChartTextColor();
    var symbolBorder = isDarkTheme() ? '#16213e' : '#ffffff';
    var pieBorder = isDarkTheme() ? '#16213e' : '#ffffff';
    var baseGrid = {{left: '3%', right: '4%', bottom: '15%', top: '15%', containLabel: true}};
    function catAxis() {{ return {{type: 'category', data: d.categories, axisLabel: {{rotate: 30, color: axisColor}}, axisLine: {{lineStyle: {{color: axisColor}}}}, splitLine: {{lineStyle: {{color: splitColor}}}}}}; }}
    function valAxis() {{ return {{type: 'value', axisLabel: {{formatter: yAxisLabelFormatter, color: axisColor}}, axisLine: {{lineStyle: {{color: axisColor}}}}, splitLine: {{lineStyle: {{color: splitColor}}}}}}; }}

    if (chartType === 'pie') {{
      var s = activeSeries[0];
      var pieData = [];
      for (var i = 0; i < d.categories.length; i++) {{
        pieData.push({{name: d.categories[i], value: (s.data[i] || 0)}});
      }}
      var total = 0;
      for (var i = 0; i < pieData.length; i++) total += pieData[i].value;
      return {{
        title: {{text: d.title, left: 'center', textStyle: {{fontSize: 14, color: titleColor}}}},
        tooltip: {{trigger: 'item', formatter: function(p){{ return p.name + ': ' + formatNumber(p.value, isPctName(p.seriesName)) + ' (' + p.percent.toFixed(1) + '%)'; }}, backgroundColor: tipBg, borderColor: tipBorder, borderWidth: 1, textStyle: {{color: tipText}}, extraCssText: tipCss}},
        legend: {{orient: 'vertical', left: 'left', textStyle: {{color: axisColor}}}},
        graphic: {{type: 'text', left: 'center', top: 'center', style: {{text: '总计\\n' + formatNumber(total, isPctName(d.title)), textAlign: 'center', textVerticalAlign: 'middle', fontSize: 13, fill: titleColor}}}},
        series: [{{
          type: 'pie', radius: ['45%', '72%'], center: ['50%', '50%'],
          avoidLabelOverlap: true, label: {{show: true, color: axisColor}},
          itemStyle: {{borderRadius: 6, borderColor: pieBorder, borderWidth: 2}},
          data: pieData,
        }}],
        animationDuration: 1000, animationEasing: 'cubicOut',
      }};
    }}

    if (chartType === 'line') {{
      var series = [];
      for (var i = 0; i < activeSeries.length; i++) {{
        var c = palette[i % palette.length];
        series.push({{
          name: activeSeries[i].name, type: 'line', data: activeSeries[i].data,
          smooth: true, symbol: 'circle', symbolSize: 8,
          lineStyle: {{width: 3, color: c}}, itemStyle: {{color: c, borderColor: symbolBorder, borderWidth: 2}},
        }});
      }}
      return {{
        title: {{text: d.title, left: 'center', textStyle: {{fontSize: 14, color: titleColor}}}},
        tooltip: {{trigger: 'axis', formatter: axisTipFormatter, backgroundColor: tipBg, borderColor: tipBorder, borderWidth: 1, textStyle: {{color: tipText}}, extraCssText: tipCss}},
        legend: {{data: legendData, bottom: 0, textStyle: {{color: axisColor}}}},
        grid: baseGrid,
        xAxis: catAxis(),
        yAxis: valAxis(),
        series: series,
        animationDuration: 1000, animationEasing: 'cubicOut',
      }};
    }}

    // 面积图：折线 + 大面积渐变填充
    if (chartType === 'area') {{
      var series = [];
      for (var i = 0; i < activeSeries.length; i++) {{
        var c = palette[i % palette.length];
        series.push({{
          name: activeSeries[i].name, type: 'line', data: activeSeries[i].data,
          smooth: true, symbol: 'circle', symbolSize: 8,
          lineStyle: {{width: 3, color: c}}, itemStyle: {{color: c, borderColor: symbolBorder, borderWidth: 2}},
          areaStyle: {{color: areaGradient(c), opacity: 0.5}},
        }});
      }}
      return {{
        title: {{text: d.title, left: 'center', textStyle: {{fontSize: 14, color: titleColor}}}},
        tooltip: {{trigger: 'axis', formatter: axisTipFormatter, backgroundColor: tipBg, borderColor: tipBorder, borderWidth: 1, textStyle: {{color: tipText}}, extraCssText: tipCss}},
        legend: {{data: legendData, bottom: 0, textStyle: {{color: axisColor}}}},
        grid: baseGrid,
        xAxis: catAxis(),
        yAxis: valAxis(),
        series: series,
        animationDuration: 1000, animationEasing: 'cubicOut',
      }};
    }}

    // 雷达图：每个分类作为一个维度
    if (chartType === 'radar') {{
      var indicators = [];
      for (var c = 0; c < d.categories.length; c++) {{
        var mx = 0;
        for (var i = 0; i < activeSeries.length; i++) {{
          var v = activeSeries[i].data[c] || 0;
          if (v > mx) mx = v;
        }}
        indicators.push({{name: d.categories[c], max: Math.ceil(mx * 1.1) || 1}});
      }}
      var radarData = [];
      for (var i = 0; i < activeSeries.length; i++) {{
        radarData.push({{name: activeSeries[i].name, value: activeSeries[i].data, itemStyle: {{color: palette[i % palette.length]}}, areaStyle: {{opacity: 0.2}}}});
      }}
      return {{
        title: {{text: d.title, left: 'center', textStyle: {{fontSize: 14, color: titleColor}}}},
        tooltip: {{trigger: 'item', backgroundColor: tipBg, borderColor: tipBorder, borderWidth: 1, textStyle: {{color: tipText}}, extraCssText: tipCss}},
        legend: {{data: legendData, bottom: 0, textStyle: {{color: axisColor}}}},
        radar: {{indicator: indicators, radius: '65%', axisName: {{color: axisColor}}, splitLine: {{lineStyle: {{color: splitColor}}}}, splitArea: {{areaStyle: {{color: [tipBg, 'transparent']}}}}}},
        series: [{{type: 'radar', data: radarData}}],
        animationDuration: 1000, animationEasing: 'cubicOut',
      }};
    }}

    // 散点图：前两个数值列作为 x/y
    if (chartType === 'scatter') {{
      var xS = activeSeries[0];
      var yS = activeSeries[1] || activeSeries[0];
      var scatterData = [];
      for (var i = 0; i < xS.data.length; i++) {{
        scatterData.push([xS.data[i] || 0, yS.data[i] || 0, d.categories[i] || ('项' + (i+1))]);
      }}
      var xName = xS.name, yName = yS.name;
      return {{
        title: {{text: d.title, left: 'center', textStyle: {{fontSize: 14, color: titleColor}}}},
        tooltip: {{trigger: 'item', formatter: function(p){{ return p.value[2] + '<br/>' + xName + ': ' + formatNumber(p.value[0], isPctName(xName)) + '<br/>' + yName + ': ' + formatNumber(p.value[1], isPctName(yName)); }}, backgroundColor: tipBg, borderColor: tipBorder, borderWidth: 1, textStyle: {{color: tipText}}, extraCssText: tipCss}},
        legend: {{data: [xName + ' vs ' + yName], bottom: 0, textStyle: {{color: axisColor}}}},
        grid: baseGrid,
        xAxis: {{type: 'value', name: xName, nameLocation: 'middle', nameGap: 30, nameTextStyle: {{color: axisColor}}, axisLabel: {{color: axisColor}}, axisLine: {{lineStyle: {{color: axisColor}}}}, splitLine: {{lineStyle: {{color: splitColor}}}}}},
        yAxis: {{type: 'value', name: yName, nameLocation: 'middle', nameGap: 40, nameTextStyle: {{color: axisColor}}, axisLabel: {{formatter: yAxisLabelFormatter, color: axisColor}}, axisLine: {{lineStyle: {{color: axisColor}}}}, splitLine: {{lineStyle: {{color: splitColor}}}}}},
        series: [{{type: 'scatter', data: scatterData, symbolSize: 10, itemStyle: {{color: palette[0]}}}}],
        animationDuration: 1000, animationEasing: 'cubicOut',
      }};
    }}

    // bar (default) —— 支持 stack 切换，柱状图用线性渐变填充
    var stackKey = (window._stackState && window._stackState[sectionId]) ? 'stack_' + sectionId : null;
    var series = [];
    for (var i = 0; i < activeSeries.length; i++) {{
      var c = palette[i % palette.length];
      var item = {{
        name: activeSeries[i].name, type: 'bar', data: activeSeries[i].data,
        barWidth: '50%', itemStyle: {{color: barGradient(c), borderRadius: [6, 6, 0, 0]}},
      }};
      if (stackKey) item.stack = stackKey;
      series.push(item);
    }}
    return {{
      title: {{text: d.title, left: 'center', textStyle: {{fontSize: 14, color: titleColor}}}},
      tooltip: {{trigger: 'axis', axisPointer: {{type: 'shadow'}}, formatter: axisTipFormatter, backgroundColor: tipBg, borderColor: tipBorder, borderWidth: 1, textStyle: {{color: tipText}}, extraCssText: tipCss}},
      legend: {{data: legendData, bottom: 0, textStyle: {{color: axisColor}}}},
      grid: baseGrid,
      xAxis: catAxis(),
      yAxis: valAxis(),
      series: series,
      animationDuration: 1000, animationEasing: 'cubicOut',
    }};
  }}

  // 列选择复选框变更时重新渲染图表
  function onColChange(sectionId) {{
    var btns = document.querySelectorAll('#' + sectionId + ' .chart-btn');
    var currentType = 'bar';
    for (var i = 0; i < btns.length; i++) {{
      if (btns[i].classList.contains('active')) {{
        var m = btns[i].getAttribute('onclick').match(/'([^']+)'/g);
        if (m && m.length >= 2) {{
          currentType = m[1].replace(/'/g, '');
        }}
      }}
    }}
    renderChart(sectionId, currentType);
  }}

  // 全选/取消全选
  function onGeoFilterAllChange(sectionId) {{
    var allBox = document.querySelector('input[name="geo_fv_' + sectionId + '"][value="__all__"]');
    var boxes = document.querySelectorAll('input[name="geo_fv_' + sectionId + '"]:not([value="__all__"])');
    for (var i = 0; i < boxes.length; i++) {{
      boxes[i].checked = allBox.checked;
    }}
    onGeoMetricChange(sectionId);
  }}

  function switchChart(sectionId, chartType) {{
    var btns = document.querySelectorAll('#' + sectionId + ' .chart-btn');
    for (var i = 0; i < btns.length; i++) {{
      btns[i].classList.remove('active');
    }}
    event.target.classList.add('active');
    renderChart(sectionId, chartType);
  }}
  
  function toggleZoom(img) {{
    img.classList.toggle('zoomed');
    document.body.style.overflow = img.classList.contains('zoomed') ? 'hidden' : '';
  }}
  
  function scrollToTop() {{
    window.scrollTo({{ top: 0, behavior: 'smooth' }});
  }}
  
  // 列选择全选/全不选（任务6）
  function onColSelectAll(sectionId, checked) {{
    var checkboxes = document.querySelectorAll('#col_sel_' + sectionId + ' input[type="checkbox"][data-series]');
    for (var i = 0; i < checkboxes.length; i++) {{
      checkboxes[i].checked = checked;
    }}
    onColChange(sectionId);
  }}

  // 表格排序：点击表头切换升/降/原序（任务7）
  var _tableOrigOrder = {{}};
  function sortTable(sectionId, colIdx) {{
    var table = document.getElementById('table_' + sectionId);
    if (!table) return;
    var tbody = table.tBodies[0];
    if (!tbody) return;
    var th = table.querySelector('th[data-col="' + colIdx + '"]');
    if (!th) return;
    if (!_tableOrigOrder[sectionId]) {{
      _tableOrigOrder[sectionId] = Array.prototype.slice.call(tbody.rows);
    }}
    var order = th.getAttribute('data-sort') || '';
    var newOrder = order === 'asc' ? 'desc' : (order === 'desc' ? '' : 'asc');
    var allTh = table.querySelectorAll('th');
    for (var i = 0; i < allTh.length; i++) {{
      allTh[i].setAttribute('data-sort', '');
      var ar = allTh[i].querySelector('.sort-arrow');
      if (ar) ar.textContent = '';
    }}
    th.setAttribute('data-sort', newOrder);
    var arrowEl = th.querySelector('.sort-arrow');
    if (arrowEl) arrowEl.textContent = newOrder === 'asc' ? '▲' : (newOrder === 'desc' ? '▼' : '');
    if (newOrder === '') {{
      var orig = _tableOrigOrder[sectionId];
      for (var i = 0; i < orig.length; i++) tbody.appendChild(orig[i]);
      return;
    }}
    var rows = Array.prototype.slice.call(tbody.rows);
    rows.sort(function(a, b) {{
      var av = a.cells[colIdx] ? a.cells[colIdx].textContent.trim() : '';
      var bv = b.cells[colIdx] ? b.cells[colIdx].textContent.trim() : '';
      var an = parseFloat(av.replace(/,/g, '')), bn = parseFloat(bv.replace(/,/g, ''));
      if (!isNaN(an) && !isNaN(bn)) return newOrder === 'asc' ? an - bn : bn - an;
      return newOrder === 'asc' ? av.localeCompare(bv) : bv.localeCompare(av);
    }});
    for (var i = 0; i < rows.length; i++) tbody.appendChild(rows[i]);
  }}

  // 表格搜索：实时过滤行（任务7）
  function filterTable(sectionId, keyword) {{
    var table = document.getElementById('table_' + sectionId);
    if (!table || !table.tBodies[0]) return;
    var kw = (keyword || '').toLowerCase().trim();
    var rows = table.tBodies[0].rows;
    for (var i = 0; i < rows.length; i++) {{
      if (!kw) {{ rows[i].classList.remove('hidden'); }}
      else {{ rows[i].classList.toggle('hidden', rows[i].textContent.toLowerCase().indexOf(kw) === -1); }}
    }}
  }}

  // 导出表格为 CSV（任务7）
  function exportTableCSV(sectionId) {{
    var table = document.getElementById('table_' + sectionId);
    if (!table) return;
    function csvCell(v) {{
      v = (v === null || v === undefined) ? '' : String(v);
      if (v.indexOf(',') >= 0 || v.indexOf('"') >= 0 || v.indexOf('\\n') >= 0) v = '"' + v.replace(/"/g, '""') + '"';
      return v;
    }}
    var rows = [];
    var headCells = table.querySelectorAll('thead th');
    var headRow = [];
    for (var i = 0; i < headCells.length; i++) headRow.push(csvCell(headCells[i].textContent.replace(/[▲▼]/g, '').trim()));
    rows.push(headRow.join(','));
    var bodyRows = table.querySelectorAll('tbody tr');
    for (var i = 0; i < bodyRows.length; i++) {{
      if (bodyRows[i].classList.contains('hidden')) continue;
      var cells = bodyRows[i].cells;
      var rowData = [];
      for (var j = 0; j < cells.length; j++) rowData.push(csvCell(cells[j].textContent));
      rows.push(rowData.join(','));
    }}
    var csv = '\\ufeff' + rows.join('\\n');
    var blob = new Blob([csv], {{type: 'text/csv;charset=utf-8;'}});
    var url = URL.createObjectURL(blob);
    var a = document.createElement('a');
    a.href = url; a.download = sectionId + '.csv';
    document.body.appendChild(a); a.click(); document.body.removeChild(a);
    URL.revokeObjectURL(url);
  }}

  // 图表全屏切换（任务8）
  function toggleChartFullscreen(sectionId) {{
    var ctr = document.getElementById('chart_ctr_' + sectionId);
    if (!ctr) return;
    ctr.classList.toggle('fullscreen');
    if (chartInstances[sectionId]) setTimeout(function(){{ chartInstances[sectionId].resize(); }}, 100);
  }}

  // 导出图表为 PNG（任务8）：背景色随主题
  function exportChartPNG(sectionId) {{
    if (!chartInstances[sectionId]) return;
    var bg = isDarkTheme() ? '#16213e' : '#ffffff';
    var url = chartInstances[sectionId].getDataURL({{type: 'png', pixelRatio: 2, backgroundColor: bg}});
    var a = document.createElement('a');
    a.href = url; a.download = sectionId + '.png';
    document.body.appendChild(a); a.click(); document.body.removeChild(a);
  }}

  // 堆叠/并排切换（仅柱状图，任务8）
  function toggleStack(sectionId) {{
    if (!window._stackState) window._stackState = {{}};
    window._stackState[sectionId] = !window._stackState[sectionId];
    onColChange(sectionId);
  }}

  // 目录展开/收起全部（任务4）
  function expandAllToc() {{
    var groups = document.querySelectorAll('.toc-group');
    for (var i = 0; i < groups.length; i++) groups[i].classList.add('open');
    saveTocState();
  }}
  function collapseAllToc() {{
    var groups = document.querySelectorAll('.toc-group');
    for (var i = 0; i < groups.length; i++) groups[i].classList.remove('open');
    saveTocState();
  }}

  // 单个目录组切换 + 持久化（任务9）
  function toggleTocGroup(el) {{
    el.parentElement.classList.toggle('open');
    saveTocState();
  }}
  function saveTocState() {{
    try {{
      var groups = document.querySelectorAll('.toc-group');
      var state = [];
      for (var i = 0; i < groups.length; i++) state.push(groups[i].classList.contains('open'));
      localStorage.setItem('toc_open_state', JSON.stringify(state));
    }} catch(e) {{}}
  }}
  function restoreTocState() {{
    try {{
      var saved = localStorage.getItem('toc_open_state');
      if (!saved) return;
      var state = JSON.parse(saved);
      var groups = document.querySelectorAll('.toc-group');
      for (var i = 0; i < groups.length && i < state.length; i++) {{
        if (state[i]) groups[i].classList.add('open');
        else groups[i].classList.remove('open');
      }}
    }} catch(e) {{}}
  }}

  // 全局搜索：高亮匹配的 section 标题并展开对应目录（任务9）
  function globalSearch(keyword) {{
    var kw = (keyword || '').toLowerCase().trim();
    var sections = document.querySelectorAll('.section');
    var links = document.querySelectorAll('.sidebar a');
    for (var i = 0; i < sections.length; i++) sections[i].classList.remove('search-hit');
    for (var i = 0; i < links.length; i++) links[i].classList.remove('search-hit');
    if (!kw) return;
    for (var i = 0; i < sections.length; i++) {{
      var titleEl = sections[i].querySelector('h2');
      if (titleEl && titleEl.textContent.toLowerCase().indexOf(kw) >= 0) {{
        sections[i].classList.add('search-hit');
        var sid = sections[i].id;
        for (var j = 0; j < links.length; j++) {{
          if ((links[j].getAttribute('href') || '') === '#' + sid) {{
            links[j].classList.add('search-hit');
            var grp = links[j].closest('.toc-group');
            if (grp) grp.classList.add('open');
          }}
        }}
      }}
    }}
  }}

  window.addEventListener('resize', function() {{
    Object.keys(chartInstances).forEach(function(key) {{
      chartInstances[key].resize();
    }});
  }});

  // ESC 退出全屏图表
  document.addEventListener('keydown', function(e) {{
    if (e.key === 'Escape' || e.keyCode === 27) {{
      var fs = document.querySelector('.chart-container.fullscreen');
      if (fs) fs.classList.remove('fullscreen');
    }}
  }});

  // 滚动监听：返回顶部按钮显示 + 目录高亮（任务9）
  var _backToTopBtn = document.getElementById('backToTopBtn');
  window.addEventListener('scroll', function() {{
    if (_backToTopBtn) _backToTopBtn.classList.toggle('show', window.pageYOffset > 300);
  }});

  function initScrollSpy() {{
    var sections = document.querySelectorAll('.section[id]');
    if (!('IntersectionObserver' in window)) return;
    var io = new IntersectionObserver(function(entries) {{
      entries.forEach(function(entry) {{
        if (entry.isIntersecting) {{
          var sid = entry.target.id;
          var links = document.querySelectorAll('.sidebar a');
          for (var i = 0; i < links.length; i++) {{
            links[i].classList.toggle('active', (links[i].getAttribute('href') || '') === '#' + sid);
          }}
        }}
      }});
    }}, {{rootMargin: '-80px 0px -70% 0px'}});
    for (var i = 0; i < sections.length; i++) io.observe(sections[i]);
  }}

  // 暗色模式：获取当前激活的图表类型
  function getCurrentChartType(sectionId) {{
    var btns = document.querySelectorAll('#' + sectionId + ' .chart-btn');
    for (var i = 0; i < btns.length; i++) {{
      if (btns[i].classList.contains('active')) {{
        var m = btns[i].getAttribute('onclick').match(/'([^']+)'$/);
        if (m) return m[1];
      }}
    }}
    return 'bar';
  }}

  // 暗色模式：切换主题 + 持久化 + 重新渲染所有图表
  function toggleTheme() {{
    var cur = document.documentElement.getAttribute('data-theme') || 'light';
    var next = cur === 'dark' ? 'light' : 'dark';
    document.documentElement.setAttribute('data-theme', next);
    try {{ localStorage.setItem('report_theme', next); }} catch(e) {{}}
    var btn = document.getElementById('themeToggleBtn');
    if (btn) btn.textContent = next === 'dark' ? '☀️ 亮色模式' : '🌙 暗色模式';
    // 重新渲染所有已存在的图表以适配主题配色
    Object.keys(chartInstances).forEach(function(sid) {{
      renderChart(sid, getCurrentChartType(sid));
    }});
  }}

  // 暗色模式：初始化主题（localStorage 优先，否则检测系统偏好）
  function initTheme() {{
    var theme = 'light';
    try {{ theme = localStorage.getItem('report_theme') || ''; }} catch(e) {{}}
    if (!theme) {{
      if (window.matchMedia && window.matchMedia('(prefers-color-scheme: dark)').matches) {{
        theme = 'dark';
      }} else {{
        theme = 'light';
      }}
    }}
    document.documentElement.setAttribute('data-theme', theme);
    var btn = document.getElementById('themeToggleBtn');
    if (btn) btn.textContent = theme === 'dark' ? '☀️ 亮色模式' : '🌙 暗色模式';
  }}

  document.addEventListener('DOMContentLoaded', function() {{
    initTheme();
    restoreTocState();
    initCharts();
    initScrollSpy();
  }});
</script>
</body>
</html>"""

    html_path = os.path.join(output_dir, f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"[HTML] 报告已生成: {html_path}")
    return html_path


def start_preview_server(html_path: str) -> Tuple[str, object]:
    import http.server
    import threading
    import socket

    for port in range(8765, 8780):
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.bind(("", port))
                break
        except OSError:
            continue

    directory = os.path.dirname(html_path)
    filename = os.path.basename(html_path)

    handler = lambda *args: http.server.SimpleHTTPRequestHandler(*args, directory=directory)
    server = http.server.HTTPServer(("0.0.0.0", port), handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()

    url = f"http://localhost:{port}/{filename}"
    print(f"[HTML] 预览服务已启动: {url}")
    return url, server
