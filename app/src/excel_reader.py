import os
import openpyxl
import pandas as pd
import glob


# ==================== 通用数据文件读取 ====================

SUPPORTED_DATA_EXTS = (".xlsx", ".xls", ".csv")


def find_data_file(data_source, config_dir):
    """
    根据数据源名称查找数据文件，支持 Excel 和 CSV。
    
    匹配逻辑（按优先级）：
    1. 精确匹配文件名（带扩展名）
    2. 在配置目录中查找匹配的文件（计算匹配度）
    
    Args:
        data_source: 数据源名称（如 "网络指标数据"、"数据.csv"）
        config_dir: 配置所在目录
        
    Returns:
        找到的文件路径，未找到返回 None
    """
    if not data_source:
        return None
        
    data_source = str(data_source).strip()
    
    # 如果是绝对路径
    if os.path.isabs(data_source):
        if os.path.exists(data_source):
            return data_source
        # 尝试添加常见扩展名
        for ext in SUPPORTED_DATA_EXTS:
            path = data_source + ext
            if os.path.exists(path):
                return path
        return None
    
    # 在配置目录中查找
    search_dir = config_dir if config_dir else os.path.dirname(data_source)
    
    # 1. 精确匹配（带扩展名）
    exact_path = os.path.join(search_dir, data_source)
    if os.path.exists(exact_path):
        return exact_path
    
    # 2. 尝试添加扩展名
    for ext in SUPPORTED_DATA_EXTS:
        path = os.path.join(search_dir, data_source + ext)
        if os.path.exists(path):
            return path
    
    # 3. 模糊匹配 - 找匹配度最高的文件
    all_files = get_candidate_data_files(search_dir)
    if not all_files:
        return None
    
    best_match = None
    best_score = 0
    
    # 移除扩展名的名称用于匹配
    name_without_ext = data_source
    for ext in SUPPORTED_DATA_EXTS:
        if data_source.lower().endswith(ext):
            name_without_ext = data_source[:-len(ext)]
            break
    
    for f in all_files:
        basename = os.path.basename(f)
        # 移除扩展名
        for ext in SUPPORTED_DATA_EXTS:
            if basename.lower().endswith(ext):
                basename = basename[:-len(ext)]
                break
        
        score = _calculate_match_score(name_without_ext, basename)
        if score > best_score and score >= 0.5:  # 至少50%匹配度
            best_score = score
            best_match = f
    
    return best_match


def _calculate_match_score(query, candidate):
    """
    计算两个字符串的匹配度分数。
    考虑：包含关系、编辑距离、关键词匹配
    """
    query_lower = query.lower()
    candidate_lower = candidate.lower()
    
    # 完全相等
    if query_lower == candidate_lower:
        return 1.0
    
    # 包含关系
    if query_lower in candidate_lower:
        return 0.8 + 0.1 * (len(query_lower) / max(len(candidate_lower), 1))
    if candidate_lower in query_lower:
        return 0.7
    
    # 关键词匹配（中文分词效果较差，用简单包含判断）
    query_parts = _split_identifier(query_lower)
    candidate_parts = _split_identifier(candidate_lower)
    
    matched = sum(1 for qp in query_parts if any(qp in cp or cp in qp for cp in candidate_parts))
    if query_parts:
        return 0.5 * (matched / len(query_parts))
    
    return 0.0


def _split_identifier(s):
    """拆分标识符为关键词部分"""
    import re
    # 按中文字符、下划线、大写字母分割
    parts = re.split(r'[\u4e00-\u9fff_]', s)
    # 进一步按大写字母分割（驼峰命名）
    result = []
    for p in parts:
        if p:
            sub_parts = re.findall(r'[a-z]+|[A-Z][a-z]*|[A-Z]+', p)
            result.extend([sp.lower() for sp in sub_parts if len(sp) > 1])
    return [p for p in result if p]


def get_candidate_data_files(config_dir):
    """获取配置目录中所有候选数据文件（Excel和CSV）"""
    if not config_dir or not os.path.exists(config_dir):
        return []
    
    files = []
    for ext in SUPPORTED_DATA_EXTS:
        pattern = os.path.join(config_dir, f"*{ext}")
        files.extend(glob.glob(pattern))
    
    # 过滤临时文件和配置类文件
    result = []
    for f in files:
        basename = os.path.basename(f)
        if basename.startswith("~$"):
            continue
        if "配置" in basename or "config" in basename.lower():
            continue
        if "template" in basename.lower() or "模板" in basename:
            continue
        result.append(f)
    
    return result


def read_data_file(file_path, sheet_name=None):
    """
    通用数据文件读取函数，自动根据扩展名选择读取方式。
    
    Args:
        file_path: 文件路径（.xlsx, .xls, .csv）
        sheet_name: Sheet名称（仅对Excel有效）
        
    Returns:
        pandas.DataFrame
    """
    if not file_path or not os.path.exists(file_path):
        raise FileNotFoundError(f"数据文件不存在: {file_path}")
    
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == ".csv":
        return pd.read_csv(file_path, encoding="utf-8-sig")
    else:
        # Excel 文件
        if sheet_name:
            return pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            return pd.read_excel(file_path)


def get_data_file_sheets(file_path):
    """获取Excel文件的所有Sheet名称，CSV文件返回None"""
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == ".csv":
        return None
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception:
        return None


def get_data_file_info(file_path):
    """获取数据文件的基本信息"""
    if not file_path or not os.path.exists(file_path):
        return None
    
    ext = os.path.splitext(file_path)[1].lower()
    basename = os.path.basename(file_path)
    
    info = {
        "path": file_path,
        "name": basename,
        "type": "csv" if ext == ".csv" else "excel",
        "size": os.path.getsize(file_path),
        "sheets": None
    }
    
    if ext != ".csv":
        info["sheets"] = get_data_file_sheets(file_path)
    
    return info


# ==================== 原有配置读取函数 ====================


def read_config(config_path, sheet_name=None):
    wb = openpyxl.load_workbook(config_path, data_only=True)
    ws = _find_or_get_sheet(wb, sheet_name, _PPT_KEYWORDS)
    config = _parse_config(ws)
    wb.close()
    return config


_PPT_KEYWORDS = {"页码", "页面类型", "页面标题", "图表类型", "布局"}
_PIVOT_KEYWORDS = {"数据源", "行维度", "列维度", "值字段", "聚合方式"}


def _find_or_get_sheet(wb, sheet_name, keywords):
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]
    for name in wb.sheetnames:
        ws = wb[name]
        row = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if len(keywords & set(row)) >= 2:
            return ws
    return wb[wb.sheetnames[0]]


def _parse_config(ws):
    all_rows = list(ws.iter_rows(values_only=True))

    if not all_rows:
        return {"general": {}, "pages": [], "colors": {}}

    header_idx = None
    for i, row in enumerate(all_rows):
        if row[0] is not None and str(row[0]).strip() in ["页码", "序号", "编号"]:
            header_idx = i
            break

    if header_idx is None:
        header_idx = 0

    headers = [str(c).strip() if c is not None else "" for c in all_rows[header_idx]]

    raw_rows = []
    for row in all_rows[header_idx + 1:]:
        if row[0] is None and all(c is None for c in row):
            continue
        item = {}
        for h, v in zip(headers, row):
            if h:
                item[h] = v
        raw_rows.append(item)

    pages = _group_rows_to_pages(raw_rows)

    return {
        "general": {},
        "colors": {},
        "pages": pages,
    }


def _group_rows_to_pages(raw_rows):
    """按页码分组，相同页码的行合并为一页，每页含charts列表"""
    pages = []
    current_page = None
    current_page_num = None
    last_sheet = None
    last_data_source = None

    for row in raw_rows:
        page_num = row.get("页码")

        if page_num is not None and str(page_num).strip() != "":
            if current_page is not None:
                pages.append(current_page)
            last_sheet = None
            last_data_source = None
            page_title = str(row.get("页面标题", "")).strip() if row.get("页面标题") is not None else ""
            sub_title = str(row.get("副标题", "")).strip() if row.get("副标题") is not None else ""
            if sub_title:
                page_title = page_title + "|" + sub_title if page_title else sub_title
            current_page = {
                "页码": page_num,
                "页面类型": str(row.get("页面类型", "内容")).strip() or "内容",
                "页面标题": page_title,
                "布局": str(row.get("布局", "")).strip() if row.get("布局") is not None else "",
                "charts": [],
            }
            current_page_num = page_num

        if current_page is None:
            continue

        chart_title = str(row.get("图表标题", "")).strip() if row.get("图表标题") else ""
        if chart_title:
            sheet_val = str(row.get("数据Sheet", "")).strip() if row.get("数据Sheet") else ""
            if sheet_val:
                last_sheet = sheet_val
            elif last_sheet:
                sheet_val = last_sheet
            else:
                sheet_val = "Sheet1"
            
            # 数据源：支持 Excel 或 CSV 文件
            data_source = str(row.get("数据源", "")).strip() if row.get("数据源") else ""
            if data_source:
                last_data_source = data_source
            elif last_data_source:
                data_source = last_data_source
            
            chart_def = {
                "图表标题": chart_title,
                "图表类型": str(row.get("图表类型", "column")).strip() if row.get("图表类型") else "column",
                "数据Sheet": sheet_val,
                "数据源": data_source,
                "X轴范围": str(row.get("X轴范围", row.get("X轴", ""))).strip() if (row.get("X轴范围") is not None or row.get("X轴") is not None) else "",
                "Y轴范围": str(row.get("Y轴范围", row.get("Y轴", ""))).strip() if (row.get("Y轴范围") is not None or row.get("Y轴") is not None) else "",
                "颜色": str(row.get("颜色", "")).strip() if row.get("颜色") is not None else "",
                "区块名": str(row.get("区块名", "")).strip() if row.get("区块名") is not None else "",
                "结论模板": str(row.get("结论模板", "")).strip() if row.get("结论模板") is not None else "",
            }
            current_page["charts"].append(chart_def)

    if current_page is not None:
        pages.append(current_page)

    return pages


def read_data(file_path, sheet_name, x_range, y_range, block_name=None):
    """
    读取数据用于图表生成，支持 Excel 和 CSV 文件。
    
    Args:
        file_path: 数据文件路径（.xlsx, .xls, .csv）
        sheet_name: Sheet名称（CSV文件忽略此参数）
        x_range: X轴列名
        y_range: Y轴列名
        block_name: 数据区块名称
    """
    x_col_names = [cn.strip() for cn in str(x_range).split(",") if cn.strip()] if x_range else []
    y_col_names = [cn.strip() for cn in str(y_range).split(",") if cn.strip()] if y_range else []

    if not file_path:
        print("    [警告] 数据文件路径为空")
        return [], {}

    ext = os.path.splitext(str(file_path))[1].lower()
    is_csv = ext == ".csv"

    if is_csv:
        if block_name and str(block_name).strip():
            print(f"    [警告] CSV文件不支持区块名，将读取全量数据")
        return _read_dataframe_columns(file_path, sheet_name, x_range, y_range)
    
    # Excel 文件使用原有逻辑
    if block_name and str(block_name).strip():
        return _read_with_block_name(file_path, sheet_name, str(block_name).strip(), x_range, y_range)

    if len(x_col_names) >= 2:
        return _read_hierarchical_multi_y(file_path, sheet_name, x_col_names, y_col_names)

    x_values = _read_axis(file_path, sheet_name, x_range, is_x=True)
    y_values = _read_axis(file_path, sheet_name, y_range, is_x=False)
    return x_values, y_values


def _read_dataframe_columns(file_path, sheet_name, x_range, y_range):
    """
    使用 pandas 读取 CSV/Excel 数据列，适用于 CSV 和简单的 Excel 读取。
    """
    df = read_data_file(file_path, sheet_name)
    
    x_col_names = [cn.strip() for cn in str(x_range).split(",") if cn.strip()] if x_range else []
    y_col_names = [cn.strip() for cn in str(y_range).split(",") if cn.strip()] if y_range else []
    
    # 查找匹配的列
    x_cols = [c for c in df.columns if any(xn in c or c in xn for xn in x_col_names)] if x_col_names else []
    y_cols = [c for c in df.columns if any(yn in c or c in yn for yn in y_col_names)] if y_col_names else []
    
    # 如果没有匹配，尝试直接使用列名
    if not x_cols:
        x_cols = [c for c in df.columns if c in x_col_names] if x_col_names else []
    if not y_cols:
        y_cols = [c for c in df.columns if c in y_col_names] if y_col_names else []
    
    if not x_cols:
        # 尝试使用第一个文本列作为 X 轴
        text_cols = [c for c in df.columns if df[c].dtype == 'object']
        if text_cols:
            x_cols = [text_cols[0]]
    
    if not x_cols:
        return [], {}
    
    # 构建 X 轴值和 Y 轴值
    x_values = []
    y_values = {}
    
    for yc in y_cols:
        y_values[yc] = []
    
    for idx, row in df.iterrows():
        x_parts = []
        for xc in x_cols:
            val = row.get(xc)
            if pd.notna(val):
                x_parts.append(str(val))
        if x_parts:
            x_values.append(" - ".join(x_parts))
        else:
            x_values.append(str(idx))
        
        for yc in y_cols:
            val = row.get(yc)
            y_values[yc].append(val if pd.notna(val) else None)
    
    # 如果只有一个 Y 轴列，返回简化格式
    if len(y_cols) == 1:
        return x_values, y_values[y_cols[0]]
    
    return x_values, y_values


def _read_with_block_name(excel_path, sheet_name, block_name, x_range, y_range):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column

    block_row = _find_block_name_row(ws, max_row, max_col, block_name)
    if block_row is None:
        wb.close()
        return [], []

    header_row = block_row + 1
    if header_row > max_row:
        wb.close()
        return [], []

    header_map = {}
    for attempt in range(3):
        hrow = header_row + attempt
        if hrow > max_row:
            break
        for c in range(1, max_col + 1):
            cell_val = ws.cell(row=hrow, column=c).value
            if cell_val is not None and str(cell_val).strip():
                col_name = str(cell_val).strip()
                if col_name not in header_map:
                    header_map[col_name] = c
        if header_map:
            break

    x_col_names = [cn.strip() for cn in str(x_range).split(",") if cn.strip()]
    y_col_names = [cn.strip() for cn in str(y_range).split(",") if cn.strip()]

    x_cols = [header_map[cn] for cn in x_col_names if cn in header_map]
    y_cols = [header_map[cn] for cn in y_col_names if cn in header_map]
    missing_y = [cn for cn in y_col_names if cn not in header_map]
    if missing_y:
        print(f"    [警告] 区块数据中未找到以下Y轴列: {missing_y}")

    if not x_cols:
        wb.close()
        return [], []

    if len(x_cols) >= 2:
        result = _read_hierarchical_multi_from_ws(ws, header_row, max_row, max_col, x_cols, y_cols, x_col_names, y_col_names)
        wb.close()
        return result

    x_values = []
    y_values = {}
    for ycn in y_col_names:
        y_values[ycn] = []

    for r in range(header_row + 1, max_row + 1):
        all_cells = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if len(set(all_cells)) == 1 and all_cells[0] is None:
            break

        if _looks_like_header(ws, r, x_cols, x_col_names):
            break

        if len(x_cols) == 1:
            v = ws.cell(row=r, column=x_cols[0]).value
            if v is not None:
                x_values.append(str(v))
        else:
            parts = [str(ws.cell(row=r, column=xc).value) for xc in x_cols
                     if ws.cell(row=r, column=xc).value is not None]
            if parts:
                x_values.append(" - ".join(parts))

        for ycn, yc in zip(y_col_names, y_cols):
            v = ws.cell(row=r, column=yc).value
            if v is not None:
                y_values[ycn].append(v)

    wb.close()

    if len(y_cols) == 1:
        y_values = y_values.get(y_col_names[0], [])

    return x_values, y_values


def _read_grouped_from_ws(ws, header_row, max_row, max_col, x_cols, y_col, x_col_names):
    cat_col = x_cols[0]
    group_col = x_cols[1]

    categories = []
    cat_set = set()
    raw_map = {}

    for r in range(header_row + 1, max_row + 1):
        all_cells = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if len(set(all_cells)) == 1 and all_cells[0] is None:
            break

        if _looks_like_header(ws, r, x_cols, x_col_names):
            break

        cat_val = ws.cell(row=r, column=cat_col).value
        grp_val = ws.cell(row=r, column=group_col).value
        y_val = ws.cell(row=r, column=y_col).value

        if cat_val is None or y_val is None:
            continue

        cat_str = str(cat_val)
        grp_str = str(grp_val) if grp_val is not None else ""

        if cat_str not in cat_set:
            categories.append(cat_str)
            cat_set.add(cat_str)

        if grp_str not in raw_map:
            raw_map[grp_str] = {}
        raw_map[grp_str][cat_str] = y_val

    y_values = {}
    for grp_str in raw_map:
        series = []
        for cat in categories:
            series.append(raw_map[grp_str].get(cat, None))
        y_values[grp_str] = series

    active_categories = set()
    for grp_str in raw_map:
        active_categories.update(raw_map[grp_str].keys())
    categories = [c for c in categories if c in active_categories]

    cat_series_count = {c: 0 for c in categories}
    for grp_str in raw_map:
        for c in categories:
            if c in raw_map[grp_str]:
                cat_series_count[c] += 1

    categories = [c for c in categories if cat_series_count.get(c, 0) >= 1]

    y_values = {}
    for grp_str in raw_map:
        series = [raw_map[grp_str].get(cat) for cat in categories]
        valid_count = sum(1 for v in series if v is not None)
        if valid_count >= 1:
            y_values[grp_str] = series

    return categories, y_values

    return categories, y_values


def _read_grouped(excel_path, sheet_name, x_col_names, y_col_name):
    df = pd.read_excel(excel_path, sheet_name=sheet_name) if not str(excel_path).lower().endswith('.csv') else pd.read_csv(excel_path)

    cat_col = x_col_names[0]
    group_col = x_col_names[1]

    if cat_col not in df.columns or group_col not in df.columns or y_col_name not in df.columns:
        return [], []

    categories = []
    cat_set = set()
    raw_map = {}

    for _, row in df.iterrows():
        cat_val = row.get(cat_col)
        grp_val = row.get(group_col)
        y_val = row.get(y_col_name)

        if pd.isna(cat_val) or pd.isna(y_val):
            continue

        cat_str = str(cat_val)
        grp_str = str(grp_val) if pd.notna(grp_val) else ""

        if cat_str not in cat_set:
            categories.append(cat_str)
            cat_set.add(cat_str)

        if grp_str not in raw_map:
            raw_map[grp_str] = {}
        raw_map[grp_str][cat_str] = y_val

    y_values = {}
    for grp_str in raw_map:
        series = []
        for cat in categories:
            series.append(raw_map[grp_str].get(cat, None))
        y_values[grp_str] = series

    active_categories = set()
    for grp_str in raw_map:
        active_categories.update(raw_map[grp_str].keys())
    categories = [c for c in categories if c in active_categories]

    cat_series_count = {c: 0 for c in categories}
    for grp_str in raw_map:
        for c in categories:
            if c in raw_map[grp_str]:
                cat_series_count[c] += 1

    categories = [c for c in categories if cat_series_count.get(c, 0) >= 1]

    y_values = {}
    for grp_str in raw_map:
        series = [raw_map[grp_str].get(cat) for cat in categories]
        valid_count = sum(1 for v in series if v is not None)
        if valid_count >= 1:
            y_values[grp_str] = series

    return categories, y_values

    return categories, y_values


def _read_hierarchical_multi_y(excel_path, sheet_name, x_col_names, y_col_names):
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    cat_col = x_col_names[0]
    sub_col = x_col_names[1]
    if cat_col not in df.columns or sub_col not in df.columns:
        return [], {}
    hierarchical = []
    y_values = {ycn: [] for ycn in y_col_names}
    current_parent = None
    current_children = []
    for _, row in df.iterrows():
        cat_val = row.get(cat_col)
        if pd.isna(cat_val):
            continue
        cat_str = str(cat_val)
        if cat_str != current_parent:
            if current_children:
                hierarchical.append((current_parent, current_children))
            current_parent = cat_str
            current_children = []
        sub_val = row.get(sub_col)
        sub_str = str(sub_val) if pd.notna(sub_val) else ""
        children_y = {}
        for ycn in y_col_names:
            if ycn in df.columns:
                v = row.get(ycn)
                if pd.notna(v):
                    children_y[ycn] = v
        current_children.append((sub_str, children_y))
        for ycn in y_col_names:
            if ycn in children_y:
                y_values[ycn].append(children_y[ycn])
            else:
                y_values[ycn].append(None)
    if current_children:
        hierarchical.append((current_parent, current_children))
    if len(y_col_names) == 1:
        y_values = {y_col_names[0]: [v for v in y_values[y_col_names[0]] if v is not None]}
    return hierarchical, y_values


def _read_hierarchical(excel_path, sheet_name, x_col_names, y_col_name):
    hierarchical, y_values = _read_hierarchical_multi_y(excel_path, sheet_name, x_col_names, [y_col_name])
    flat = y_values.get(y_col_name, [])
    return hierarchical, {y_col_name: flat}


def _read_hierarchical_multi_from_ws(ws, header_row, max_row, max_col, x_cols, y_cols, x_col_names, y_col_names):
    cat_col = x_cols[0]
    sub_col = x_cols[1]
    hierarchical = []
    y_values = {ycn: [] for ycn, yc in zip(y_col_names, y_cols) if yc is not None}
    current_parent = None
    current_children = []
    for r in range(header_row + 1, max_row + 1):
        all_cells = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if len(set(all_cells)) == 1 and all_cells[0] is None:
            if current_children:
                hierarchical.append((current_parent, current_children))
            break
        if _looks_like_header(ws, r, x_cols, x_col_names):
            if current_children:
                hierarchical.append((current_parent, current_children))
            break
        cat_val = ws.cell(row=r, column=cat_col).value
        if cat_val is None:
            continue
        cat_str = str(cat_val)
        if cat_str != current_parent:
            if current_children:
                hierarchical.append((current_parent, current_children))
            current_parent = cat_str
            current_children = []
        sub_val = ws.cell(row=r, column=sub_col).value
        sub_str = str(sub_val) if sub_val is not None else ""
        children_y = {}
        for ycn, yc in zip(y_col_names, y_cols):
            if yc is not None:
                v = ws.cell(row=r, column=yc).value
                if v is not None:
                    children_y[ycn] = v
        current_children.append((sub_str, children_y))
        for ycn, yc in zip(y_col_names, y_cols):
            if ycn in children_y:
                y_values[ycn].append(children_y[ycn])
            else:
                y_values[ycn].append(None)
    if current_children:
        hierarchical.append((current_parent, current_children))
    if len(y_col_names) == 1:
        y_values = {y_col_names[0]: [v for v in y_values[y_col_names[0]] if v is not None]}
    return hierarchical, y_values


def _read_hierarchical_from_ws(ws, header_row, max_row, max_col, x_cols, y_col, x_col_names, y_col_name):
    hierarchical, y_values = _read_hierarchical_multi_from_ws(ws, header_row, max_row, max_col, x_cols, [y_col], x_col_names, [y_col_name])
    return hierarchical, y_values


def _find_block_name_row(ws, max_row, max_col, block_name):
    """查找区块名所在行，首列精确匹配优先，再回退子串匹配。"""
    bn = block_name.strip().lower()
    for r in range(1, max_row + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val is not None and bn == str(cell_val).strip().lower():
            return r
    if not bn:
        return None
    for r in range(1, max_row + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val is not None and bn in str(cell_val).strip().lower():
            return r
    return None


def _looks_like_header(ws, row, x_cols, x_col_names):
    for xc, xn in zip(x_cols, x_col_names):
        val = ws.cell(row=row, column=xc).value
        if val is not None and str(val).strip() == xn:
            return True
    return False


def _is_range_format(s):
    s = str(s).strip()
    return ":" in s and any(c.isalpha() for c in s.split(":")[0])


def _read_axis(excel_path, sheet_name, ref, is_x=False):
    if ref is None or str(ref).strip() == "":
        return []

    ref_str = str(ref).strip()

    if _is_range_format(ref_str):
        return _read_range_by_range(excel_path, sheet_name, ref_str, combine_multi_col=is_x)

    return _read_range_by_columns(excel_path, sheet_name, ref_str, combine_multi_col=is_x)


def _fuzzy_match_column(cn, df_columns):
    """模糊匹配列名，精确匹配优先于子串匹配。"""
    s = str(cn).strip()
    for c in df_columns:
        if s == str(c).strip():
            return c
    for c in df_columns:
        if s in str(c):
            return c
    for c in df_columns:
        if str(c).strip() in s:
            return c
    return None


def _read_range_by_columns(excel_path, sheet_name, col_names_str, combine_multi_col=False):
    col_names = [c.strip() for c in str(col_names_str).split(",") if c.strip()]
    if not col_names:
        return []

    df = pd.read_excel(excel_path, sheet_name=sheet_name)

    available = []
    for cn in col_names:
        matched = _fuzzy_match_column(cn, df.columns)
        if matched:
            available.append(matched)

    if not available:
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=1)
        available = []
        for cn in col_names:
            matched = _fuzzy_match_column(cn, df.columns)
            if matched:
                available.append(matched)

    if not available:
        return []

    if len(available) == 1:
        return df[available[0]].dropna().tolist()

    if not combine_multi_col:
        series_cols = []
        for cn in col_names:
            matched = _fuzzy_match_column(cn, df.columns)
            if matched and matched not in series_cols:
                series_cols.append(matched)
        if len(series_cols) == 1:
            return df[series_cols[0]].dropna().tolist()
        if len(series_cols) > 1:
            result = {}
            for cn in series_cols:
                result[cn] = df[cn].dropna().tolist()
            return result

    combined = []
    for idx, row in df.iterrows():
        parts = []
        all_none = True
        for cn in col_names:
            matched = _fuzzy_match_column(cn, df.columns)
            if matched:
                val = row[matched]
                if pd.notna(val):
                    all_none = False
                    parts.append(str(val))
                else:
                    parts.append("")
        if not all_none:
            valid_parts = [p for p in parts if p != ""]
            combined.append(" - ".join(valid_parts))

    return combined


def _read_range_by_range(excel_path, sheet_name, cell_range, combine_multi_col=False):
    range_str = str(cell_range).strip()
    if "!" in range_str:
        sheet_name, range_str = range_str.split("!", 1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    start_cell, end_cell = range_str.split(":")
    start_col = _col_letter_to_index(start_cell)
    end_col = _col_letter_to_index(end_cell)
    start_row = int("".join(c for c in start_cell if c.isdigit()))
    end_row = int("".join(c for c in end_cell if c.isdigit()))

    num_cols = end_col - start_col + 1

    if num_cols <= 1 or not combine_multi_col:
        values = []
        for row in ws[range_str]:
            for cell in row:
                values.append(cell.value)
        wb.close()
        return values

    result = []
    for r in range(start_row, end_row + 1):
        parts = []
        all_none = True
        for c in range(start_col, end_col + 1):
            val = ws.cell(row=r, column=c + 1).value
            if val is not None:
                all_none = False
                parts.append(str(val))
            else:
                parts.append("")
        if not all_none:
            valid_parts = [p for p in parts if p != ""]
            result.append(" - ".join(valid_parts))

    wb.close()
    return result


def _col_letter_to_index(cell_ref):
    letters = "".join(c for c in cell_ref if c.isalpha()).upper()
    idx = 0
    for char in letters:
        idx = idx * 26 + (ord(char) - ord("A") + 1)
    return idx - 1


def read_data_as_dataframe(file_path, sheet_name=None):
    """
    读取数据文件为 DataFrame，支持 Excel 和 CSV。
    CSV 文件忽略 sheet_name 参数。
    """
    return read_data_file(file_path, sheet_name)


def read_geo_data(excel_path, sheet_name, x_range, y_range):
    x_cols = [cn.strip() for cn in str(x_range).split(",") if cn.strip()] if x_range else []
    y_cols = [cn.strip() for cn in str(y_range).split(",") if cn.strip()] if y_range else []

    all_cols = x_cols + y_cols
    if not all_cols:
        return None

    df = read_data_file(excel_path, sheet_name)
    available = [c for c in all_cols if c in df.columns]
    if not available:
        return None

    geo_df = df[available].copy()
    extra_cols = ["site_id", "cell_id", "siteid", "cellid", "站点ID", "小区ID"]
    for ec in extra_cols:
        if ec in df.columns and ec not in available:
            geo_df[ec] = df[ec]

    return geo_df
