import os
import re
import openpyxl
import pandas as pd
import numpy as np
from src.safe_math import evaluate_numeric_expression


# 已知的格式化字符串集合（与 template_filler.py 的 _KNOWN_FMTS 保持一致）
# 聚合方式的 |数字格式 和 PPT 占位符的 |格式 共用同一套语法
_KNOWN_FMTS = {
    # 小数位格式
    ".0f", ".1f", ".2f", ".3f", ".4f", ".5f", ".6f",
    # 整数格式
    "int", "d",
    # 百分比格式（自动 ×100 加 %）
    ".0%", ".1%", ".2%", ".3%", ".4%",
    # 科学计数法
    ".0e", ".1e", ".2e", ".3e", ".0E", ".1E", ".2E", ".3E",
    # 千分位
    ",.0f", ",.1f", ",.2f", ",.3f", ",.0%", ",.1%", ",.2%", ",.3%",
}


def _col_round_precision(col_name) -> int:
    """根据列名确定数值精度。经纬度坐标需要6位小数，其他默认2位。"""
    name = str(col_name).lower()
    if any(k in name for k in ["经度", "纬度", "lat", "lon", "lng", "longitude", "latitude"]):
        return 6
    return 2


AGG_MAP = {
    "sum": "sum",
    "avg": "mean",
    "mean": "mean",
    "count": "count",
    "max": "max",
    "min": "min",
    "nunique": "nunique",
    "去重计数": "nunique",
    "distinct": "nunique",
    "pct": "pct",
    "占比": "pct",
    "percentage": "pct",
    "count_pct": "count_pct",
    "计数占比": "count_pct",
    "join": "join",
    "拼接": "join",
    "去重拼接": "join",
    "group_concat": "join",
}


# join 聚合的默认分隔符
_JOIN_DEFAULT_SEP = "+"

# join 聚合的关键字集合（用于识别含分隔符参数的写法，如 "去重拼接|、"）
_JOIN_KEYWORDS = {"join", "拼接", "去重拼接", "group_concat"}


def _parse_join_sep(agg_str):
    """解析 join 聚合的分隔符参数。

    支持写法：
      "去重拼接"          → ("去重拼接", "+")  默认加号
      "去重拼接|、"        → ("去重拼接", "、") 指定分隔符
      "join|, "           → ("join", ", ")

    返回 (原始agg关键字, 分隔符)。非 join 聚合原样返回 (agg_str, None)。
    """
    if "|" not in agg_str:
        return agg_str, None
    head, tail = agg_str.split("|", 1)
    head = head.strip()
    tail = tail.strip()
    if head.lower() in _JOIN_KEYWORDS or head in _JOIN_KEYWORDS:
        sep = tail if tail else _JOIN_DEFAULT_SEP
        return head, sep
    return agg_str, None


def _parse_agg_format(agg_str):
    """解析聚合方式的数字格式后缀（用 ``|`` 分隔，与 join 的 ``|`` 正交）。

    与 PPT 模板占位符的格式语法完全一致（复用 _KNOWN_FMTS 白名单）：
      "sum|.2f"          → ("sum", ".2f")
      "avg|.1%"          → ("avg", ".1%")
      "sum|,.0f"         → ("sum", ",.0f")  千分位整数
      "去重拼接|、"       → ("去重拼接|、", None)  join 的 | 是分隔符，非已知格式
      "sum"              → ("sum", None)

    仅当 | 后内容命中 _KNOWN_FMTS 时才视为格式串，否则 | 视为 join 分隔符或表达式一部分。

    :return: (聚合方式字符串（保留 join 的 | 分隔符）, 数字格式)。无格式返回 (agg_str, None)。
    """
    if "|" not in agg_str:
        return agg_str, None
    head, tail = agg_str.split("|", 1)
    head = head.strip()
    tail = tail.strip()
    if not tail:
        return head, None
    # 仅当 | 后是已知格式才视为数字格式后缀
    if tail in _KNOWN_FMTS:
        return head, tail
    # 否则 | 是 join 分隔符或其他用途，原样返回
    return agg_str, None


# 已知聚合关键字集合，用于 _split_agg_funcs 识别多聚合分隔符
_AGG_KEYWORDS_SET = set(AGG_MAP.keys())


def _match_agg_keyword(s):
    """检查 s 是否以已知聚合关键字开头（后跟 , | 或结尾）。返回关键字或 None。"""
    for kw in _AGG_KEYWORDS_SET:
        if s.lower().startswith(kw.lower()):
            after = s[len(kw):]
            if not after or after[0] in (",", "|"):
                return kw
    return None


def _split_agg_funcs(agg_str):
    """按逗号拆分多聚合方式，保护 ``|`` 后参数内的逗号不被误判为分隔符。

    PPT 格式串可能含逗号（如 ``,.2f`` 千分位），直接 split 会被切断。
    本函数遇到 ``|`` 后进入参数模式，参数内的逗号属于参数；只有当逗号后
    紧跟已知聚合关键字时，才认为是多聚合分隔符。

    例：
      'sum,avg'                → ['sum', 'avg']
      'sum|,.0f,sum|.2f'       → ['sum|,.0f', 'sum|.2f']
      '去重拼接|、,sum|.2f'     → ['去重拼接|、', 'sum|.2f']
    """
    parts = []
    current = ""
    i = 0
    n = len(agg_str)
    while i < n:
        ch = agg_str[i]
        if ch == "|":
            # 进入参数模式，收集直到"逗号+聚合关键字"或结尾
            current += ch
            i += 1
            while i < n:
                ch2 = agg_str[i]
                if ch2 == ",":
                    rest = agg_str[i + 1:].lstrip()
                    if _match_agg_keyword(rest):
                        break  # 逗号后是聚合关键字，这是多聚合分隔符
                    current += ch2
                    i += 1
                else:
                    current += ch2
                    i += 1
            if current.strip():
                parts.append(current.strip())
            current = ""
        elif ch == ",":
            if current.strip():
                parts.append(current.strip())
            current = ""
            i += 1
        else:
            current += ch
            i += 1
    if current.strip():
        parts.append(current.strip())
    return parts


def _join_unique(series, sep: str = _JOIN_DEFAULT_SEP) -> str:
    """对 series 做去重拼接，排序后输出，跳过 NaN/空值。

    排序保证同一组数据无论行顺序如何，拼接结果一致。
    """
    seen_set = set()
    for v in series:
        if pd.isna(v):
            continue
        s = str(v).strip()
        if s:
            seen_set.add(s)
    return sep.join(sorted(seen_set))


def read_pivot_config(config_path, sheet_name=None):
    wb = openpyxl.load_workbook(config_path, data_only=True)

    from src.excel_reader import _find_or_get_sheet, _PIVOT_KEYWORDS
    ws = _find_or_get_sheet(wb, sheet_name, _PIVOT_KEYWORDS)

    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [str(c).strip() if c else "" for c in rows[0]]

    tasks = []
    for row in rows[1:]:
        if all(c is None for c in row):
            continue
        item = {}
        for h, v in zip(headers, row):
            if h:
                item[h] = v

        task = {}
        task["序号"] = item.get("序号", "")
        task["数据源"] = str(item.get("数据源", "")).strip() if item.get("数据源") else ""
        task["sheet"] = str(item.get("Sheet", item.get("sheet", ""))).strip() if (item.get("Sheet") or item.get("sheet")) else "Sheet1"
        task["行维度"] = str(item.get("行维度", "")).strip().replace("，", ",").replace("\n", ",").replace("\r", ",") if item.get("行维度") else ""
        task["列维度"] = str(item.get("列维度", "")).strip().replace("，", ",").replace("\n", ",").replace("\r", ",") if item.get("列维度") else ""
        task["值字段"] = str(item.get("值字段", "")).strip().replace("，", ",").replace("\n", ",").replace("\r", ",") if item.get("值字段") else ""
        task["聚合方式"] = str(item.get("聚合方式", "sum")).strip().replace("，", ",").replace("\n", ",").replace("\r", ",") if item.get("聚合方式") else "sum"
        task["结果Sheet"] = str(item.get("结果Sheet", "")).strip() if item.get("结果Sheet") else f"结果{task.get('序号','')}"
        task["区块名"] = str(item.get("区块名", item.get("备注", ""))).strip() if (item.get("区块名") or item.get("备注")) else ""
        task["行映射"] = str(item.get("行映射", "")).strip() if item.get("行映射") else (str(item.get("行维度映射", "")).strip() if item.get("行维度映射") else str(item.get("映射表", "")).strip() if item.get("映射表") else "")
        task["列映射"] = str(item.get("列映射", "")).strip() if item.get("列映射") else (str(item.get("列维度映射", "")).strip() if item.get("列维度映射") else "")
        task["值映射"] = str(item.get("值映射", "")).strip().replace("，", ",").replace("\n", ",").replace("\r", ",") if item.get("值映射") else (str(item.get("值字段映射", "")).strip().replace("，", ",").replace("\n", ",").replace("\r", ",") if item.get("值字段映射") else "")
        task["分箱"] = str(item.get("分箱", "")).strip().replace("，", ",").replace("\n", ",").replace("\r", ",") if item.get("分箱") else ""
        task["值计算"] = str(item.get("值计算", "")).strip().replace("，", ",").replace("\n", ",").replace("\r", ",") if item.get("值计算") else ""
        task["是否计算"] = str(item.get("是否计算", "是")).strip() if item.get("是否计算") else "是"
        task["过滤条件"] = str(item.get("过滤条件", "")).strip() if item.get("过滤条件") else ""

        if not task["行维度"] and not task["值字段"]:
            continue

        tasks.append(task)

    wb.close()
    return tasks


def _resolve_data_path(data_source, config_dir):
    if not data_source:
        return None
    if os.path.isabs(data_source):
        return data_source
    return os.path.join(config_dir, data_source)


def _resolve_block_reference(data_source, block_results):
    """解析前序透视结果区块引用。

    支持语法（按精确度从高到低）：
      - ``{区块名.结果Sheet名}`` → 精确匹配：同时匹配区块名和结果Sheet名（推荐，区块名不唯一时使用）
      - ``{pivot}`` / ``{透视结果}`` / ``{prev}`` / ``{上一个}`` → 取 block_results 中最后插入的 DataFrame
      - ``{区块名}`` → 取 block_results[区块名]（区块名唯一时使用）
      - ``{pivot:区块名}`` → 同上，显式指代透视结果中的某区块

    :return: 命中则返回 DataFrame.copy()；未命中或 block_results 为空则返回 None
    """
    if not block_results:
        return None
    ds = str(data_source).strip()
    if not ds:
        return None

    # {pivot:区块名} 形式
    m = re.match(r"^\{pivot[:：](.+)\}$", ds, re.IGNORECASE)
    if m:
        block_name = m.group(1).strip()
        if block_name in block_results:
            return block_results[block_name].copy()
        return None

    # 单层 {…} 引用
    m = re.match(r"^\{(.+)\}$", ds)
    if not m:
        return None
    inner = m.group(1).strip()
    inner_lower = inner.lower()

    # 通配：上一个/前序结果
    if inner_lower in ("pivot", "透视结果", "prev", "上一个", "上一个结果"):
        # 取 block_results 中最后插入的 DataFrame（Python 3.7+ dict 保序）
        last_df = None
        for df in block_results.values():
            last_df = df
        return last_df.copy() if last_df is not None else None

    # {区块名.结果Sheet名} 组合精确匹配（区块名不唯一时使用）
    # 注意：区块名和结果Sheet名本身不应包含 "."，若包含会导致解析歧义
    if "." in inner:
        parts = inner.split(".", 1)
        block_name_part = parts[0].strip()
        sheet_name_part = parts[1].strip()
        # 在 block_results 中查找组合 key (区块名, 结果Sheet名)
        for k, df in block_results.items():
            if isinstance(k, tuple) and len(k) == 2:
                k_block, k_sheet = k
                if str(k_block) == block_name_part and str(k_sheet) == sheet_name_part:
                    return df.copy()
        # 组合 key 未命中，尝试宽松匹配：区块名或Sheet名任一匹配且另一项也匹配
        return None

    # 指定区块名（单层引用）
    if inner in block_results:
        return block_results[inner].copy()
    # 区块名大小写不敏感兜底
    for k, df in block_results.items():
        if not isinstance(k, tuple) and k.lower() == inner_lower:
            return df.copy()

    return None


def _parse_join_spec(data_source):
    # 修复 P1 问题8：JOIN 检测对换行/制表符友好（Excel Alt+Enter 换行场景）
    if not re.search(r"\bJOIN\b", data_source, re.IGNORECASE):
        return None

    tokens = _tokenize_join(data_source)
    if not tokens:
        return None

    parts = []
    i = 0
    while i < len(tokens):
        if i == 0:
            left_table = tokens[i]
            i += 1
        else:
            left_table = parts[-1]["right"]

        if i >= len(tokens):
            break

        if tokens[i].upper() not in ("JOIN", "INNER JOIN", "LEFT JOIN", "RIGHT JOIN", "OUTER JOIN"):
            i += 1
            if i >= len(tokens):
                break

        join_type = tokens[i]
        i += 1
        if i >= len(tokens):
            break

        right_table = tokens[i]
        i += 1

        if i + 1 <= len(tokens) and i < len(tokens) and tokens[i].upper() == "ON":
            i += 1
            on_parts = []
            while i < len(tokens) and tokens[i].upper() not in ("JOIN", "INNER JOIN", "LEFT JOIN", "RIGHT JOIN", "OUTER JOIN"):
                on_parts.append(tokens[i])
                i += 1
            on_expr = " ".join(on_parts)

            left_keys = []
            right_keys = []
            and_conditions = re.split(r"\s+AND\s+", on_expr, flags=re.IGNORECASE)
            for cond in and_conditions:
                cond = cond.strip()
                # 修复 P1 问题7：只支持等值 JOIN（pandas merge 本身只支持等值）
                # 非等值条件（!=, >, <, >=, <=）显式报错而非静默丢弃
                non_eq_ops = ["!=", ">=", "<=", ">", "<"]
                if any(op in cond for op in non_eq_ops):
                    raise ValueError(
                        f"JOIN ON 仅支持等值条件(=)，不支持非等值: '{cond}'。"
                        f"如需非等值过滤，请在「过滤条件」列配置"
                    )
                if "=" in cond:
                    eq_parts = cond.split("=", 1)
                    left_keys.append(eq_parts[0].strip())
                    right_keys.append(eq_parts[1].strip())

            left_key = ",".join(left_keys) if left_keys else ""
            right_key = ",".join(right_keys) if right_keys else ""
        else:
            left_keys = []
            right_keys = []
            left_key = ""
            right_key = ""

        how = "inner"
        jt = join_type.upper()
        if "LEFT" in jt:
            how = "left"
        elif "RIGHT" in jt:
            how = "right"
        elif "OUTER" in jt:
            how = "outer"

        left_file, left_sheet = _split_table_token(left_table)
        right_file, right_sheet = _split_table_token(right_table)

        parts.append({
            "left": left_file,
            "right": right_file,
            "left_sheet": left_sheet,
            "right_sheet": right_sheet,
            "left_key": left_key,
            "right_key": right_key,
            "left_keys": left_keys,
            "right_keys": right_keys,
            "how": how,
        })

    return parts


def _split_table_token(token):
    if "@" in token:
        idx = token.index("@")
        return token[:idx], token[idx + 1:]
    return token, None


def _tokenize_join(data_source):
    s = data_source.strip()
    keywords = ["INNER JOIN", "LEFT JOIN", "RIGHT JOIN", "OUTER JOIN", "JOIN", "ON", "AND"]
    tokens = []
    remaining = s
    while remaining:
        found = False
        for kw in keywords:
            pattern = re.compile(r"^\s*" + re.escape(kw) + r"\s+", re.IGNORECASE)
            m = pattern.match(remaining)
            if m:
                tokens.append(kw)
                remaining = remaining[m.end():]
                found = True
                break
        if not found:
            m = re.match(r"^\s*(.+?)(?=\s+(?:INNER\s+JOIN|LEFT\s+JOIN|RIGHT\s+JOIN|OUTER\s+JOIN|JOIN|ON|AND)\b|$)", remaining, re.IGNORECASE)
            if m:
                raw = m.group(1).strip()
                tokens.append(raw)
                remaining = remaining[m.end():]
            else:
                remaining = remaining.strip()
                if remaining:
                    tokens.append(remaining)
                break
    return tokens


def _load_joined_dataframe(config_dir, data_source, sheet_name, block_results=None):
    """加载数据，支持 Excel/CSV 文件、JOIN 语句、以及前序透视结果区块引用。

    区块引用语法（将前序任务的输出 DataFrame 作为当前任务的输入数据源，实现二次透视）：
      - ``{pivot}`` / ``{透视结果}`` / ``{prev}`` / ``{上一个}`` → 上一个成功任务的输出
      - ``{区块名}`` → 引用指定区块名的结果（区块名来自前序任务的"区块名"或"结果Sheet"）
      - ``{pivot:区块名}`` → 显式指代透视结果中的某区块

    JOIN 语法（多表 JOIN 用空格连接，ON 条件用 AND 分隔多键）：
      - ``表A.xlsx JOIN 表B.xlsx ON k1=k1``
      - ``表A.xlsx@Sheet1 JOIN 表B.xlsx@Sheet2 ON k1=k1 AND k2=k2``
      - ``{区块名} JOIN 表B.xlsx ON k1=k1``（左表为前序透视结果区块引用）
      - 支持 INNER/LEFT/RIGHT/OUTER JOIN（默认 INNER）

    :param block_results: 前序任务累积的结果 {区块名: DataFrame}，由 _run_pivot_mode 滚动传入
    :raises ValueError: JOIN 右表加载失败、ON 列名不存在、左右键数量不一致时抛出
    """
    from src.excel_reader import find_data_file, read_data_file

    # 优先处理前序透视结果区块引用（{pivot}/{区块名}/{pivot:区块名}）
    pivot_df = _resolve_block_reference(data_source, block_results)
    if pivot_df is not None:
        return pivot_df

    join_parts = _parse_join_spec(data_source)

    if not join_parts:
        # 非 JOIN 情况：直接查找文件
        file_path = _resolve_data_path(data_source, config_dir)
        if not file_path or not os.path.exists(file_path):
            # 尝试模糊查找
            file_path = find_data_file(data_source, config_dir)
        if not file_path or not os.path.exists(file_path):
            return None
        return read_data_file(file_path, sheet_name)

    candidate_files = _collect_candidate_xlsx(config_dir)

    def _resolve_table_source(table_name, sheet_hint, config_dir, candidate_files, block_results):
        """解析单表数据源：优先区块引用，否则按文件名+sheet_hint 加载。

        :param sheet_hint: 由 _split_table_token 解析出的 @Sheet 提示，优先于模糊匹配
        """
        # 先处理区块引用（{pivot}/{区块名}）
        df = _resolve_block_reference(table_name, block_results)
        if df is not None:
            return df, None
        file_path, sheet = _resolve_join_table(table_name, config_dir, candidate_files)
        if file_path is not None:
            # sheet_hint 优先（@Sheet 语法），其次用 _resolve_join_table 的模糊匹配结果
            use_sheet = sheet_hint or sheet
            if use_sheet is None:
                use_sheet = "Sheet1"
            df = read_data_file(file_path, use_sheet)
            return df, use_sheet
        return None, None

    # 第一段 JOIN 的左表（使用 @Sheet 提示）
    first_left = join_parts[0]["left"]
    first_left_sheet = join_parts[0]["left_sheet"]
    df, left_sheet = _resolve_table_source(first_left, first_left_sheet, config_dir, candidate_files, block_results)
    if df is None:
        raise ValueError(f"JOIN 左表加载失败: {first_left}")

    for jp in join_parts:
        right_name = jp["right"]
        right_sheet_hint = jp["right_sheet"]
        df_right, right_sheet = _resolve_table_source(right_name, right_sheet_hint, config_dir,
                                                       candidate_files, block_results)
        if df_right is None:
            # 修复 P0 问题2：右表加载失败必须报错，不再静默跳过
            raise ValueError(f"JOIN 右表加载失败: {right_name}")

        left_on = jp["left_key"]
        right_on = jp["right_key"]
        left_keys = jp.get("left_keys", [])
        right_keys = jp.get("right_keys", [])

        # 修复 P0 问题3：多键 JOIN 左右键数量必须一致
        if len(left_keys) != len(right_keys):
            raise ValueError(
                f"JOIN ON 条件左右键数量不一致: 左 {left_keys}({len(left_keys)}个) vs 右 {right_keys}({len(right_keys)}个)"
            )

        # 统一走多键分支（单键也用列表形式处理，逻辑更清晰）
        # strip 列名两端空格
        left_keys = [k.strip() for k in left_keys]
        right_keys = [k.strip() for k in right_keys]

        # 校验列名存在
        missing_left = [k for k in left_keys if k not in df.columns]
        missing_right = [k for k in right_keys if k not in df_right.columns]
        if missing_left or missing_right:
            # 修复 P0 问题2：列名不存在必须报错
            msg_parts = []
            if missing_left:
                msg_parts.append(f"左表缺少列 {missing_left}")
            if missing_right:
                msg_parts.append(f"右表({right_name})缺少列 {missing_right}")
            raise ValueError(f"JOIN ON 列名不存在: {'; '.join(msg_parts)}")

        df = pd.merge(df, df_right, left_on=left_keys, right_on=right_keys,
                      how=jp["how"], suffixes=("", "_r"))

    return df


def _resolve_join_table(table_name, config_dir, candidate_files):
    """解析 JOIN 表名，支持 Excel 和 CSV 文件"""
    from src.excel_reader import find_data_file, get_data_file_sheets
    
    # 先尝试直接查找
    file_path = find_data_file(table_name, config_dir)
    if file_path:
        # 确定 sheet 名称
        sheets = get_data_file_sheets(file_path)
        if sheets:
            # 尝试匹配 sheet
            clean_name = table_name
            for ext in (".xlsx", ".csv"):
                if clean_name.lower().endswith(ext):
                    clean_name = clean_name[:-len(ext)]
                    break
            # 匹配最相似的 sheet
            for s in sheets:
                if clean_name.lower() in s.lower() or s.lower() in clean_name.lower():
                    return file_path, s
            return file_path, sheets[0]  # 默认第一个 sheet
        return file_path, None
    
    return None, None


def _sheet_exists(file_path, sheet_name):
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        result = sheet_name in wb.sheetnames
        wb.close()
        return result
    except Exception:
        return False


def _collect_candidate_xlsx(config_dir):
    """保留原有函数名，但改为收集所有数据文件（Excel + CSV）"""
    from src.excel_reader import get_candidate_data_files
    return get_candidate_data_files(config_dir)


def _apply_filter(df, filter_expr):
    """
    应用过滤条件到 DataFrame。
    
    支持的语法：
    - 简单比较: 列名=值, 列名>值, 列名<值, 列名>=值, 列名<=值, 列名!=值
    - 包含匹配: 列名包含'值', 列名 in ('a','b')
    - 多条件: 用 AND/OR 连接，支持括号
    
    示例:
        频段='n78' AND 上行PRB利用率>60
        地区 in ('北京','上海') OR 频段='n41'
        (频段='n78' AND 上行PRB利用率>60) OR 地区='北京'
    """
    if not filter_expr or not filter_expr.strip():
        return df
    
    expr = filter_expr.strip()
    
    # 将中文引号替换为英文引号
    expr = expr.replace("'", "'").replace("'", "'")
    
    # 使用 pandas 的 query 方法，但先做一些安全处理
    # 我们用 eval 方式更灵活，但需要注意安全
    # 为了安全，我们使用 pandas query
    try:
        result = df.query(expr)
        return result
    except Exception:
        # 如果 query 失败，尝试更简单的解析
        pass
    
    # 简单的自定义解析：处理 AND/OR 连接的简单条件
    return _simple_filter(df, expr)


def _simple_filter(df, expr):
    """简单过滤：支持 AND/OR 连接的简单比较条件"""
    # 按 AND 拆分（优先级：先 AND 后 OR）
    # 简化处理：先按 OR 拆分，每部分内按 AND 处理
    
    # 用正则处理带括号的情况太复杂，这里简化为只处理顶层 OR
    # 更完善的方案后续可以用 AST
    
    or_parts = _split_by_keyword(expr, "OR")
    
    if len(or_parts) > 1:
        # OR 条件：合并各部分结果
        result = None
        for part in or_parts:
            part = _strip_brackets(part.strip())
            sub = _simple_filter(df, part)
            if result is None:
                result = sub
            else:
                result = pd.concat([result, sub])
                result = result[~result.index.duplicated(keep='first')]
        return result if result is not None else df
    
    # 按 AND 拆分
    and_parts = _split_by_keyword(expr, "AND")
    if len(and_parts) > 1:
        result = df
        for part in and_parts:
            part = _strip_brackets(part.strip())
            result = _simple_filter(result, part)
        return result
    
    # 单个条件
    return _parse_single_condition(df, expr)


def _strip_brackets(s):
    while s.startswith("(") and s.endswith(")"):
        s = s[1:-1]
    return s


def _split_by_keyword(expr, keyword):
    """按关键字拆分表达式，忽略引号内的内容"""
    parts = []
    current = []
    in_quote = False
    quote_char = None
    i = 0
    kw_len = len(keyword)
    
    while i < len(expr):
        c = expr[i]
        if c in ("'", '"') and not in_quote:
            in_quote = True
            quote_char = c
            current.append(c)
        elif c == quote_char and in_quote:
            in_quote = False
            quote_char = None
            current.append(c)
        elif not in_quote and expr[i:i+kw_len].upper() == keyword.upper():
            # 检查前后是否是边界（空格或括号）
            before_ok = (i == 0 or expr[i-1] in (' ', '(', ')'))
            after_ok = (i + kw_len >= len(expr) or expr[i+kw_len] in (' ', '(', ')'))
            if before_ok and after_ok:
                parts.append("".join(current))
                current = []
                i += kw_len
                continue
            else:
                current.append(c)
        else:
            current.append(c)
        i += 1
    
    if current:
        parts.append("".join(current))
    
    return parts if len(parts) > 1 else [expr]


def _parse_single_condition(df, condition):
    """解析单个条件"""
    condition = condition.strip()
    
    # 支持的运算符
    operators = [">=", "<=", "!=", "=", ">", "<"]
    
    for op in operators:
        if op in condition:
            idx = condition.index(op)
            col = condition[:idx].strip()
            val = condition[idx+len(op):].strip()
            
            # 去除引号
            if (val.startswith("'") and val.endswith("'")) or (val.startswith('"') and val.endswith('"')):
                val = val[1:-1]
            
            if col not in df.columns:
                raise ValueError(f"过滤条件列不存在: {col}")
            
            if op == "=":
                return df[df[col].astype(str) == str(val)]
            elif op == "!=":
                return df[df[col].astype(str) != str(val)]
            elif op == ">":
                return df[pd.to_numeric(df[col], errors="coerce") > float(val)]
            elif op == "<":
                return df[pd.to_numeric(df[col], errors="coerce") < float(val)]
            elif op == ">=":
                return df[pd.to_numeric(df[col], errors="coerce") >= float(val)]
            elif op == "<=":
                return df[pd.to_numeric(df[col], errors="coerce") <= float(val)]
    
    # 包含关系: col in ('a','b') 或 列名包含'值'
    if " in (" in condition.lower():
        idx = condition.lower().index(" in (")
        col = condition[:idx].strip()
        vals_str = condition[idx+4:].strip().strip("()")
        vals = [v.strip().strip("'").strip('"') for v in vals_str.split(",")]
        if col not in df.columns:
            raise ValueError(f"过滤条件列不存在: {col}")
        return df[df[col].astype(str).isin(vals)]
    
    if "包含" in condition:
        idx = condition.index("包含")
        col = condition[:idx].strip()
        val = condition[idx+2:].strip().strip("'").strip('"')
        if col not in df.columns:
            raise ValueError(f"过滤条件列不存在: {col}")
        return df[df[col].astype(str).str.contains(val, na=False)]
    
    raise ValueError(f"无法解析过滤条件: {condition}")


def validate_pivot_config(tasks, config_dir):
    """
    校验透视分析配置，返回校验结果列表。
    
    Returns:
        list of dict: 每个元素是一个校验项，包含:
            - task_seq: 任务序号
            - level: error/warning/info
            - message: 校验信息
            - column: 问题所在列（可选）
    """
    results = []

    # 列名/字段名禁用的特殊字符（会破坏配置解析或值计算表达式）
    # - , 逗号：值字段/值映射/聚合方式/值计算的多元素分隔符
    # - = 等号：值计算的 表达式=结果列名 语法
    # - @ at：已废弃的标量引用前缀（保留禁用以防误用）
    # - {} 花括号：值映射的 {区块.列.行值|格式} 占位符
    # - \n \r 换行：值计算会把换行转成逗号
    _FORBIDDEN_NAME_CHARS = [",", "=", "@", "{", "}", "\n", "\r"]

    def _check_forbidden_chars(name_str, col_label, seq):
        """检查字段名列表是否含禁用字符，有则追加 error 到 results。"""
        if not name_str:
            return
        items = [s.strip() for s in str(name_str).split(",") if s.strip()]
        for item in items:
            bad_chars = [c for c in _FORBIDDEN_NAME_CHARS if c in item]
            if bad_chars:
                # 换行符显示为 \n 方便阅读
                display_chars = [repr(c)[1:-1] if c in "\n\r" else c for c in bad_chars]
                results.append({
                    "task_seq": seq,
                    "level": "error",
                    "message": f"字段名 '{item}' 含禁用字符 {display_chars}（会破坏配置解析或值计算）",
                    "column": col_label
                })

    if not tasks:
        results.append({
            "task_seq": "-",
            "level": "error",
            "message": "没有有效的透视分析配置为空",
            "column": ""
        })
        return results
    
    for task in tasks:
        seq = task.get("序号", "?")
        data_source = task.get("数据源", "")
        sheet_name = task.get("sheet", "Sheet1")
        行维度_str = task.get("行维度", "")
        列维度_str = task.get("列维度", "")
        值字段_str = task.get("值字段", "")
        聚合方式_str = task.get("聚合方式", "sum")
        是否计算 = task.get("是否计算", "是")
        
        # 跳过不计算的任务
        should_calc = str(是否计算).strip().lower()
        if should_calc in ("否", "no", "false", "0", "不计算", "跳过", "skip"):
            results.append({
                "task_seq": seq,
                "level": "info",
                "message": "已设置为不计算，将跳过",
                "column": "是否计算"
            })
            continue
        
        # 1. 检查必填项
        if not 值字段_str:
            results.append({
                "task_seq": seq,
                "level": "error",
                "message": "值字段不能为空",
                "column": "值字段"
            })

        # 1.1 检查字段名禁用字符（在读取数据前校验，避免数据文件不存在时漏检）
        _check_forbidden_chars(值字段_str, "值字段", seq)
        _check_forbidden_chars(行维度_str, "行维度", seq)
        _check_forbidden_chars(列维度_str, "列维度", seq)

        if not 行维度_str and not 列维度_str:
            results.append({
                "task_seq": seq,
                "level": "warning",
                "message": "行维度和列维度都为空，将生成单行汇总",
                "column": "行维度/列维度"
            })
        
        # 2. 检查聚合方式（用 _split_agg_funcs 拆分，保护 | 后参数内的逗号）
        agg_funcs = _split_agg_funcs(聚合方式_str)
        invalid_aggs = []
        for a in agg_funcs:
            a_no_fmt, _ = _parse_agg_format(a)  # 先剥离 |数字格式
            a_key, _ = _parse_join_sep(a_no_fmt)
            mapped = AGG_MAP.get(a_key, a_key)
            valid_aggs = {"sum", "mean", "count", "max", "min", "nunique", "pct", "count_pct", "join"}
            if mapped not in valid_aggs:
                invalid_aggs.append(a)
        if invalid_aggs:
            results.append({
                "task_seq": seq,
                "level": "error",
                "message": f"不支持的聚合方式: {invalid_aggs}",
                "column": "聚合方式"
            })
        
        # 3. 检查数据源文件是否存在
        if not data_source:
            results.append({
                "task_seq": seq,
                "level": "error",
                "message": "数据源不能为空",
                "column": "数据源"
            })
        else:
            # 检查是否是前序透视结果区块引用（{pivot}/{透视结果}/{prev}/{区块名}/{pivot:区块名}）
            # 区块引用的存在性在执行时由 block_results 决定，校验阶段跳过文件存在性检查
            ds_lower = str(data_source).strip().lower()
            is_pivot_ref = ds_lower in ("{pivot}", "pivot", "透视结果", "{prev}", "{上一个}", "{上一个结果}")
            # 识别 {区块名} / {pivot:区块名} 形式
            if not is_pivot_ref:
                ds_str = str(data_source).strip()
                if re.match(r"^\{.+\}$", ds_str):
                    is_pivot_ref = True
            if not is_pivot_ref:
                # 处理 JOIN 语法：逐表校验文件存在性
                # 修复 P0 问题4：JOIN 中左/右表为 {...} 区块引用时跳过文件检查
                ds_check = str(data_source).strip()
                join_parts = _parse_join_spec(ds_check)
                if join_parts:
                    # 收集 JOIN 中所有非区块引用的表名，逐个校验文件存在性
                    tables_to_check = []
                    # 第一段的左表
                    left_first = join_parts[0]["left"]
                    left_first_sheet = join_parts[0].get("left_sheet")
                    if not re.match(r"^\{.+\}$", left_first.strip()):
                        tables_to_check.append((left_first, left_first_sheet, "左表"))
                    # 每段的右表
                    for jp in join_parts:
                        right_name = jp["right"]
                        right_sheet = jp.get("right_sheet")
                        if not re.match(r"^\{.+\}$", right_name.strip()):
                            tables_to_check.append((right_name, right_sheet, "右表"))

                    from src.excel_reader import find_data_file, get_data_file_sheets
                    all_files_ok = True
                    for tbl_name, tbl_sheet, tbl_role in tables_to_check:
                        file_path = _resolve_data_path(tbl_name, config_dir)
                        if not file_path or not os.path.exists(file_path):
                            file_path = find_data_file(tbl_name, config_dir)
                        if not file_path or not os.path.exists(file_path):
                            results.append({
                                "task_seq": seq,
                                "level": "error",
                                "message": f"JOIN {tbl_role}文件不存在: {tbl_name}",
                                "column": "数据源"
                            })
                            all_files_ok = False
                        else:
                            # 校验 @Sheet 提示的 sheet 是否存在
                            if tbl_sheet:
                                sheets = get_data_file_sheets(file_path)
                                if sheets and tbl_sheet not in sheets:
                                    results.append({
                                        "task_seq": seq,
                                        "level": "error",
                                        "message": f"JOIN {tbl_role}Sheet '{tbl_sheet}' 不存在。可用Sheet: {sheets}",
                                        "column": "数据源"
                                    })
                                    all_files_ok = False
                    # JOIN 任务跳过列名校验（列可能来自右表，需执行时合并后才能判断）
                else:
                    # 非 JOIN、非区块引用：原有文件存在性检查
                    file_path = _resolve_data_path(ds_check, config_dir)
                    if not file_path or not os.path.exists(file_path):
                        from src.excel_reader import find_data_file
                        file_path = find_data_file(ds_check, config_dir)
                    if not file_path or not os.path.exists(file_path):
                        results.append({
                            "task_seq": seq,
                            "level": "error",
                            "message": f"数据源文件不存在: {data_source}",
                            "column": "数据源"
                        })
                    else:
                        # 4. 检查 Sheet 是否存在
                        from src.excel_reader import get_data_file_sheets
                        sheets = get_data_file_sheets(file_path)
                        if sheets and sheet_name not in sheets:
                            results.append({
                                "task_seq": seq,
                                "level": "error",
                                "message": f"Sheet '{sheet_name}' 不存在。可用Sheet: {sheets}",
                                "column": "Sheet"
                            })
                        else:
                            # 5. 检查列名是否存在（JOIN 任务跳过：列可能来自右表）
                            from src.excel_reader import read_data_file
                            try:
                                df = read_data_file(file_path, sheet_name)
                                if df is not None and not df.empty:
                                    all_cols = list(df.columns)

                                # 检查行维度
                                row_dims = [d.strip() for d in 行维度_str.split(",") if d.strip()]
                                missing_row = [d for d in row_dims if d not in all_cols]
                                if missing_row:
                                    results.append({
                                        "task_seq": seq,
                                        "level": "error",
                                        "message": f"行维度列不存在: {missing_row}",
                                        "column": "行维度"
                                    })

                                # 检查列维度
                                col_dims = [d.strip() for d in 列维度_str.split(",") if d.strip()]
                                missing_col = [d for d in col_dims if d not in all_cols]
                                if missing_col:
                                    results.append({
                                        "task_seq": seq,
                                        "level": "error",
                                        "message": f"列维度列不存在: {missing_col}",
                                        "column": "列维度"
                                    })

                                # 检查值字段
                                value_cols = [v.strip() for v in 值字段_str.split(",") if v.strip()]
                                missing_val = [v for v in value_cols if v not in all_cols]
                                if missing_val:
                                    results.append({
                                        "task_seq": seq,
                                        "level": "error",
                                        "message": f"值字段列不存在: {missing_val}",
                                        "column": "值字段"
                                    })
                                else:
                                    # 检查聚合方式与列数据类型是否兼容
                                    _check_agg_dtype_compat(
                                        results, seq, df, value_cols,
                                        _split_agg_funcs(聚合方式_str)
                                    )
                            except Exception as e:
                                results.append({
                                    "task_seq": seq,
                                    "level": "warning",
                                    "message": f"读取数据失败: {str(e)}",
                                    "column": "数据源"
                                })
        
        # 6. 值映射数量检查（对比实际输出列数，考虑单字段多聚合）
        val_map_str = task.get("值映射", "")
        val_maps = [m.strip() for m in val_map_str.split(",") if m.strip()]

        # 6.1 值映射禁用字符检查（值映射文本合法含 {} 占位符和 , 分隔符，只禁 = @ \n）
        _VALMAP_FORBIDDEN = ["=", "@", "\n", "\r"]
        for vm in val_maps:
            bad = [c for c in _VALMAP_FORBIDDEN if c in vm]
            if bad:
                display = [repr(c)[1:-1] if c in "\n\r" else c for c in bad]
                results.append({
                    "task_seq": seq,
                    "level": "error",
                    "message": f"值映射 '{vm}' 含禁用字符 {display}（花括号是占位符语法，= @ 会破坏表达式）",
                    "column": "值映射"
                })

        value_cols = [v.strip() for v in 值字段_str.split(",") if v.strip()]
        if val_maps:
            agg_funcs_val = _split_agg_funcs(聚合方式_str) if 聚合方式_str else []
            if len(value_cols) == 1 and len(agg_funcs_val) > 1:
                expected_cols = len(agg_funcs_val)
            else:
                expected_cols = len(value_cols)
            if len(val_maps) != expected_cols:
                results.append({
                    "task_seq": seq,
                    "level": "warning",
                    "message": f"值映射数量({len(val_maps)})与输出列数({expected_cols}, 来自{len(value_cols)}个值字段+{len(agg_funcs_val) if agg_funcs_val else 'sum'}聚合)不一致",
                    "column": "值映射"
                })

        # 7. 值计算结果列名校验（表达式本身合法含 =+-*/()，但 = 后的结果列名不能含禁用字符）
        val_calc_str = task.get("值计算", "")
        if val_calc_str:
            # 值计算把换行转逗号再按逗号拆分，模拟解析
            vc = val_calc_str.replace("\n", ",").replace("\r", ",")
            for c_item in [c.strip() for c in vc.split(",") if c.strip()]:
                if "=" in c_item:
                    result_name = c_item.split("=", 1)[1].strip()
                    bad = [ch for ch in [",", "@", "\n", "\r"] if ch in result_name]
                    if bad:
                        display = [repr(ch)[1:-1] if ch in "\n\r" else ch for ch in bad]
                        results.append({
                            "task_seq": seq,
                            "level": "error",
                            "message": f"值计算结果列名 '{result_name}' 含禁用字符 {display}（结果列名不能含 , @ 或换行）",
                            "column": "值计算"
                        })

    return results


# 仅这些聚合方式强制要求数值列；max/min 对字符串列合法（字典序），可用于取路径等场景
NUMERIC_AGGS = {"sum", "mean", "avg", "pct"}


def _check_agg_dtype_compat(results, task_seq, df, value_cols, agg_funcs):
    """检查聚合方式与列数据类型是否兼容。"""
    if not value_cols:
        return
    for vcol in value_cols:
        if vcol not in df.columns:
            continue
        af = _get_agg_for_col(vcol, agg_funcs, value_cols)
        af_no_fmt, _ = _parse_agg_format(af)  # 先剥离 |数字格式
        af_key, _ = _parse_join_sep(af_no_fmt)
        af_mapped = AGG_MAP.get(af_key, af_key)
        if af_mapped in NUMERIC_AGGS and not pd.api.types.is_numeric_dtype(df[vcol]):
            results.append({
                "task_seq": task_seq,
                "level": "error",
                "message": f"列「{vcol}」不是数值类型，不能使用 {af} 聚合（可用 count / nunique / count_pct）",
                "column": "聚合方式"
            })


def print_validation_results(results):
    """格式化打印校验结果"""
    if not results:
        print("[校验] 全部通过 ✓")
        return True
    
    error_count = sum(1 for r in results if r["level"] == "error")
    warning_count = sum(1 for r in results if r["level"] == "warning")
    info_count = sum(1 for r in results if r["level"] == "info")
    
    print(f"\n{'='*60}")
    print(f"  配置校验结果")
    print(f"{'='*60}")
    print(f"  错误: {error_count}  |  警告: {warning_count}  |  提示: {info_count}")
    print(f"{'='*60}")
    
    for r in results:
        icon = {"error": "✗", "warning": "⚠", "info": "ℹ"}.get(r["level"], " ")
        col_info = f" [{r['column']}]" if r.get("column") else ""
        print(f"  {icon} [任务{r['task_seq']}{col_info}: {r['message']}")
    
    print(f"{'='*60}")
    
    if error_count > 0:
        print(f"  ❌ 发现 {error_count} 个错误，请修正后再执行")
    elif warning_count > 0:
        print(f"  ⚠  发现 {warning_count} 个警告，可继续执行但可能影响结果")
    print()
    
    return error_count == 0


def run_analysis(task, config_dir, scalar_context=None, block_results=None):
    """执行单个透视任务。

    :param scalar_context: 无行维度任务产生的标量 {列名: 值}，供值计算引用
    :param block_results: 前面任务的结果 {区块名: DataFrame}，供值映射的 {区块.列.行值} 占位符引用
    """
    序号 = task.get("序号", "?")
    data_source = task.get("数据源", "")
    行维度_str = task.get("行维度", "")
    列维度_str = task.get("列维度", "")
    值字段_str = task.get("值字段", "")
    聚合方式_str = task.get("聚合方式", "sum")
    sheet_name = task.get("sheet", "Sheet1")

    if not 值字段_str:
        return None, f"[任务{序号}] 未指定值字段"

    行维度_str = 行维度_str or ""
    列维度_str = 列维度_str or ""
    值字段_str = 值字段_str or ""
    聚合方式_str = 聚合方式_str or "sum"

    行维度 = [d.strip() for d in 行维度_str.split(",") if d.strip()]
    列维度 = [d.strip() for d in 列维度_str.split(",") if d.strip()]
    值字段 = [v.strip() for v in 值字段_str.split(",") if v.strip()]
    # 解析每个聚合方式的 |数字格式 后缀，剥离格式后保留 join 的 |分隔符
    # number_formats 按值字段位置对应，供 excel_writer 应用自定义显示格式
    # 用 _split_agg_funcs 拆分多聚合，保护格式内的逗号（如 #,##0）不被切断
    number_formats = []
    聚合函数 = []
    for a in _split_agg_funcs(聚合方式_str):
        a_no_fmt, fmt = _parse_agg_format(a)
        聚合函数.append(a_no_fmt)
        number_formats.append(fmt)

    try:
        df = _load_joined_dataframe(config_dir, data_source, sheet_name, block_results=block_results)
    except ValueError as e:
        # JOIN 加载失败（右表缺失/列名不存在/键数量不一致等）显式报错
        return None, f"[任务{序号}] {str(e)}"
    if df is None or df.empty:
        return None, f"[任务{序号}] 数据为空或文件不存在: {data_source}"

    join_happened = _parse_join_spec(str(data_source)) is not None

    # 应用过滤条件
    filter_expr = task.get("过滤条件", "")
    if filter_expr:
        try:
            df = _apply_filter(df, filter_expr)
            if df.empty:
                return None, f"[任务{序号}] 过滤后数据为空，过滤条件: {filter_expr}"
        except Exception as e:
            return None, f"[任务{序号}] 过滤条件执行失败: {str(e)}"

    mapping_str = task.get("映射表", "")
    row_map_str = task.get("行映射", "")
    col_map_str = task.get("列映射", "")
    val_map_str = task.get("值映射", "")
    col_map, val_map = _parse_mapping(mapping_str, row_map_str, col_map_str, val_map_str, 行维度, 列维度, 值字段)
    if col_map or val_map:
        df = _apply_mapping(df, col_map, val_map)
        行维度 = [_translate_name(d, col_map) for d in 行维度]
        列维度 = [_translate_name(d, col_map) for d in 列维度]
        值字段 = [_translate_name(v, col_map) for v in 值字段]

    bin_spec = task.get("分箱", "")
    if bin_spec:
        行维度, 列维度, df = _apply_binning(bin_spec, 行维度, 列维度, df, col_map)

    missing_cols = []
    all_dims = 行维度 + 列维度
    for col in all_dims:
        if col not in df.columns:
            missing_cols.append(col)
    for col in 值字段:
        if col not in df.columns:
            missing_cols.append(col)

    # 智能列名映射：级联引用前序分组聚合结果时，用户可能写原字段名（如"5G用户数"），
    # 但实际列名带聚合后缀（如"5G用户数_求和"）。自动做前缀匹配映射，避免报错。
    # 仅当数据源是区块引用（{...}）时启用，避免对原始文件误映射。
    is_block_ref = bool(re.match(r"^\{.+\}$", str(data_source).strip())) or \
                   _resolve_block_reference(str(data_source), block_results) is not None
    if missing_cols and is_block_ref:
        available_cols = list(df.columns)
        col_map_smart = {}  # {用户写的列名: 实际列名}
        for mc in missing_cols[:]:
            mc_str = str(mc)
            # 精确匹配跳过
            if mc_str in available_cols:
                continue
            # 前缀匹配：可用列以 "用户列名_聚合后缀" 形式存在
            candidates = [str(c) for c in available_cols
                          if str(c).startswith(mc_str + "_")]
            if len(candidates) == 1:
                # 唯一匹配，自动映射
                col_map_smart[mc_str] = candidates[0]
                missing_cols.remove(mc)
            elif len(candidates) > 1:
                # 多个候选（如"5G用户数_求和"和"5G用户数_均值"），不自动映射，留给报错建议
                pass
        # 应用自动映射
        if col_map_smart:
            行维度 = [col_map_smart.get(d, d) for d in 行维度]
            列维度 = [col_map_smart.get(d, d) for d in 列维度]
            值字段 = [col_map_smart.get(v, v) for v in 值字段]
            # 同步更新 task 中的值，供后续 _cross_pivot/_group_aggregate 使用
            task["行维度"] = ",".join(行维度)
            task["列维度"] = ",".join(列维度)
            task["值字段"] = ",".join(值字段)

    if missing_cols:
        available = list(df.columns)
        # 增强报错：显示列总数和完整列名，方便排查（列多时截断显示前10+后5）
        if len(available) <= 15:
            avail_str = str(available)
        else:
            avail_str = f"[{available[:10]} ... {available[-5:]}] (共{len(available)}列)"
        # 智能建议：缺失列是某可用列前缀时（如写"5G用户数"，实际列"5G用户数_求和"）
        suggestions = []
        for mc in missing_cols:
            mc_str = str(mc)
            # 1. 前缀+后缀匹配（聚合后缀场景）
            candidates = [str(c) for c in available
                          if str(c).startswith(mc_str + "_") or str(c) == mc_str]
            if candidates and mc_str not in candidates:
                suggestions.append(f"'{mc_str}' → 可能是 {candidates[:3]}")
            else:
                # 2. 模糊包含匹配（值映射重命名场景：原字段名"销售额" → 值映射名"总销售额"）
                fuzzy = [str(c) for c in available if mc_str in str(c) and str(c) != mc_str]
                if fuzzy:
                    suggestions.append(f"'{mc_str}' → 可能是 {fuzzy[:3]}（前序任务可能用值映射重命名了列）")
        hint_parts = []
        if suggestions:
            hint_parts.append("；".join(suggestions))
        # 提示是否引用了交叉透视结果（列被展开）
        if any("_" in str(c) for c in missing_cols):
            hint_parts.append("若引用前序交叉透视结果，原列维度值可能已被展开为列名（如'地区'→'华东/华北/华南'）")
        # 区块引用场景的特殊提示
        if is_block_ref and not suggestions:
            hint_parts.append("前序任务可能通过「值映射」重命名了列，请检查前序任务的值映射配置，用映射后的列名引用")
        hint = "。提示：" + "；".join(hint_parts) if hint_parts else ""
        return None, f"[任务{序号}] 列不存在: {missing_cols}。可用列: {avail_str}{hint}"

    for col in 值字段:
        af_for_col = _get_agg_for_col(col, 聚合函数, 值字段)
        af_key, _ = _parse_join_sep(af_for_col)
        af_mapped = AGG_MAP.get(af_key, af_key)
        # max/min 对字符串列合法（字典序），不强制转数值（如图片路径取值场景）
        # join 专门处理字符串列去重拼接，也不强制转数值
        if af_mapped not in ("count", "nunique", "pct", "count_pct", "max", "min", "join"):
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.dropna(subset=值字段, how="all")
    if df.empty:
        return None, f"[任务{序号}] 值字段全为空"

    if 列维度 and 列维度[0]:
        result = _cross_pivot(df, 行维度, 列维度, 值字段, 聚合函数, task)
    else:
        result = _group_aggregate(df, 行维度, 值字段, 聚合函数, task)

    # 聚合后应用列名映射
    raw_val_maps = [m.strip() for m in val_map_str.split(",") if m.strip()] if val_map_str else []
    
    if result and raw_val_maps:
        map_idx = 0
        for key, df in result.items():
            if isinstance(df, pd.DataFrame):
                rename = {}
                for c in df.columns:
                    if c in 行维度:
                        continue
                    if map_idx < len(raw_val_maps):
                        # 值映射支持 {区块.列.行值|格式} 占位符，用前面任务结果替换
                        new_name = _resolve_value_map_placeholder(raw_val_maps[map_idx], block_results)
                        rename[c] = new_name
                        map_idx += 1
                    else:
                        break

                if rename:
                    df.rename(columns=rename, inplace=True)

    val_calc = task.get("值计算", "")
    if val_calc and result:
        result = _apply_value_calc(result, val_calc, 值字段, 聚合函数, scalar_context, raw_val_maps, 值字段)

    if 行维度:
        task["行维度"] = ",".join(行维度)
    if 列维度:
        task["列维度"] = ",".join(列维度)

    # 记录百分比列（pct/count_pct 产生的列），供 excel_writer 精确识别，避免靠列名猜测
    # pct/count_pct 结果的数值列均为 0~1 小数，所有非分组数值列都属于百分比列
    if any(af in ("pct", "count_pct") for af in 聚合函数):
        pct_cols = _collect_pct_columns(result, 行维度)
        if pct_cols:
            task["_pct_columns"] = pct_cols

    # 记录自定义数字格式（聚合方式 ::格式 语法），供 excel_writer 应用
    # 优先级：自定义格式 > pct 默认 > 0.00 默认
    if any(number_formats):
        col_fmts = _collect_number_formats(result, 行维度, number_formats)
        if col_fmts:
            task["_number_formats"] = col_fmts

    if join_happened and not df.empty:
        结果Sheet = task.get("结果Sheet", f"结果{序号}")
        result[f"_JOIN中间表_{结果Sheet}"] = df.copy()

    return result, None


def _collect_pct_columns(result, row_dims):
    """收集 pct/count_pct 产生的最终列名（值映射/值计算后的列名）。
    规则：结果 DataFrame 中所有非分组维度列、非合计列、非指标列的数值列。
    """
    pct_cols = set()
    if not isinstance(result, dict):
        return []
    for key, df in result.items():
        if not isinstance(df, pd.DataFrame):
            continue
        skip = set(row_dims) | {"合计", "总计", "指标", "值"}
        for col in df.columns:
            if col in skip:
                continue
            # 交叉表的列可能是数值型列名（如产品A的销量），需排除分组维度
            pct_cols.add(str(col))
    return list(pct_cols)


def _collect_number_formats(result, row_dims, number_formats):
    """收集每个数值列对应的自定义数字格式，按值映射后的列名索引。

    number_formats 按值字段位置对应；若结果列数多于值字段数（如交叉表展开），
    超出部分用最后一个有效格式，保证交叉表所有数值列格式一致。

    :param result: run_analysis 的结果
    :param row_dims: 行维度列表
    :param number_formats: 每个值字段对应的数字格式列表（None 表示用默认）
    :return: {列名: 数字格式}
    """
    fmt_map = {}
    if not number_formats or not isinstance(result, dict):
        return fmt_map
    for key, df in result.items():
        if not isinstance(df, pd.DataFrame):
            continue
        skip = set(row_dims) | {"合计", "总计", "指标", "值"}
        map_idx = 0
        for col in df.columns:
            if col in skip:
                continue
            # 按值字段位置对应；交叉表展开时超出位置用最后一个有效格式
            if map_idx < len(number_formats):
                fmt = number_formats[map_idx]
            else:
                fmt = None
                for f in reversed(number_formats):
                    if f:
                        fmt = f
                        break
            if fmt:
                fmt_map[str(col)] = fmt
            map_idx += 1
    return fmt_map


def _cross_pivot(df, row_dims, col_dims, value_cols, agg_funcs, task):
    results = {}
    for vcol in value_cols:
        for af in agg_funcs:
            af_key, join_sep = _parse_join_sep(af)
            a_mapped = AGG_MAP.get(af_key, af_key)
            is_sum_pct = a_mapped == "pct"
            is_count_pct = a_mapped == "count_pct"
            is_join = a_mapped == "join"

            if is_join:
                # join 聚合：自定义 aggfunc，空值填充空字符串
                actual_af = lambda s, _sep=join_sep or _JOIN_DEFAULT_SEP: _join_unique(s, _sep)
                fill_val = ""
            else:
                actual_af = "sum" if is_sum_pct else ("count" if is_count_pct else af_key)
                fill_val = 0

            pivot = pd.pivot_table(
                df,
                values=vcol,
                index=row_dims,
                columns=col_dims,
                aggfunc=actual_af,
                fill_value=fill_val,
            )

            if is_sum_pct or is_count_pct:
                grand = pivot.values.sum()
                if grand != 0:
                    pivot = pivot / grand
                    pivot = pivot.map(lambda x: round(x, 4))
                else:
                    pivot = pivot.map(lambda x: 0.0)

            pivot = pivot.reset_index()

            dim_cols = [c for c in pivot.columns if c not in pivot.select_dtypes(include=np.number).columns]
            if not dim_cols:
                dim_cols = [pivot.columns[0]]
            numeric_cols = [c for c in pivot.columns if c in pivot.select_dtypes(include=np.number).columns]

            if is_join:
                # join 结果是字符串，不计算合计行/列，不做数值转换
                agg_label = "拼接"
                if _is_display_name(vcol):
                    key = vcol
                elif len(value_cols) * len(agg_funcs) > 1:
                    key = f"{vcol}_{agg_label}"
                else:
                    key = vcol
                results[key] = pivot
                continue

            col_sums = {}
            for nc in numeric_cols:
                col_sums[nc] = pivot[nc].sum()

            total_row_data = {}
            for c in pivot.columns:
                if c in col_sums:
                    total_row_data[c] = col_sums[c]
                elif c == dim_cols[0]:
                    total_row_data[c] = "合计"
                else:
                    total_row_data[c] = ""
            pivot = pd.concat([pivot, pd.DataFrame([total_row_data])], ignore_index=True)

            for c in pivot.columns:
                if c not in dim_cols and c != "合计":
                    try:
                        pivot[c] = pd.to_numeric(pivot[c], errors="coerce")
                    except Exception:
                        pass

            row_sums = pd.Series(0.0, index=pivot.index)
            for nc in numeric_cols:
                row_sums = row_sums + pd.to_numeric(pivot[nc], errors="coerce").fillna(0)
            if "合计" in pivot.columns:
                row_sums = row_sums + pd.to_numeric(pivot["合计"], errors="coerce").fillna(0)
            pivot["合计"] = row_sums

            agg_label = {"sum": "求和", "mean": "均值", "avg": "均值", "count": "计数", "max": "最大值", "min": "最小值", "nunique": "去重计数", "pct": "占比", "join": "拼接"}.get(af_key, af_key)
            if _is_display_name(vcol):
                key = vcol
            elif len(value_cols) * len(agg_funcs) > 1:
                key = f"{vcol}_{agg_label}"
            else:
                key = vcol
            results[key] = pivot

    return results


def _group_aggregate(df, group_cols, value_cols, agg_funcs, task):
    if not group_cols:
        row_data = {}
        # 按位置 1:1 对应（与有行维度一致）；单值多聚合时全组合
        if len(value_cols) == 1 and len(agg_funcs) > 1:
            vcol = value_cols[0]
            for af in agg_funcs:
                af_key, join_sep = _parse_join_sep(af)
                a_mapped = AGG_MAP.get(af_key, af_key)
                if a_mapped == "join":
                    val = _join_unique(df[vcol], join_sep or _JOIN_DEFAULT_SEP)
                elif a_mapped in ("pct", "count_pct"):
                    val = df[vcol].agg("sum" if a_mapped == "pct" else "count")
                    total = float(df[vcol].agg("sum" if a_mapped == "pct" else "count"))
                    val = round(float(val) / total, 4) if total != 0 else 0.0
                elif a_mapped == "nunique":
                    val = df[vcol].nunique()
                else:
                    val = df[vcol].agg(a_mapped)
                agg_label = _get_agg_label(af_key)
                key = f"{vcol}_{agg_label}"
                row_data[key] = round(float(val), 4) if isinstance(val, (int, float)) else val
        else:
            for idx, vcol in enumerate(value_cols):
                if idx < len(agg_funcs):
                    af = agg_funcs[idx]
                else:
                    af = "sum"
                af_key, join_sep = _parse_join_sep(af)
                a_mapped = AGG_MAP.get(af_key, af_key)
                if a_mapped == "join":
                    val = _join_unique(df[vcol], join_sep or _JOIN_DEFAULT_SEP)
                elif a_mapped in ("pct", "count_pct"):
                    val = df[vcol].agg("sum" if a_mapped == "pct" else "count")
                    total = float(df[vcol].agg("sum" if a_mapped == "pct" else "count"))
                    val = round(float(val) / total, 4) if total != 0 else 0.0
                elif a_mapped == "nunique":
                    val = df[vcol].nunique()
                else:
                    val = df[vcol].agg(a_mapped)
                agg_label = _get_agg_label(af_key)
                key = f"{vcol}_{agg_label}"
                row_data[key] = round(float(val), 4) if isinstance(val, (int, float)) else val
        result_df = pd.DataFrame([row_data])
        return {"结果": result_df}

    agg_dict = {}
    has_sum_pct = False
    has_count_pct = False
    
    agg_dict = {}
    field_funcs = {}
    pct_tmp_cols = {}
    count_pct_tmp_cols = {}
    tmp_col_counter = 0
    join_specs = []  # [(vcol, sep, agg_label), ...] join 聚合单独处理

    for idx, vcol in enumerate(value_cols):
        if len(value_cols) == 1 and len(agg_funcs) > 1:
            funcs_for_vcol = agg_funcs
        elif idx < len(agg_funcs):
            funcs_for_vcol = [agg_funcs[idx]]
        else:
            funcs_for_vcol = ["sum"]

        for a in funcs_for_vcol:
            a_key, join_sep = _parse_join_sep(a)
            a_mapped = AGG_MAP.get(a_key, a_key)
            if a_mapped == "join":
                # join 聚合单独记录，不走标准 agg_dict（需要分隔符且结果为字符串）
                join_specs.append((vcol, join_sep or _JOIN_DEFAULT_SEP, _get_agg_label(a_key)))
            elif a_mapped == "pct":
                tmp_col = f"__pct_{tmp_col_counter}__"
                tmp_col_counter += 1
                pct_tmp_cols[tmp_col] = vcol
                field_funcs[tmp_col] = ["sum"]
                has_sum_pct = True
            elif a_mapped == "count_pct":
                tmp_col = f"__cntpct_{tmp_col_counter}__"
                tmp_col_counter += 1
                count_pct_tmp_cols[tmp_col] = vcol
                field_funcs[tmp_col] = ["count"]
                has_count_pct = True
            else:
                if vcol not in field_funcs:
                    field_funcs[vcol] = []
                field_funcs[vcol].append(a_mapped)
    
    if has_sum_pct or has_count_pct:
        df = df.copy()
        for tmp_col, orig_col in pct_tmp_cols.items():
            df[tmp_col] = df[orig_col].astype(float)
        # count_pct 只做行数计数，与数据类型无关，直接复制原始列（支持字符串列）
        for tmp_col, orig_col in count_pct_tmp_cols.items():
            df[tmp_col] = df[orig_col]
    
    for vcol, funcs in field_funcs.items():
        agg_dict[vcol] = funcs

    if agg_dict:
        grouped = df.groupby(group_cols, as_index=False, observed=True).agg(agg_dict)
    else:
        # 所有聚合都是 join，无标准 agg；只获取分组键
        grouped = df[group_cols].drop_duplicates().reset_index(drop=True)

    if has_sum_pct or has_count_pct:
        orig_tuples = [col for col in grouped.columns.values]
        for i, col in enumerate(orig_tuples):
            if isinstance(col, tuple) and len(col) >= 2:
                field_name = str(col[0]).strip()
                agg_part = str(col[1]).strip()
                
                if field_name in pct_tmp_cols and agg_part == 'sum':
                    total = grouped[col].sum()
                    grouped[col] = grouped[col].astype(float)
                    if total != 0:
                        grouped[col] = (grouped[col] / total).round(4)
                    else:
                        grouped[col] = 0.0
                
                elif field_name in count_pct_tmp_cols and agg_part == 'count':
                    total = grouped[col].sum()
                    grouped[col] = grouped[col].astype(float)
                    if total != 0:
                        grouped[col] = (grouped[col] / total).round(4)
                    else:
                        grouped[col] = 0.0
                
                elif field_name not in pct_tmp_cols and field_name not in count_pct_tmp_cols:
                    if pd.api.types.is_numeric_dtype(grouped[col]):
                        grouped[col] = grouped[col].round(_col_round_precision(field_name))
            elif col not in group_cols and pd.api.types.is_numeric_dtype(grouped[col]):
                grouped[col] = grouped[col].round(_col_round_precision(str(col)))
        
        new_cols = list(grouped.columns)
        for i, col in enumerate(new_cols):
            if isinstance(col, tuple) and len(col) >= 2:
                field_name = str(col[0]).strip()
                if field_name in pct_tmp_cols:
                    orig = pct_tmp_cols[field_name]
                    new_cols[i] = (orig, 'pct')
                elif field_name in count_pct_tmp_cols:
                    orig = count_pct_tmp_cols[field_name]
                    new_cols[i] = (orig, 'count_pct')
        for i in range(len(new_cols)):
            if not isinstance(new_cols[i], tuple):
                new_cols[i] = (new_cols[i], '')
        grouped.columns = pd.MultiIndex.from_tuples(new_cols)
    else:
        for col in grouped.columns:
            if col not in group_cols and pd.api.types.is_numeric_dtype(grouped[col]):
                grouped[col] = grouped[col].round(_col_round_precision(str(col)))

    # 检查是否有值映射配置
    val_map_str = task.get("值映射", "") if task else ""
    raw_val_maps = [m.strip() for m in val_map_str.split(",") if m.strip()] if val_map_str else []
    has_val_map = bool(raw_val_maps)
    
    # 如果有值映射，使用原始 key 作为列名（不加聚合后缀），后续步骤会应用映射
    # 如果没有值映射，使用 _flatten_col 添加聚合后缀
    orig_tuples = [col for col in grouped.columns.values]
    if has_val_map:
        # 有值映射时，直接在 _group_aggregate 中应用映射，不走后续重命名逻辑
        # 值映射按顺序对应值列（跳过分组列）
        new_columns = []
        val_map_idx = 0
        for col in orig_tuples:
            # 先判断是否是分组列（分组列的 tuple 第二个元素通常是空字符串）
            if isinstance(col, tuple):
                # 检查是否是分组列
                if len(col) >= 2 and str(col[1]).strip() == '':
                    # 这是分组列，保留原名
                    new_columns.append(str(col[0]).strip())
                elif val_map_idx < len(raw_val_maps):
                    # 值列，使用值映射
                    new_columns.append(raw_val_maps[val_map_idx])
                    val_map_idx += 1
                else:
                    new_columns.append(_agg_key_to_name(col))
            elif col in group_cols:
                new_columns.append(col)
            else:
                new_columns.append(col)
        grouped.columns = new_columns
    else:
        grouped.columns = [_flatten_col(col) for col in grouped.columns.values]

    # 处理重复列名
    seen = {}
    for i, name in enumerate(grouped.columns):
        if name in seen and name not in group_cols:
            prev = seen[name]
            curr_tuple = orig_tuples[i] if i < len(orig_tuples) else None
            prev_tuple = orig_tuples[prev] if prev < len(orig_tuples) else None
            if isinstance(curr_tuple, tuple) and len(curr_tuple) >= 2:
                agg_part = str(curr_tuple[1]).strip()
                if agg_part:
                    grouped.columns.values[i] = f"{name}_{agg_part}"
                    continue
            if isinstance(prev_tuple, tuple) and len(prev_tuple) >= 2:
                agg_part = str(prev_tuple[1]).strip()
                if agg_part:
                    grouped.columns.values[prev] = f"{name}_{agg_part}"
        else:
            seen[name] = i

    # join 聚合单独计算（在列名展平后添加，保持首次出现顺序去重拼接）
    if join_specs:
        for vcol, sep, agg_label in join_specs:
            # 确定 join 列名
            if has_val_map and val_map_idx < len(raw_val_maps):
                col_name = raw_val_maps[val_map_idx]
                val_map_idx += 1
            else:
                col_name = f"{vcol}_{agg_label}"

            # 用 groupby + apply 计算，转成 dict 按分组键映射，确保与 grouped 行对齐
            join_series = df.groupby(group_cols, observed=True)[vcol].apply(
                lambda s, _sep=sep: _join_unique(s, _sep)
            )
            join_map = {k: v for k, v in join_series.items()}

            if len(group_cols) == 1:
                grouped[col_name] = grouped[group_cols[0]].map(join_map)
            else:
                grouped[col_name] = grouped.apply(
                    lambda row: join_map.get(tuple(row[c] for c in group_cols), ""), axis=1
                )

    grouped = grouped.sort_values(group_cols)

    return {"结果": grouped}


def _agg_key_to_name(col):
    """将聚合后的列名 tuple 转为字符串，保留原始字段名（不加聚合后缀）"""
    if isinstance(col, tuple):
        # 取第一个元素（字段名），可能带序号后缀
        parts = [str(p).strip() for p in col if str(p).strip()]
        if parts:
            return parts[0]
        return str(col)
    return str(col).strip()


def _flatten_col(col):
    if isinstance(col, tuple):
        parts = [str(p).strip() for p in col if str(p).strip()]
        if len(parts) >= 2:
            field_name = parts[0]
            agg_func = parts[1] if len(parts) > 1 else ""
            
            # 还原原始字段名（去掉我们添加的 _N 后缀）
            import re
            m = re.match(r'^(.+)_(\d+)$', field_name)
            if m:
                # field_name 是 "销售额_1" 这样的格式，还原为 "销售额"
                field_name = m.group(1)
                suffix = m.group(2)  # 保留序号
            else:
                suffix = None
            
            # 如果聚合函数是 sum/mean 等英文，转为中文
            agg_cn = {"sum": "求和", "mean": "均值", "avg": "均值", "count": "计数",
                      "max": "最大值", "min": "最小值", "nunique": "去重",
                      "pct": "占比", "count_pct": "计数占比"}.get(agg_func, agg_func)
            
            if suffix is not None:
                return f"{field_name}_{agg_cn}_{suffix}"
            return f"{field_name}_{agg_cn}"
        return "_".join(parts) if parts else str(col)
    return str(col).strip()


def _is_display_name(name):
    return any(c in name for c in "()（）") or any("\u4e00" <= c <= "\u9fff" for c in name)


def _get_agg_for_col(col, agg_funcs, value_cols):
    idx = value_cols.index(col) if col in value_cols else 0
    if idx < len(agg_funcs):
        return agg_funcs[idx]
    return agg_funcs[0] if agg_funcs else "sum"


def _get_agg_label(af):
    return {
        "sum": "求和", "mean": "均值", "avg": "均值", "count": "计数",
        "max": "最大值", "min": "最小值", "nunique": "去重计数",
        "pct": "占比", "count_pct": "计数占比",
        "join": "拼接", "拼接": "拼接", "去重拼接": "拼接", "group_concat": "拼接",
    }.get(af, af)


def _align_funcs_to_fields(value_cols, agg_funcs):
    if len(value_cols) == len(agg_funcs):
        return {vc: [agg_funcs[i]] for i, vc in enumerate(value_cols)}
    return {}


def _auto_find_data_files(config_dir, config_path):
    import glob
    config_name = os.path.basename(config_path).lower()
    xlsx_files = glob.glob(os.path.join(config_dir, "*.xlsx"))
    candidates = []
    for f in xlsx_files:
        name = os.path.basename(f)
        if name.startswith("~$"):
            continue
        if name.lower() == config_name:
            continue
        if "配置" in name or "config" in name.lower():
            continue
        candidates.append(f)
    return candidates


# 公开别名，供 main.py 等外部模块调用（避免导入下划线私有成员）
def find_data_files(config_dir, config_path):
    """查找配置目录下的候选数据文件（公开 API）。"""
    return _auto_find_data_files(config_dir, config_path)


def _parse_mapping(mapping_str, row_map_str, col_map_str, val_map_str, 行维度, 列维度, 值字段):
    parts_legacy = [p.strip() for p in mapping_str.split(",") if p.strip()] if mapping_str and mapping_str.strip() else []
    if parts_legacy and any("=" in p for p in parts_legacy):
        return _parse_old_new_mapping(parts_legacy)

    col_map = {}
    if row_map_str or col_map_str:
        col_map = _map_fields(row_map_str, 行维度, col_map)
        col_map = _map_fields(col_map_str, 列维度, col_map)
    elif parts_legacy:
        all_fields = 行维度 + 列维度 + 值字段
        for i, new_name in enumerate(parts_legacy):
            if i < len(all_fields) and new_name:
                col_map[all_fields[i]] = new_name

    return col_map, {}


def _map_fields(map_str, fields, col_map):
    if not map_str or not map_str.strip():
        return col_map
    parts = [p.strip() for p in map_str.split(",") if p.strip()]
    for i, part in enumerate(parts):
        if not part:
            continue
        if "=" in part:
            src, dst = part.split("=", 1)
            src = src.strip()
            dst = dst.strip()
            if src and dst:
                col_map[src] = dst
        elif i < len(fields) and part:
            col_map[fields[i]] = part
    return col_map


def _parse_old_new_mapping(parts):
    col_map = {}
    val_map = {}
    for pair in parts:
        if "=" not in pair:
            continue
        src, dst = pair.split("=", 1)
        src = src.strip()
        dst = dst.strip()
        if not src or not dst:
            continue
        if " " not in dst and not any(c in dst for c in "():（）"):
            col_map[src] = dst
        val_map[src] = dst
    return col_map, val_map


def _translate_name(name, col_map):
    if name in col_map:
        return col_map[name]
    return name


def _apply_mapping(df, col_map, val_map):
    rename = {}
    for src, dst in col_map.items():
        if src in df.columns:
            rename[src] = dst
        else:
            # 处理聚合后的列名（如 "销售额_求和"，需要映射到 "新名称_求和"）
            for c in df.columns:
                # 精确匹配原始列名
                if c == src:
                    rename[c] = dst
                    break
                # 匹配 "原列名_聚合后缀" 格式（如 "销售额_求和" -> "新名称_求和"）
                if c.startswith(src + "_"):
                    suffix = c[len(src):]
                    rename[c] = dst + suffix
                    break
    df = df.rename(columns=rename)
    if val_map:
        for col in df.columns:
            if pd.api.types.is_string_dtype(df[col]) or pd.api.types.is_object_dtype(df[col]):
                df[col] = df[col].astype(str).map(lambda x: val_map.get(x, x))
    return df


def _apply_binning(bin_spec, row_dims, col_dims, df, col_map):
    spec = bin_spec.strip()
    if "=" not in spec:
        return row_dims, col_dims, df
    
    custom_name = None
    if "|" in spec:
        spec, custom_name = spec.split("|", 1)
        custom_name = custom_name.strip()
    
    col_name, params_str = spec.split("=", 1)
    col_name = col_name.strip()
    params_str = params_str.strip()
    if col_name in col_map:
        col_name = col_map[col_name]
    if col_name not in df.columns:
        return row_dims, col_dims, df
    parts = [p.strip() for p in params_str.split(",")]
    if len(parts) < 3:
        return row_dims, col_dims, df

    try:
        if len(parts) == 3:
            start = float(parts[0])
            end = float(parts[1])
            bins = int(parts[2])
            if bins < 1:
                return row_dims, col_dims, df
            bin_edges = np.linspace(start, end, bins + 1)
        else:
            bin_edges = np.array([float(p) for p in parts])
            if len(np.unique(bin_edges)) < 2:
                return row_dims, col_dims, df
            bin_edges = np.sort(np.unique(bin_edges))
    except ValueError:
        return row_dims, col_dims, df

    labels = []
    for i in range(len(bin_edges) - 1):
        labels.append(f"{_fmt_num(bin_edges[i])}~{_fmt_num(bin_edges[i+1])}")
    
    if custom_name:
        binned_col = custom_name
    else:
        binned_col = col_name
    df[binned_col] = pd.cut(df[col_name], bins=bin_edges, labels=labels, include_lowest=True)
    if binned_col != col_name and col_name in df.columns:
        df.drop(col_name, axis=1, inplace=True)
    new_row_dims = [binned_col if d == col_name else d for d in row_dims]
    new_col_dims = [binned_col if d == col_name else d for d in col_dims]
    return new_row_dims, new_col_dims, df


def _fmt_num(val):
    if val == int(val):
        return str(int(val))
    return f"{val:.1f}"


def _resolve_scalar_or_number(text, scalar_context=None):
    """解析值为数字或标量。先尝试 float，失败则查 scalar_context。
    返回数值或 None（无法解析）。
    """
    try:
        return float(text.strip())
    except (ValueError, TypeError):
        pass
    scalar_context = scalar_context or {}
    key = text.strip()
    if key in scalar_context:
        return float(scalar_context[key])
    return None


def _mask_token(working_expr, token):
    """将 working_expr[0] 中 token 替换为等长空格，防止后续短token子串误匹配"""
    working_expr[0] = working_expr[0].replace(token, " " * len(token))


def _apply_value_calc(result, val_calc, value_cols, agg_funcs, scalar_context=None, raw_val_maps=None, orig_value_cols=None):
    """
    值计算：支持单列与常数运算，以及多列组合运算。
    
    语法格式：
    1. 单列与常数：*100, /1000, +10, -5
    2. 多列组合：销售额/销量, 利润/销售额*100
    3. 指定结果列名：销售额/销量=单价
    4. 引用历史标量：销售额/总销售额（总销售额来自之前的无行维度任务）
    
    示例：
    - 值计算="销售额,销量"
    - 表达式列："销售额/销量=单价"
    """
    scalar_context = scalar_context or {}
    raw_val_maps = raw_val_maps or []
    orig_value_cols = orig_value_cols or list(value_cols)
    # 防御性处理：将换行符统一为逗号（Excel Alt+Enter 产生的换行会导致多个表达式被合并为一个）
    val_calc = val_calc.replace("\n", ",").replace("\r", ",")
    calcs = [c.strip() for c in val_calc.split(",") if c.strip()]
    calc_map = {}
    for i, expr in enumerate(calcs):
        if not expr:
            continue
        if i < len(value_cols):
            calc_map[value_cols[i]] = expr

    for key, df in result.items():
        if not isinstance(df, pd.DataFrame):
            continue
        
        new_cols = {}
        processed_expressions = set()
        
        for col in df.columns:
            if not pd.api.types.is_numeric_dtype(df[col]):
                continue
            expr = _find_matching_calc(col, calc_map, value_cols, raw_val_maps)
            if expr:
                expr_key = expr.split("=")[0].strip()
                if expr_key in processed_expressions:
                    continue
                processed_expressions.add(expr_key)
                result_col_name = None
                if "=" in expr:
                    expr_part, result_col_name = expr.split("=", 1)
                    expr = expr_part.strip()
                    result_col_name = result_col_name.strip()
                
                # 判断走多列分支还是单列分支：
                # - 单列分支：表达式以运算符开头（如 *100, /标量, +10），对当前列做单值运算
                # - 多列分支：表达式是完整算式（如 销售额/销量, 总销售额/总销量），用 eval 计算
                #   （标量自动识别：表达式中的标识符若不在列名中但在 scalar_context 中，自动作为标量）
                is_single_col_op = bool(re.match(r"^[+\-*/]", expr))
                has_multi_col = not is_single_col_op

                try:
                    if has_multi_col:
                        calc_df = df.copy()

                        # 构建列名映射：从已知列名/值字段/标量中查找表达式内出现的完整标识符
                        # 不使用正则token拆分，避免括号等特殊字符破坏列名完整性
                        # 使用渐进式占位替换防止子串误匹配（如 销售额 误匹配 总销售额 的子串）
                        # 标量自动识别：candidates 包含 scalar_context.keys()，匹配后映射为 __SCALAR__
                        col_mapping = {}
                        unmatched_tokens = []

                        candidates = list(calc_df.columns)
                        candidates.extend(value_cols)
                        if raw_val_maps:
                            candidates.extend(raw_val_maps)
                        if scalar_context:
                            candidates.extend(scalar_context.keys())
                        candidates = sorted(set(str(c) for c in candidates), key=len, reverse=True)

                        working_expr = [expr]  # 用列表包装实现可变引用
                        for token in candidates:
                            if token not in working_expr[0]:
                                continue
                            if token in calc_df.columns:
                                col_mapping[token] = token
                            elif raw_val_maps and token in raw_val_maps and token in calc_df.columns:
                                col_mapping[token] = token
                            elif token in value_cols:
                                if raw_val_maps:
                                    idx = value_cols.index(token)
                                    if idx < len(raw_val_maps):
                                        mapped = raw_val_maps[idx]
                                        if mapped in calc_df.columns:
                                            col_mapping[token] = mapped
                                            _mask_token(working_expr, token)
                                            continue
                                for c in calc_df.columns:
                                    if c.startswith(token + "_"):
                                        col_mapping[token] = c
                                        break
                                if token not in col_mapping:
                                    unmatched_tokens.append(token)
                                _mask_token(working_expr, token)
                                continue
                            else:
                                for c in calc_df.columns:
                                    if c.startswith(token + "_"):
                                        col_mapping[token] = c
                                        break
                                if token not in col_mapping and scalar_context and token in scalar_context:
                                    col_mapping[token] = "__SCALAR__"
                            # 将已匹配token在working_expr中屏蔽，防止子串误匹配
                            _mask_token(working_expr, token)

                        if not col_mapping or unmatched_tokens:
                            print(f"    [警告] 值计算表达式'{expr}'解析失败:" if unmatched_tokens
                                  else f"    [警告] 值计算表达式'{expr}'无法匹配任何列名")
                            if unmatched_tokens:
                                print(f"           未识别标识符: {unmatched_tokens}")
                            print(f"           可用列名: {list(calc_df.columns)}")
                            print(f"           值字段原始名: {value_cols}")
                            if raw_val_maps:
                                print(f"           值映射: {raw_val_maps}")
                            if scalar_context:
                                print(f"           可用标量: {list(scalar_context.keys())}")
                            if not unmatched_tokens and not col_mapping:
                                continue
                            elif unmatched_tokens:
                                # 部分匹配失败也跳过，避免表达式不完整
                                continue

                        # 使用 AST 白名单一次性计算 Series，避免逐行 iterrows/eval。
                        safe_expr = expr
                        namespace = {}
                        for i, (orig_name, actual_col) in enumerate(col_mapping.items()):
                            temp_name = f"_C{i}_"
                            safe_expr = safe_expr.replace(orig_name, temp_name)
                            if actual_col == "__SCALAR__":
                                namespace[temp_name] = float(scalar_context[orig_name])
                            else:
                                namespace[temp_name] = pd.to_numeric(calc_df[actual_col], errors="coerce")
                        try:
                            calc_result = evaluate_numeric_expression(safe_expr, namespace)
                            if not isinstance(calc_result, pd.Series):
                                calc_result = pd.Series(calc_result, index=calc_df.index)
                        except Exception as e:
                            print(f"    [警告] 值计算表达式'{expr}'执行失败: {e}")
                            print(f"           替换后表达式: {safe_expr}")
                            print(f"           列名映射: {col_mapping}")
                            continue

                        if calc_result is None:
                            continue

                        if result_col_name:
                            new_col_name = result_col_name
                        else:
                            # 自动生成列名（运算符替换为中文使列名更整洁）
                            new_col_name = expr.replace("/", "_除_").replace("*", "_乘_").replace("+", "_加_").replace("-", "_减_")
                            new_col_name = new_col_name.replace("(", "").replace(")", "").replace("%", "")

                        new_cols[new_col_name] = calc_result
                    else:
                        # 单列与常数运算（支持引用标量，标量名直接写在运算符后，如 *总销售额）
                        operand_expr = expr
                        if operand_expr.startswith("*"):
                            factor = _resolve_scalar_or_number(operand_expr[1:], scalar_context)
                            if factor is None:
                                print(f"    [警告] 值计算无法解析: {expr}")
                                continue
                            df[col] = df[col] * factor
                        elif operand_expr.startswith("/"):
                            divisor = _resolve_scalar_or_number(operand_expr[1:], scalar_context)
                            if divisor is None:
                                print(f"    [警告] 值计算无法解析: {expr}")
                                continue
                            if divisor == 0:
                                print(f"    [警告] 值计算除数为0，跳过: {expr}")
                                continue
                            df[col] = df[col] / divisor
                        elif operand_expr.startswith("+"):
                            addend = _resolve_scalar_or_number(operand_expr[1:], scalar_context)
                            if addend is None:
                                print(f"    [警告] 值计算无法解析: {expr}")
                                continue
                            df[col] = df[col] + addend
                        elif operand_expr.startswith("-"):
                            subtrahend = _resolve_scalar_or_number(operand_expr[1:], scalar_context)
                            if subtrahend is None:
                                print(f"    [警告] 值计算无法解析: {expr}")
                                continue
                            df[col] = df[col] - subtrahend
                        elif _resolve_scalar_or_number(operand_expr, scalar_context) is not None:
                            df[col] = float(_resolve_scalar_or_number(operand_expr, scalar_context))
                        elif operand_expr.replace(".", "").replace("-", "").isnumeric():
                            df[col] = float(operand_expr)
                except Exception as e:
                    print(f"    [警告] 值计算表达式解析失败: {expr}, 错误: {e}")
        
        # 添加新计算的列
        for new_col_name, calc_result in new_cols.items():
            df[new_col_name] = calc_result

    return result


def _find_matching_calc(col, calc_map, value_cols, raw_val_maps=None):
    for vcol, expr in calc_map.items():
        if col == vcol or col.startswith(vcol + "_"):
            return expr
    if raw_val_maps:
        for vcol, expr in calc_map.items():
            if col in raw_val_maps:
                map_idx = raw_val_maps.index(col)
                if map_idx < len(value_cols) and value_cols[map_idx] == vcol:
                    return expr
    # 多列表达式兜底：非运算符开头的表达式，只要当前列是数值列就返回
    # （多列分支的入口不依赖某个特定列，只需要任意数值列作为触发）
    for vcol, expr in calc_map.items():
        if not re.match(r"^[+\-*/]", expr.split("=")[0].strip()):
            return expr
    return None


def _format_numeric_value(val, fmt: str) -> str:
    """按格式后缀格式化数值，与模板填充的文本占位符格式语法一致。

    支持的格式：
        .Nf     → 保留 N 位小数（如 .2f → 4800.00）
        .N%     → 百分比，保留 N 位小数（如 .2% → 37.65%，输入值应为 0~1 的小数）
        int     → 整数
        无格式  → 原始值
    """
    if val is None:
        return ""
    try:
        num = float(val)
    except (ValueError, TypeError):
        return str(val)

    fmt = (fmt or "").strip()
    if not fmt:
        # 整数则去尾零，否则保持原样
        if num == int(num):
            return str(int(num))
        return str(num)

    if fmt == "int":
        return str(int(num))

    if fmt.startswith(".") and fmt.endswith("f"):
        try:
            digits = int(fmt[1:-1])
            return f"{num:.{digits}f}"
        except ValueError:
            return str(num)

    if fmt.startswith(".") and fmt.endswith("%"):
        try:
            digits = int(fmt[1:-1])
            return f"{num * 100:.{digits}f}%"
        except ValueError:
            return str(num)

    return str(num)


def _resolve_value_map_placeholder(text: str, block_results: dict) -> str:
    """解析值映射文本中的 {区块.列.行值|格式} 占位符，替换为前面任务的计算数值。

    语法与模板填充的文本占位符一致：
        {区块.列名}             → 该列所有行的合计值
        {区块.列名.行值}        → 精确取某行某列
        {区块.列名|格式}        → 合计值 + 格式化（如 |.2f）
        {区块.列名.行值|格式}   → 精确取值 + 格式化
        {标量名}               → 从 scalar_context 取（block_results 里标量区块首行）
        {标量名|格式}           → 标量 + 格式化

    :param text: 值映射原始文本（如 "华东占{按地区汇总.总销售额.华东|.0f}元"）
    :param block_results: {区块名: DataFrame} 字典
    :return: 替换后的文本，未找到的占位符保留原样
    """
    if not text or not block_results:
        return text

    # 匹配 {...} 占位符，不支持嵌套
    pattern = re.compile(r"\{([^{}]+)\}")

    def _replace(m):
        expr = m.group(1).strip()
        if not expr:
            return m.group(0)

        # 分离格式后缀
        fmt = ""
        if "|" in expr:
            parts = expr.split("|", 1)
            expr = parts[0].strip()
            fmt = parts[1].strip()

        # 拆分 区块.列.行值
        segments = [s.strip() for s in expr.split(".") if s.strip()]
        if not segments:
            return m.group(0)

        # 尝试匹配：1段=标量名/区块名，2段=区块.列，3段=区块.列.行值
        val = None
        resolved = False

        if len(segments) == 1:
            # 标量名：在 block_results 找单行 DataFrame 的第一个数值列
            name = segments[0]
            df = block_results.get(name)
            if df is not None and hasattr(df, "shape") and len(df) == 1:
                for col in df.columns:
                    try:
                        val = float(df[col].iloc[0])
                        resolved = True
                        break
                    except (ValueError, TypeError):
                        continue

        elif len(segments) == 2:
            block_name, col_name = segments
            df = block_results.get(block_name)
            if df is not None and col_name in df.columns:
                try:
                    s = pd.to_numeric(df[col_name], errors="coerce").dropna()
                    if len(s) > 0:
                        val = float(s.sum())
                        resolved = True
                except Exception:
                    pass

        elif len(segments) >= 3:
            block_name, col_name = segments[0], segments[1]
            row_value = ".".join(segments[2:])  # 行值本身可能含点
            df = block_results.get(block_name)
            if df is not None and len(df) > 0:
                row_dim_col = df.columns[0]
                mask = df[row_dim_col].astype(str).str.strip() == row_value
                matched = df[mask]
                if len(matched) > 0 and col_name in matched.columns:
                    try:
                        val = float(pd.to_numeric(matched[col_name].iloc[0], errors="coerce"))
                        resolved = True
                    except (ValueError, TypeError):
                        pass

        if not resolved:
            return m.group(0)  # 未找到，保留原样

        return _format_numeric_value(val, fmt)

    return pattern.sub(_replace, text)


def collect_task_scalars(result):
    """从无行维度任务的横向一行结果中收集标量值。
    
    返回 {列名: 值} 字典，供后续任务的值计算公式引用。
    有行维度的任务（多行分组结果）不产生标量。
    """
    scalars = {}
    if not isinstance(result, dict):
        return scalars
    for key, df in result.items():
        if not isinstance(df, pd.DataFrame):
            continue
        if len(df) != 1:
            continue
        for col in df.columns:
            if col in ("合计", "总计", "指标", "值"):
                continue
            try:
                val = df[col].iloc[0]
                try:
                    float_val = float(val)
                    scalars[str(col)] = float_val
                except (ValueError, TypeError):
                    pass
            except (ValueError, TypeError, IndexError):
                pass
    return scalars
