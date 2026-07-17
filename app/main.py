"""
Excel 统一分析工具 — PPT 生成 + 透视分析 + 模板填充
用法：
  python main.py ppt 文件夹路径           ← PPT 生成（自动找配置和数据）
  python main.py ppt -c 配置.xlsx         ← PPT 生成（指定配置）
  python main.py ppt -c 配置.xlsx -o out.pptx
  python main.py pivot 文件夹路径         ← 透视分析（自动找配置和数据）
  python main.py pivot -c 配置.xlsx       ← 透视分析（指定配置）
  python main.py pivot -c 配置.xlsx -o out.xlsx
  python main.py template 模板.pptx --pivot 透视结果.xlsx   ← 基于PPT模板填充数据
  python main.py template 模板.pptx                           ← 自动找模板目录下最新 xlsx
  python main.py template 模板.pptx --pivot latest            ← 同上，显式触发
  python main.py template 模板.pptx --pivot-dir D:\数据        ← 指定扫描目录
  python main.py 文件夹路径               ← 自动检测配置类型，分发到对应模式
"""
import os
import sys
import argparse
import glob
from datetime import datetime

# 版本信息
__VERSION__ = "2.54.3"
__UPDATE_DATE__ = "2026-07-17"

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))


def find_config_file(folder):
    xlsx_files = glob.glob(os.path.join(folder, "*.xlsx"))
    xlsx_files = [f for f in xlsx_files if not os.path.basename(f).startswith("~$")]
    if not xlsx_files:
        return None

    # 优先：文件名含「配置」/「config」的候选
    config_candidates = [f for f in xlsx_files
                         if "配置" in os.path.basename(f) or "config" in os.path.basename(f).lower()]

    # 无「配置」候选时，若只有一个 xlsx 则直接用
    if not config_candidates:
        if len(xlsx_files) == 1:
            return xlsx_files[0]
        # 多个 xlsx 但都不含「配置」→ 也纳入候选让用户选
        config_candidates = xlsx_files

    # 只有一个候选，直接返回
    if len(config_candidates) == 1:
        return config_candidates[0]

    # 多个候选：交互式选择
    return _prompt_select_config(config_candidates)


def _prompt_select_config(candidates):
    """多个配置文件时让用户选择。非交互环境取第一个并警告。"""
    # 非交互式（管道/重定向），不能 input()，取第一个并警告
    if not sys.stdin.isatty():
        chosen = candidates[0]
        print(f"[警告] 检测到 {len(candidates)} 个配置文件，非交互环境自动选择: {os.path.basename(chosen)}")
        print(f"        其他配置文件: {', '.join(os.path.basename(f) for f in candidates[1:])}")
        return chosen

    print(f"\n[信息] 检测到 {len(candidates)} 个配置文件，请选择:")
    for i, f in enumerate(candidates, 1):
        print(f"  {i}. {os.path.basename(f)}")

    while True:
        try:
            choice = input(f"请输入序号 (1-{len(candidates)})，回车默认选 1: ").strip()
            if not choice:
                return candidates[0]
            idx = int(choice)
            if 1 <= idx <= len(candidates):
                return candidates[idx - 1]
            print(f"[错误] 序号超出范围，请输入 1-{len(candidates)}")
        except ValueError:
            print("[错误] 请输入数字")
        except (EOFError, KeyboardInterrupt):
            print("\n[信息] 已取消选择")
            return None


def _auto_find_data_file(config_dir, config_path):
    """
    自动查找数据文件，支持 Excel 和 CSV。
    优先查找与配置同目录下的非配置数据文件。
    """
    from src.excel_reader import get_candidate_data_files
    candidates = get_candidate_data_files(config_dir)
    return candidates[0] if candidates else None


def _ensure_output_dir(config_dir):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = os.path.join(config_dir, f"output_{timestamp}")
    os.makedirs(out_dir, exist_ok=True)
    return out_dir


def _auto_find_latest_pivot(search_dir):
    """递归扫描目录及子目录下所有 xlsx，按文件创建时间取最新。

    排除：临时文件（~$）、配置文件（含"配置"/"config"）、模板文件（含"模板"/"template"）。
    """
    candidates = []
    for root, dirs, files in os.walk(search_dir):
        for f in files:
            if not f.lower().endswith(".xlsx"):
                continue
            if f.startswith("~$"):
                continue
            basename_lower = f.lower()
            if "配置" in basename_lower or "config" in basename_lower:
                continue
            if "模板" in basename_lower or "template" in basename_lower:
                continue
            full_path = os.path.join(root, f)
            candidates.append(full_path)

    if not candidates:
        return None

    candidates.sort(key=os.path.getctime, reverse=True)
    return candidates[0]


def _resolve_config_path(args):
    """解析配置文件路径。

    为避免歧义，位置参数只接受配置文件路径，不再自动扫描文件夹找配置。
    如需指定文件夹，请用 --data-dir 参数。
    """
    config_path = args.config
    folder_or_config = getattr(args, "folder_or_config", None)
    if not config_path and folder_or_config:
        if os.path.isdir(folder_or_config):
            # 传入文件夹时给出明确错误，避免自动选择错误的配置文件
            print(f"[错误] 请直接指定配置文件路径，而非文件夹: {folder_or_config}")
            print(f"       用法: python main.py pivot -c 配置文件.xlsx --data-dir 数据目录 -o 输出.xlsx")
            sys.exit(1)
        elif os.path.isfile(folder_or_config):
            config_path = folder_or_config
    if not config_path:
        print("[错误] 必须指定配置文件路径（位置参数或 -c/--config）")
        sys.exit(1)
    if not os.path.exists(config_path):
        print(f"[错误] 配置文件不存在: {config_path}")
        sys.exit(1)
    return os.path.abspath(config_path)


def _require_arg(value, arg_name, mode):
    """校验必填参数，未提供时报错退出"""
    if not value:
        print(f"[错误] {mode} 模式必须指定 {arg_name}")
        sys.exit(1)


def _apply_timestamp(output_path: str, add_ts: bool) -> str:
    """给输出文件路径追加时间戳。

    未启用时原样返回。启用时在文件名 stem 和扩展名之间插入 _YYYYMMDD_HHMMSS。
    示例: out/报告.pptx → out/报告_20260714_173000.pptx
    """
    if not add_ts or not output_path:
        return output_path
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dirname = os.path.dirname(output_path)
    basename = os.path.basename(output_path)
    if "." in basename:
        stem, ext = basename.rsplit(".", 1)
        new_basename = f"{stem}_{ts}.{ext}"
    else:
        new_basename = f"{basename}_{ts}"
    return os.path.join(dirname, new_basename) if dirname else new_basename


def _validate_required_args(args, mode):
    """校验各模式的必填参数"""
    if mode == "template":
        # template 模式在调用处单独校验
        return
    # pivot/ppt/auto 模式：必须提供 --data-dir 和 -o
    _require_arg(getattr(args, 'data_dir', None), "--data-dir <数据目录>", mode)
    _require_arg(getattr(args, 'output', None), "-o/--output <输出路径>", mode)


def _detect_mode(config_path):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(config_path, read_only=True)
    except Exception as e:
        print(f"[错误] 无法读取配置文件: {e}")
        sys.exit(1)
    ppt_keywords = {"页码", "页面类型", "页面标题", "图表类型"}
    pivot_keywords = {"数据源", "行维度", "列维度", "值字段", "聚合方式"}
    ppt_found = False
    pivot_found = False
    for name in wb.sheetnames:
        ws = wb[name]
        # 检查前 5 行（PPT 配置第一行可能是主题设置行）
        all_texts = set()
        for row in ws.iter_rows(min_row=1, max_row=5):
            for cell in row:
                v = cell.value
                if v is not None:
                    all_texts.add(str(v).strip())
        if len(ppt_keywords & all_texts) >= 2:
            ppt_found = True
        if len(pivot_keywords & all_texts) >= 2:
            pivot_found = True
    wb.close()
    if ppt_found and pivot_found:
        return "all"
    if pivot_found:
        return "pivot"
    if ppt_found:
        return "ppt"
    return "unknown"


def _append_chart(chart_map, chart_title, chart_def, page_title):
    """追加图表到 chart_map（list），同名时打印警告但不丢弃。"""
    same_name = [c for c in chart_map if c.get("图表标题") == chart_title]
    if same_name:
        print(f"    [警告] [{page_title}] 图表标题 '{chart_title}' 与已有图表重名，两者均保留")
    chart_map.append(chart_def)


def _run_ppt_mode(config_path, output_path=None, pivot_data_file=None, validate_only=False, data_dir=None):
    """
    生成PPT。
    :param pivot_data_file: 透视分析结果文件路径，用于 {pivot} 数据源引用
    :param validate_only: 仅校验配置不执行生成
    :param data_dir: 数据文件所在目录（数据源相对路径基于此目录，默认配置文件所在目录）
    """
    from src.excel_reader import (
        read_config, read_data, read_geo_data, find_data_file, get_data_file_info
    )
    from src.ppt_builder import build_ppt, validate_ppt_config, print_ppt_validation_results

    config_dir = os.path.dirname(config_path)
    # data_dir 必填（由 _validate_required_args 保证），直接使用
    eff_data_dir = data_dir
    print(f"    → 数据目录: {data_dir}")
    print(f"[PPT/1] 读取配置: {config_path}")
    config = read_config(config_path)
    general = config.get("general", {})
    pages = config.get("pages", [])
    colors = config.get("colors", {})
    print(f"    → PPT页面: {len(pages)} 页")

    # 配置校验
    print(f"[PPT/校验] 检查配置...")
    val_results, all_ok = validate_ppt_config(config, eff_data_dir, pivot_data_file)
    print_ppt_validation_results(val_results)

    if validate_only:
        print("[校验] 仅校验模式，不执行生成")
        return None

    if not all_ok:
        print("[警告] 配置校验发现以上错误，将继续执行（运行时可能失败）")

    # 跳过「是否生成=否」的页面
    skip_keywords = ("否", "no", "false", "0", "不生成", "跳过", "skip")
    filtered_pages = []
    for page in pages:
        should_gen = str(page.get("是否生成", "是")).strip().lower()
        if should_gen in skip_keywords:
            print(f"    [SKIP] 第{page.get('页码','?')}页：是否生成=否")
            continue
        filtered_pages.append(page)
    pages = filtered_pages
    config["pages"] = pages

    if pivot_data_file:
        print(f"    → 透视结果数据源: {os.path.basename(pivot_data_file)}")
    else:
        # 检测是否引用了透视结果但未提供文件，提前给出明确提示
        pivot_keywords = ("{pivot}", "pivot", "透视结果", "透视分析")
        refs_pivot = any(
            str(cd.get("数据源", "")).strip().lower() in pivot_keywords
            for page in pages for cd in page.get("charts", [])
        )
        if refs_pivot:
            print(f"    [警告] 配置引用了 {{pivot}} 数据源但未提供 --pivot-file，相关图表将被跳过")
            print(f"           建议：使用 'python main.py 配置.xlsx' 自动模式（先透视后PPT），")
            print(f"                 或先用 pivot 子命令生成结果，再用 'ppt --pivot-file 结果.xlsx' 指定")

    # 全局数据文件路径
    default_data_file = general.get("数据文件", general.get("excel_path", ""))
    if not default_data_file:
        default_data_file = _auto_find_data_file(eff_data_dir, config_path)
        if default_data_file:
            print(f"    → 自动找到数据文件: {os.path.basename(default_data_file)}")
    elif not os.path.isabs(str(default_data_file)):
        default_data_file = os.path.join(eff_data_dir, str(default_data_file))

    print(f"    → 数据文件: {os.path.basename(default_data_file) if default_data_file else '无'}")

    print(f"[PPT/2] 读取图表数据...")
    total_charts = 0
    chart_map = []  # 使用 list 避免同名图表键覆盖

    for page_def in pages:
        page_title = str(page_def.get("页面标题", f"第{page_def.get('页码', '?')}页"))
        for chart_def in page_def.get("charts", []):
            chart_title = chart_def.get("图表标题", "")
            chart_type = str(chart_def.get("图表类型", "")).strip().lower()
            data_sheet = chart_def.get("数据Sheet", "Sheet1")
            data_source = chart_def.get("数据源", "")
            x_range = chart_def.get("X轴范围", "")
            y_range = chart_def.get("Y轴范围", "")
            block_name = chart_def.get("区块名", "")

            # 查找数据文件：根据图表指定的数据源或全局默认
            if data_source:
                ds_lower = str(data_source).strip().lower()
                # 支持 {pivot} / pivot / 透视结果 引用透视分析结果
                if ds_lower in ("{pivot}", "pivot", "透视结果", "透视分析"):
                    if pivot_data_file and os.path.exists(pivot_data_file):
                        file_path = pivot_data_file
                        print(f"    [信息] [{page_title}] {chart_title} 使用透视结果数据")
                    else:
                        print(f"    ! [{page_title}] {chart_title} 引用透视结果但未找到透视输出文件")
                        continue
                else:
                    file_path = find_data_file(data_source, eff_data_dir)
                    if file_path:
                        print(f"    [信息] [{page_title}] {chart_title} 使用数据源: {os.path.basename(file_path)}")
                    else:
                        print(f"    ! [{page_title}] {chart_title} 数据源 '{data_source}' 未找到")
                        continue
            else:
                file_path = default_data_file
            
            if not file_path or not os.path.exists(str(file_path)):
                print(f"    ! [{page_title}] {chart_title} 数据文件不存在: {file_path}")
                continue

            try:
                if chart_type in ("map", "heatmap"):
                    geo_df = read_geo_data(file_path, data_sheet, x_range, y_range)
                    if geo_df is not None and len(geo_df) > 0:
                        chart_def["_geo_df"] = geo_df
                        chart_def["_is_map"] = True
                        _append_chart(chart_map, chart_title, chart_def, page_title)
                        total_charts += 1
                        print(f"    [OK] [{page_title}] {chart_title} (地图)")
                    else:
                        print(f"    - [{page_title}] {chart_title} (无地理数据)")
                    continue

                x_values, y_values = read_data(file_path, data_sheet, x_range, y_range, block_name)
                if x_values and y_values:
                    chart_def["_categories"] = x_values
                    chart_def["_values"] = y_values
                    if isinstance(x_values, list) and x_values and isinstance(x_values[0], (tuple, list)):
                        chart_def["_is_hierarchical"] = True
                    _append_chart(chart_map, chart_title, chart_def, page_title)
                    total_charts += 1
                    print(f"    [OK] [{page_title}] {chart_title}")
                else:
                    print(f"    - [{page_title}] {chart_title} (无数据)")
            except KeyError as e:
                print(f"    ! [{page_title}] {chart_title} 列名未找到: {e}")
                print(f"      建议: 检查 X轴='{x_range}' Y轴='{y_range}' 是否与数据 Sheet '{data_sheet}' 的列名完全一致")
            except (FileNotFoundError, PermissionError) as e:
                print(f"    ! [{page_title}] {chart_title} 数据文件访问失败: {e}")
            except Exception as e:
                msg = str(e).lower()
                if "sheet" in msg or "worksheet" in msg:
                    print(f"    ! [{page_title}] {chart_title} Sheet不存在: {e}")
                    print(f"      建议: 检查 数据Sheet='{data_sheet}' 是否存在于 {os.path.basename(file_path)}")
                elif "column" in msg or "key" in msg:
                    print(f"    ! [{page_title}] {chart_title} 列名匹配失败: {e}")
                    print(f"      建议: 检查 X轴='{x_range}' Y轴='{y_range}' 列名是否精确匹配")
                else:
                    print(f"    ! [{page_title}] {chart_title} (错误: {e})")

    print(f"    → 共 {total_charts} 个图表")
    if total_charts == 0 and not pages:
        print("[错误] 没有有效内容。")
        sys.exit(1)

    if not output_path:
        # 有透视结果文件时，复用其所在目录，避免 Excel 和 PPT 分散到不同目录
        if pivot_data_file and os.path.exists(str(pivot_data_file)):
            output_dir = os.path.dirname(os.path.abspath(str(pivot_data_file)))
        else:
            output_dir = _ensure_output_dir(config_dir)
        base_name = os.path.splitext(os.path.basename(config_path))[0]
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(output_dir, f"{base_name}_报告_{timestamp_str}.pptx")
    else:
        output_dir = os.path.dirname(output_path) or "."
        os.makedirs(output_dir, exist_ok=True)

    if os.path.exists(output_path):
        print(f"    [警告] 输出文件已存在，将被覆盖: {os.path.basename(output_path)}")
    print(f"[PPT/3] 生成PPT: {output_path}")
    build_ppt(config, chart_map, output_path)
    print(f"\n[OK] 完成！PPT已保存至: {output_path}")
    print(f"   共 {len(pages)} 页, {total_charts} 个原生图表")
    print(f"   提示: 双击PPT中的图表可查看和编辑数据")


def _run_html_from_pivot(pivot_path, output_path=None, tasks=None):
    """从透视结果 Excel 自动生成 HTML 报告（无需额外配置）。

    每个 sheet 自动生成一个图表 section，图表类型自动推断，支持柱状/折线/饼图/表格切换。
    tasks: 透视配置任务列表（可选），用于提取行维度信息以精确识别分类列。
    """
    from src.html_builder import generate_html_report

    if not pivot_path or not os.path.exists(pivot_path):
        print(f"[HTML] 透视结果文件不存在，跳过 HTML 生成: {pivot_path}")
        return None

    if not output_path:
        base = os.path.basename(pivot_path).rsplit(".", 1)[0]
        out_dir = os.path.dirname(pivot_path) or "."
        output_path = os.path.join(out_dir, f"{base}_报告.html")

    # 从 tasks 提取 {结果Sheet名: [行维度列名]}，用于精确识别分类列
    dim_map = {}
    if tasks:
        for t in tasks:
            sheet = t.get("结果Sheet", "")
            dims_str = str(t.get("行维度", "") or "").strip()
            if sheet and dims_str:
                dims = [d.strip() for d in dims_str.replace("，", ",").split(",") if d.strip()]
                if dims:
                    dim_map[sheet] = dims

    print(f"[HTML] 从透视结果生成 HTML 报告: {os.path.basename(pivot_path)}")
    html_path = generate_html_report(
        excel_path=pivot_path,
        ppt_config=None,  # 无 PPT 配置 → 纯 Excel 自动生成模式
        output_dir=os.path.dirname(output_path) or ".",
        report_title="透视分析报告",
        report_subtitle="自动生成",
        dim_map=dim_map or None,
    )
    if html_path:
        # generate_html_report 内部用自己的文件名，重命名为用户指定的
        if os.path.abspath(html_path) != os.path.abspath(output_path) and os.path.exists(html_path):
            os.replace(html_path, output_path)
            html_path = output_path
        print(f"[OK] HTML 报告已保存至: {html_path}")
    return html_path


def _write_join_intermediate(output_path, join_intermediate):
    """把 JOIN 中间表写入独立 Excel 文件，方便检查多表 JOIN 后的数据是否正确。

    :param output_path: 主输出文件路径，JOIN 中间表文件名基于此推导
    :param join_intermediate: {sheet名: DataFrame}，每个 JOIN 任务一个条目
    :return: JOIN 中间表文件路径，失败返回 None
    """
    import openpyxl
    import pandas as pd
    from src.excel_writer import _safe_sheet_name, _auto_fit_columns

    dirname = os.path.dirname(output_path)
    basename = os.path.basename(output_path)
    if "." in basename:
        stem, ext = basename.rsplit(".", 1)
    else:
        stem, ext = basename, "xlsx"
    join_path = os.path.join(dirname, f"{stem}_JOIN中间表.{ext}")

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sheet_name, df in join_intermediate.items():
            safe_name = _safe_sheet_name(str(sheet_name))
            ws = wb.create_sheet(safe_name)
            # 写表头
            ws.append(list(df.columns))
            # 写数据行
            for _, row in df.iterrows():
                ws.append([row[c] if not pd.isna(row[c]) else None for c in df.columns])
            # 冻结首行 + 自动列宽
            ws.freeze_panes = "A2"
            _auto_fit_columns(ws)
        wb.save(join_path)
        return join_path
    except Exception as e:
        print(f"    [警告] JOIN 中间表写入失败: {e}")
        return None


def _merge_same_block_results(prev_df, new_df, block_name, seq):
    """同名区块结果横向合并（按行维度列对齐）。

    修复级联透视"找不到字段"根因：多行配置相同区块名时，旧逻辑直接覆盖，
    导致后续 {区块名} 引用只能拿到最后一个任务的列，前面任务的列丢失。

    合并策略：
    1. 找两个 df 的公共列（通常是行维度列）
    2. 有公共列 → 基于公共列做 outer merge，列变多，行按公共列对齐
    3. 无公共列 → 退化为 concat(axis=0) 纵向堆叠（行变多），并打印警告
    """
    import pandas as pd
    import traceback
    try:
        prev_cols = list(prev_df.columns)
        new_cols = list(new_df.columns)
        common_cols = [c for c in prev_cols if c in new_cols]
        if common_cols:
            # 基于公共列（行维度）outer merge，保留两边所有行和列
            merged = pd.merge(prev_df, new_df, on=common_cols, how="outer")
            # 去除 merge 可能产生的重复列（非公共列同名时 merge 会加 _x/_y 后缀）
            merged = merged.loc[:, ~merged.columns.duplicated()]
            print(f"    [信息] 区块名 '{block_name}' 同名任务横向合并（任务{seq}）")
            print(f"           公共列: {common_cols}，合并后列数: {len(merged.columns)}，行数: {len(merged)}")
            return merged
        else:
            # 无公共列，纵向堆叠
            merged = pd.concat([prev_df, new_df], axis=0, ignore_index=True)
            print(f"    [警告] 区块名 '{block_name}' 同名任务无公共列，纵向堆叠（任务{seq}）")
            print(f"           前序列: {prev_cols}，当前列: {new_cols}")
            print(f"           建议改用 {{结果Sheet名.区块名}} 分别精确引用")
            return merged
    except Exception as e:
        print(f"    [错误] 区块名 '{block_name}' 同名任务合并失败（任务{seq}），回退为覆盖")
        print(f"           异常: {type(e).__name__}: {e}")
        print(f"           前序列: {list(prev_df.columns) if hasattr(prev_df, 'columns') else type(prev_df)}")
        print(f"           当前列: {list(new_df.columns) if hasattr(new_df, 'columns') else type(new_df)}")
        traceback.print_exc()
        return new_df


def _run_pivot_mode(config_path, output_path=None, validate_only=False, data_dir=None):
    from src.pivot_analyzer import (
        read_pivot_config, run_analysis, find_data_files,
        validate_pivot_config, print_validation_results,
        collect_task_scalars
    )
    from src.excel_writer import write_results

    config_dir = os.path.dirname(config_path)
    # data_dir 必填（由 _validate_required_args 保证），直接使用
    eff_data_dir = data_dir
    print(f"[Pivot/1] 读取配置: {config_path}")
    print(f"    → 数据目录: {data_dir}")
    tasks = read_pivot_config(config_path)
    print(f"    → 共 {len(tasks)} 个分析任务")

    if not tasks:
        print("[错误] 没有有效的分析任务。")
        sys.exit(1)

    data_files = find_data_files(eff_data_dir, config_path)
    if data_files:
        print(f"    → 找到数据文件: {[os.path.basename(f) for f in data_files]}")

    # 配置校验
    print(f"[Pivot/校验] 检查配置...")
    val_results = validate_pivot_config(tasks, eff_data_dir)
    all_ok = print_validation_results(val_results)
    
    if validate_only:
        print("[校验] 仅校验模式，不执行分析")
        return None
    
    if not all_ok:
        print("[警告] 配置校验发现以上错误，将继续执行（运行时可能失败）")

    print(f"[Pivot/2] 执行透视分析...")
    results = []
    errors = []
    skipped = 0
    scalar_context: dict = {}
    block_results: dict = {}  # {区块名: DataFrame}，供后续任务的值映射占位符引用
    join_intermediate: dict = {}  # {sheet名: DataFrame}，JOIN 中间表单独收集，写入独立 Excel 供检查

    for task in tasks:
        seq = task.get("序号", "?")
        sheet_name = task.get("结果Sheet", f"结果{seq}")

        should_calc = str(task.get("是否计算", "是")).strip()
        if should_calc.lower() in ("否", "no", "false", "0", "不计算", "跳过", "skip"):
            print(f"    [SKIP] [任务{seq}] 已设置为不计算，跳过")
            results.append(None)
            skipped += 1
            continue

        try:
            result, error = run_analysis(task, eff_data_dir, scalar_context, block_results)
            if error:
                print(f"    [FAIL] [任务{seq}] {error}")
                errors.append({"序号": seq, "错误": error})
                results.append(None)
            else:
                # 剥离 JOIN 中间表：单独收集到 join_intermediate，不混入正常结果
                # 修复"JOIN 任务执行两次"问题：原逻辑遍历 result.items() 会把 _JOIN中间表_ 当正常结果打印和写入
                if isinstance(result, dict):
                    join_keys_to_pop = [k for k in result.keys() if str(k).startswith("_JOIN中间表_")]
                    for jk in join_keys_to_pop:
                        join_df = result.pop(jk)
                        # 提取原始 sheet 名（去掉 _JOIN中间表_ 前缀）
                        orig_sheet = str(jk).replace("_JOIN中间表_", "", 1)
                        # key 加任务序号防止多个 JOIN 任务结果Sheet同名时互相覆盖
                        # sheet 名用 "任务{seq}_{orig_sheet}" 防止 Excel sheet 重名（Excel sheet 名必须唯一）
                        join_intermediate[f"任务{seq}_{orig_sheet}"] = join_df
                        print(f"    [JOIN] [任务{seq}] 中间表已收集: {orig_sheet} -> {join_df.shape[0]}行 x {join_df.shape[1]}列")

                # 明细模式：结果仅作中间数据供后续任务引用，不写入主结果 Excel
                # 但写入 _JOIN中间表.xlsx 供人工检查（用 [明细] 前缀区分 JOIN 中间表）
                # 适用场景：原表 JOIN 前序拼接结果等"中间步骤"，不应污染最终输出
                _agg_check = str(task.get("聚合方式", "sum")).strip().lower()
                _is_detail_mode = _agg_check in ("明细", "raw", "passthrough", "detail", "原样", "不聚合")

                if isinstance(result, dict):
                    for key, df in result.items():
                        if hasattr(df, "shape"):
                            tag = "[明细]" if _is_detail_mode else "[OK]"
                            detail_note = ""
                            if _is_detail_mode:
                                # 写入中间表 Excel 供检查
                                detail_key = f"[明细]任务{seq}_{sheet_name}"
                                join_intermediate[detail_key] = df
                                detail_note = "（仅中间数据，不写入主Excel，已保存至中间表）"
                            print(f"    {tag} [任务{seq}] {sheet_name} -> {df.shape[0]}行 x {df.shape[1]}列{detail_note}")
                        else:
                            print(f"    [OK] [任务{seq}] {sheet_name} (标量)")
                else:
                    print(f"    [OK] [任务{seq}] {sheet_name}")
                # 明细模式：不 append 到 results（不写入主 Excel），但 block_results 仍存入供后续引用
                results.append(None if _is_detail_mode else result)
                # 收集无行维度任务产生的标量，供后续任务公式引用
                task_scalars = collect_task_scalars(result)
                if task_scalars:
                    scalar_context.update(task_scalars)
                # 收集本任务结果到 block_results，供后续任务的值映射占位符引用
                # 区块名优先用 task 的"区块名"，否则用"结果Sheet"
                block_name = task.get("区块名", "") or sheet_name
                if isinstance(result, dict):
                    # 收集所有 DataFrame 类型的结果（result 已剥离 _JOIN中间表_，无需再过滤）
                    dfs_in_result = [(key, df) for key, df in result.items() if hasattr(df, "shape")]
                    if dfs_in_result:
                        if len(dfs_in_result) == 1:
                            # 单结果：直接存入
                            merged_df = dfs_in_result[0][1]
                        else:
                            # 多结果（多值字段交叉透视等场景）：横向合并所有 DataFrame
                            # 保证后续 {区块名} 引用能看到所有值字段的列
                            if len(dfs_in_result) > 1:
                                import pandas as _pd
                                # 按行维度列对齐横向合并，相同列名用后缀区分
                                merged_df = _pd.concat([df for _, df in dfs_in_result], axis=1)
                                # 去除合并后可能重复的行维度列（取第一次出现的）
                                merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]
                            else:
                                merged_df = dfs_in_result[0][1]
                        # 同名区块处理：不再覆盖，而是按行维度对齐横向合并
                        # 修复级联透视"找不到字段"根因：多行配置相同区块名，最后一行覆盖前面的，
                        # 导致后续 {区块名} 引用只能拿到最后一个任务的列，前面任务的列丢失
                        # 组合 key 始终指向本任务自己的原始结果（未合并），用于 {结果Sheet名.区块名} 精确引用
                        block_results[(block_name, sheet_name)] = merged_df
                        if block_name in block_results:
                            prev_df = block_results[block_name]
                            merged_df = _merge_same_block_results(prev_df, merged_df, block_name, seq)
                        block_results[block_name] = merged_df
                        # 同时用结果Sheet名作为别名存入，使后续任务既可用 {区块名}
                        # 也可用 {结果Sheet名} 引用本任务输出（区块名与结果Sheet不同时两者都生效）
                        if sheet_name and sheet_name != block_name:
                            block_results[sheet_name] = merged_df
        except Exception as e:
            print(f"    [FAIL] [任务{seq}] 异常: {e}")
            errors.append({"序号": seq, "错误": str(e)})
            results.append(None)

    valid_results = [r for r in results if r is not None]
    if not valid_results:
        print("[错误] 所有任务均失败，未生成输出。")
        sys.exit(1)

    if not output_path:
        output_dir = _ensure_output_dir(config_dir)
        base_name = os.path.splitext(os.path.basename(config_path))[0]
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(output_dir, f"{base_name}_分析_{timestamp_str}.xlsx")
    else:
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    if os.path.exists(output_path):
        print(f"    [警告] 输出文件已存在，将被覆盖: {os.path.basename(output_path)}")
    print(f"[Pivot/3] 输出结果: {output_path}")
    valid_tasks = [t for t, r in zip(tasks, results) if r is not None]
    write_results(valid_tasks, valid_results, errors, output_path)

    # 中间表单独写入 Excel（含 JOIN 中间表 + 明细模式中间结果），方便检查
    # 文件名：{输出文件stem}_JOIN中间表.xlsx
    if join_intermediate:
        join_output_path = _write_join_intermediate(output_path, join_intermediate)
        if join_output_path:
            join_count = sum(1 for k in join_intermediate if not str(k).startswith("[明细]"))
            detail_count = sum(1 for k in join_intermediate if str(k).startswith("[明细]"))
            parts = []
            if join_count:
                parts.append(f"{join_count} 个 JOIN")
            if detail_count:
                parts.append(f"{detail_count} 个明细")
            print(f"   中间表已保存至: {join_output_path}（{' + '.join(parts)}）")

    print(f"\n[OK] 完成！分析结果已保存至: {output_path}")
    print(f"   共 {len(tasks)} 个任务: {len(valid_tasks)} 个成功" + (f", {skipped} 个跳过" if skipped else "") + (f", {len(errors)} 个失败" if errors else ""))
    if errors:
        print(f"   失败详情见「错误信息」Sheet")
    return output_path, valid_tasks


class _GuiLogRedirector:
    """重定向 stdout 到 GUI Text 控件"""
    def __init__(self, text_widget, tag=""):
        self.text_widget = text_widget
        self.tag = tag
        self._lock = False

    def write(self, s):
        if self._lock or not s:
            return
        self._lock = True
        try:
            self.text_widget.after(0, lambda: self._append(s))
        finally:
            self._lock = False

    def _append(self, s):
        self.text_widget.insert("end", s, self.tag)
        self.text_widget.see("end")

    def flush(self):
        pass


class _StdoutTee:
    """把 stdout/stderr 同时写到原流和日志文件，方便远程排查问题。

    CLI 模式下安装：所有 print / traceback 都会同时写入 logs/last_run.log。
    GUI 模式下不安装（GUI 用 _GuiLogRedirector 写到日志框）。
    """
    def __init__(self, original, log_file):
        self.original = original
        self.log_file = log_file

    def write(self, s):
        try:
            self.original.write(s)
        except Exception:
            pass
        if s:
            try:
                self.log_file.write(s)
                self.log_file.flush()
            except Exception:
                pass

    def flush(self):
        try:
            self.original.flush()
        except Exception:
            pass
        try:
            self.log_file.flush()
        except Exception:
            pass


def _install_cli_log_tee():
    """安装 CLI 日志 Tee，返回 (old_stdout, old_stderr) 或 None。"""
    try:
        project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        logs_dir = os.path.join(project_dir, "logs")
        if not os.path.exists(logs_dir):
            os.makedirs(logs_dir, exist_ok=True)
        log_path = os.path.join(logs_dir, "last_run.log")
        # 用 utf-8 编码，避免中文/特殊字符乱码
        log_file = open(log_path, "w", encoding="utf-8", buffering=1)
        log_file.write(f"===== excel2ppt 运行日志 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} =====\n")
        log_file.write(f"版本: {__VERSION__}\n")
        log_file.write(f"参数: {' '.join(sys.argv[1:])}\n")
        log_file.write("=" * 60 + "\n\n")
        old_stdout = sys.stdout
        old_stderr = sys.stderr
        sys.stdout = _StdoutTee(old_stdout, log_file)
        sys.stderr = _StdoutTee(old_stderr, log_file)
        return old_stdout, old_stderr
    except Exception as e:
        print(f"[警告] 无法安装日志 Tee: {e}")
        return None


def _scan_files(folder):
    """扫描文件夹，返回检测到的配置和数据文件信息"""
    if not folder or not os.path.isdir(folder):
        return []

    xlsx_files = glob.glob(os.path.join(folder, "*.xlsx"))
    csv_files = glob.glob(os.path.join(folder, "*.csv"))
    all_data_files = [f for f in (xlsx_files + csv_files)
                      if not os.path.basename(f).startswith("~$")]

    results = []

    # 查找配置文件
    config_candidates = []
    for f in all_data_files:
        name = os.path.basename(f)
        if "配置" in name or "config" in name.lower():
            config_candidates.append(f)

    if not config_candidates and xlsx_files:
        config_candidates = [f for f in xlsx_files if not os.path.basename(f).startswith("~$")]

    import openpyxl
    ppt_keywords = {"页码", "页面类型", "页面标题", "图表类型"}
    pivot_keywords = {"数据源", "行维度", "列维度", "值字段", "聚合方式"}

    for cfg in config_candidates:
        name = os.path.basename(cfg)
        try:
            wb = openpyxl.load_workbook(cfg, read_only=True)
            ppt_found = False
            pivot_found = False
            sheet_names = wb.sheetnames
            for sn in sheet_names:
                ws = wb[sn]
                try:
                    row = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
                except StopIteration:
                    continue
                row_set = set(row)
                if len(ppt_keywords & row_set) >= 2:
                    ppt_found = True
                if len(pivot_keywords & row_set) >= 2:
                    pivot_found = True
            wb.close()

            if ppt_found and pivot_found:
                mode = "综合配置 (PPT + 透视)"
            elif pivot_found:
                mode = "透视分析配置"
            elif ppt_found:
                mode = "PPT报告配置"
            else:
                continue

            results.append({"type": "config", "name": name, "mode": mode, "path": cfg})
        except Exception:
            continue

    # 查找数据文件
    data_candidates = []
    for f in all_data_files:
        name = os.path.basename(f)
        if name.startswith("~$"):
            continue
        is_config = any("配置" in r["name"] or "config" in r["name"].lower() for r in results)
        if results and any(name == r["name"] for r in results):
            continue
        if "配置" in name or "config" in name.lower():
            continue
        data_candidates.append(f)

    for df in data_candidates:
        name = os.path.basename(df)
        ftype = "CSV" if name.lower().endswith(".csv") else "Excel"
        results.append({"type": "data", "name": name, "mode": f"数据文件 ({ftype})", "path": df})

    return results


def _draw_gradient_banner(canvas, height, top_hex, bottom_hex):
    """在 Canvas 上绘制从上到下的渐变矩形"""
    r1, g1, b1 = int(top_hex[1:3], 16), int(top_hex[3:5], 16), int(top_hex[5:7], 16)
    r2, g2, b2 = int(bottom_hex[1:3], 16), int(bottom_hex[3:5], 16), int(bottom_hex[5:7], 16)
    w = canvas.winfo_reqwidth() or 900
    for i in range(height):
        t = i / max(height - 1, 1)
        r = int(r1 + (r2 - r1) * t)
        g = int(g1 + (g2 - g1) * t)
        b = int(b1 + (b2 - b1) * t)
        color = f"#{r:02x}{g:02x}{b:02x}"
        canvas.create_rectangle(0, i, w + 200, i + 1, fill=color, outline="")
    canvas.bind("<Configure>", lambda e: _redraw_gradient(canvas, height, top_hex, bottom_hex))


def _redraw_gradient(canvas, height, top_hex, bottom_hex):
    canvas.delete("all")
    _draw_gradient_banner_content(canvas, height, top_hex, bottom_hex)


def _draw_gradient_banner_content(canvas, height, top_hex, bottom_hex):
    w = canvas.winfo_width() or 900
    for i in range(height):
        r1, g1, b1 = int(top_hex[1:3], 16), int(top_hex[3:5], 16), int(top_hex[5:7], 16)
        r2, g2, b2 = int(bottom_hex[1:3], 16), int(bottom_hex[3:5], 16), int(bottom_hex[5:7], 16)
        t = i / max(height - 1, 1)
        r = int(r1 + (r2 - r1) * t)
        g = int(g1 + (g2 - g1) * t)
        b = int(b1 + (b2 - b1) * t)
        color = f"#{r:02x}{g:02x}{b:02x}"
        canvas.create_rectangle(0, i, w + 200, i + 1, fill=color, outline="")
    canvas.create_text(16, 18, text="Excel 统一分析工具", anchor="w",
                       font=("Microsoft YaHei", 15, "bold"), fill="#FFFFFF")
    canvas.create_text(16, 38, text=f"PPT 报告生成 + 透视分析  |  v{__VERSION__}  |  更新: {__UPDATE_DATE__}",
                       anchor="w", font=("Microsoft YaHei", 8), fill="#C0D4F0")


def _select_folder():
    try:
        import tkinter as tk
        from tkinter import filedialog, ttk, scrolledtext
        import threading

        root = tk.Tk()
        root.title("Excel 统一分析工具")
        root.geometry("720x720")
        root.minsize(600, 560)
        root.configure(bg="#f7f8fa")

        # 样式
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TFrame", background="#f7f8fa")
        style.configure("Header.TLabel", font=("Microsoft YaHei", 16, "bold"), foreground="#182B49", background="#f7f8fa")
        style.configure("Sub.TLabel", font=("Microsoft YaHei", 11, "bold"), foreground="#182B49", background="#f7f8fa")
        style.configure("Info.TLabel", font=("Microsoft YaHei", 9), foreground="#666666", background="#f7f8fa")
        style.configure("Guide.TButton", font=("Microsoft YaHei", 10), padding=6)
        style.configure("Run.TButton", font=("Microsoft YaHei", 11, "bold"), padding=8)
        style.configure("Pick.TButton", font=("Microsoft YaHei", 10), padding=6)
        style.configure("Scan.TButton", font=("Microsoft YaHei", 10), padding=6)

        # 主容器
        main_frame = tk.Frame(root, bg="#f7f8fa", padx=16, pady=12)
        main_frame.pack(fill="both", expand=True)

        # 渐变色标题横幅（Canvas 绘制）
        header_canvas = tk.Canvas(main_frame, height=56, bg="#f7f8fa", highlightthickness=0)
        header_canvas.pack(fill="x", pady=(0, 8))
        _draw_gradient_banner(header_canvas, 56, "#182B49", "#2E75B6")
        header_canvas.create_text(16, 18, text="Excel 统一分析工具", anchor="w",
                                  font=("Microsoft YaHei", 15, "bold"), fill="#FFFFFF")
        header_canvas.create_text(16, 38, text=f"PPT 报告生成 + 透视分析  |  v{__VERSION__}  |  更新: {__UPDATE_DATE__}",
                                  anchor="w", font=("Microsoft YaHei", 8), fill="#C0D4F0")

        # 路径选择行（卡片内）
        path_outer = tk.Frame(main_frame, bg="#ffffff", highlightbackground="#e0e4ea", highlightthickness=1, padx=10, pady=10)
        path_outer.pack(fill="x", pady=(0, 8))
        path_frame = tk.Frame(path_outer, bg="#ffffff")
        path_frame.pack(fill="x")

        tk.Label(path_frame, text="📁 选择文件夹", font=("Microsoft YaHei", 11, "bold"),
                 fg="#182B49", bg="#ffffff").pack(anchor="w", pady=(0, 4))

        path_row = tk.Frame(path_frame, bg="#ffffff")
        path_row.pack(fill="x")
        path_var = tk.StringVar(value="")
        path_entry = tk.Entry(path_row, textvariable=path_var, font=("Microsoft YaHei", 10), state="readonly",
                              relief="solid", borderwidth=1, bg="#f5f7fa")
        path_entry.pack(side="left", fill="x", expand=True, ipady=4)

        ttk.Button(path_row, text="选择文件夹", style="Pick.TButton",
                   command=lambda: _browse_folder(path_var, file_list_text, root)).pack(side="left", padx=(8, 0))
        ttk.Button(path_row, text="扫描文件", style="Scan.TButton",
                   command=lambda: _scan_and_display(path_var.get(), file_list_text)).pack(side="left", padx=(6, 0))

        # 检测到的文件区域（卡片内）
        file_outer = tk.Frame(main_frame, bg="#ffffff", highlightbackground="#e0e4ea", highlightthickness=1, padx=10, pady=10)
        file_outer.pack(fill="x", pady=(0, 8))
        file_frame = tk.Frame(file_outer, bg="#ffffff")
        file_frame.pack(fill="x")
        tk.Label(file_frame, text="📄 检测到的文件", font=("Microsoft YaHei", 11, "bold"),
                 fg="#182B49", bg="#ffffff").pack(anchor="w", pady=(0, 4))

        file_list_text = tk.Text(file_frame, height=6, wrap="word", font=("Microsoft YaHei", 10),
                                  bg="#ffffff", fg="#333333", relief="solid", borderwidth=1)
        file_list_text.pack(fill="x", expand=True)

        # 配置文件选择区域
        config_select_frame = tk.Frame(file_frame, bg="#ffffff")
        config_combo_var = tk.StringVar()
        ttk.Label(config_select_frame, text="选择配置文件:", style="Info.TLabel").pack(side="left", padx=(0, 6))
        config_combo = ttk.Combobox(config_select_frame, textvariable=config_combo_var,
                                    state="readonly", width=45, font=("Microsoft YaHei", 10))
        config_combo.pack(side="left")
        file_list_text._config_select_frame = config_select_frame
        file_list_text._config_combo = config_combo
        file_list_text._config_combo_var = config_combo_var
        file_list_text._config_paths = []

        # 按钮行（放在日志区域之前 pack，确保始终可见）——卡片包裹
        btn_outer = tk.Frame(main_frame, bg="#ffffff", highlightbackground="#e0e4ea", highlightthickness=1, padx=10, pady=10)
        btn_outer.pack(fill="x", pady=(4, 0))
        tk.Label(btn_outer, text="🎯 操作", font=("Microsoft YaHei", 11, "bold"),
                 fg="#182B49", bg="#ffffff").pack(anchor="w", pady=(0, 6))
        btn_frame = tk.Frame(btn_outer, bg="#ffffff")
        btn_frame.pack(fill="x")

        run_state = {"running": False}

        def do_run():
            folder = path_var.get()
            if not folder:
                log_text.insert("end", "[错误] 请先选择文件夹\n")
                log_text.see("end")
                return
            if run_state["running"]:
                log_text.insert("end", "[警告] 分析正在运行中，请等待完成\n")
                log_text.see("end")
                return
            run_state["running"] = True
            run_btn.config(state="disabled", text="运行中...")
            log_text.delete("1.0", "end")

            # 读取选中的配置文件（多配置文件时由用户在下拉框选择）
            raw_args = [folder]
            selected_idx = getattr(file_list_text, "_config_combo", None)
            if selected_idx is not None:
                combo = file_list_text._config_combo
                idx = combo.current()
                paths = getattr(file_list_text, "_config_paths", [])
                if idx >= 0 and idx < len(paths):
                    raw_args = ["-c", paths[idx]]

            threading.Thread(target=_run_analysis_thread,
                             args=(raw_args, log_text, run_btn, run_state),
                             daemon=True).start()

        ttk.Button(btn_frame, text="操作指南", style="Guide.TButton",
                   command=_open_guide).pack(side="left")

        run_btn = ttk.Button(btn_frame, text="开始分析", style="Run.TButton", command=do_run)
        run_btn.pack(side="left", fill="x", expand=True, padx=(8, 0))

        ttk.Button(btn_frame, text="🌐 在浏览器中打开", style="Guide.TButton",
                   command=_open_web).pack(side="left", padx=(8, 0))

        ttk.Label(main_frame, text="提示：选择文件夹后点击「扫描文件」检测合规文件，再点击「开始分析」",
                  style="Info.TLabel").pack(anchor="w", pady=(8, 0))

        # 日志区域（放在最后 pack，expand=True 填满剩余空间）——卡片包裹
        log_outer = tk.Frame(main_frame, bg="#ffffff", highlightbackground="#e0e4ea", highlightthickness=1, padx=10, pady=10)
        log_outer.pack(fill="both", expand=True, pady=(8, 0))
        tk.Label(log_outer, text="📋 运行日志", font=("Microsoft YaHei", 11, "bold"),
                 fg="#182B49", bg="#ffffff").pack(anchor="w", pady=(0, 4))
        log_frame = tk.Frame(log_outer, bg="#ffffff")
        log_frame.pack(fill="both", expand=True)

        log_text = scrolledtext.ScrolledText(log_frame, wrap="word", font=("Consolas", 10),
                                              bg="#1d2b3a", fg="#c9d1d9", insertbackground="#fff",
                                              relief="solid", borderwidth=1)
        log_text.pack(fill="both", expand=True)

        # 尝试绑定拖拽（需要 tkinterdnd2，非必需）
        try:
            root.drop_target_register("DND_Files")
            root.dnd_bind("<<Drop>>", lambda e: _on_drop(e, path_var, file_list_text, root))
        except Exception:
            pass

        root.mainloop()
    except Exception as e:
        print(f"GUI 启动失败: {e}")
        import traceback
        traceback.print_exc()
        return None
    return None


def _browse_folder(path_var, file_list_text, root):
    from tkinter import filedialog
    folder = filedialog.askdirectory(title="选择包含配置 Excel 和数据文件的文件夹", parent=root)
    if folder:
        path_var.set(folder)
        _scan_and_display(folder, file_list_text)


def _scan_and_display(folder, file_list_text):
    file_list_text.delete("1.0", "end")
    if not folder:
        file_list_text.insert("end", "请先选择文件夹\n")
        return
    results = _scan_files(folder)
    if not results:
        file_list_text.insert("end", f"在 {folder} 中未检测到合规的配置或数据文件\n")
        file_list_text.insert("end", "提示：配置文件需要包含「页码」「数据源」等列头\n")
        return

    config_count = sum(1 for r in results if r["type"] == "config")
    data_count = sum(1 for r in results if r["type"] == "data")
    file_list_text.insert("end", f"检测到 {config_count} 个配置文件, {data_count} 个数据文件\n\n")
    for r in results:
        icon = "[配置]" if r["type"] == "config" else "[数据]"
        file_list_text.insert("end", f"{icon} {r['name']}  —  {r['mode']}\n")

    # 多配置文件时显示下拉选择框
    config_files = [r for r in results if r["type"] == "config"]
    config_frame = getattr(file_list_text, "_config_select_frame", None)
    config_combo = getattr(file_list_text, "_config_combo", None)
    if config_frame is not None and config_combo is not None:
        if len(config_files) > 1:
            file_list_text._config_paths = [r["path"] for r in config_files]
            config_combo["values"] = [r["name"] for r in config_files]
            config_combo.current(0)
            config_frame.pack(fill="x", pady=(6, 0))
        else:
            file_list_text._config_paths = []
            config_combo["values"] = []
            config_combo_var = getattr(file_list_text, "_config_combo_var", None)
            if config_combo_var:
                config_combo_var.set("")
            config_frame.pack_forget()


def _on_drop(event, path_var, file_list_text, root):
    """处理拖拽事件"""
    data = event.data
    if not data:
        return
    # 拖拽可能包含引号，去掉
    folder = data.strip().strip('"').strip("'")
    if os.path.isdir(folder):
        path_var.set(folder)
        _scan_and_display(folder, file_list_text)


def _open_guide():
    import webbrowser
    guide_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "guide.html")
    webbrowser.open(f"file:///{guide_path.replace(os.sep, '/')}")


def _open_web():
    import webbrowser, subprocess, threading
    base_dir = os.path.dirname(os.path.abspath(__file__))
    server_path = os.path.join(base_dir, "server.py")
    if not os.path.exists(server_path):
        log_text.insert("end", "[错误] Web 服务文件不存在\n")
        log_text.see("end")
        return

    log_text.insert("end", "🌐 正在启动 Web 服务...\n")
    log_text.see("end")

    def launch():
        subprocess.run([sys.executable, server_path, "--no-browser"], cwd=base_dir)
    threading.Thread(target=launch, daemon=True).start()

    def open_browser():
        import time
        time.sleep(1.5)
        webbrowser.open("http://localhost:8899")
        log_text.insert("end", "   → 浏览器已打开，访问 http://localhost:8899\n")
        log_text.see("end")
    threading.Thread(target=open_browser, daemon=True).start()


def _run_analysis_thread(raw_args, log_text, run_btn, run_state):
    """在子线程中运行分析，不阻塞 GUI"""
    import threading
    redirector = _GuiLogRedirector(log_text)
    old_stdout = sys.stdout
    sys.stdout = redirector

    try:
        _dispatch_from_gui(raw_args)
    except SystemExit:
        pass
    except Exception as e:
        print(f"\n[错误] 分析过程异常: {e}")
        import traceback
        traceback.print_exc()
    finally:
        sys.stdout = old_stdout
        run_state["running"] = False
        log_text.after(0, lambda: run_btn.config(state="normal", text="开始分析"))
        log_text.after(0, lambda: log_text.insert("end", "\n[完成] 分析结束，可选择其他文件夹再次运行\n"))
        log_text.after(0, lambda: log_text.see("end"))


def _make_config_parser():
    """创建配置类模式（auto/ppt/pivot）的公共参数 parser。

    template 模式因位置参数语义不同（模板文件 vs 配置文件夹），单独处理。
    所有子命令共享 --pivot-file/--pivot、--data-dir、--output、--check 等参数，
    消除原先 4 套 parser 的重复定义。
    """
    p = argparse.ArgumentParser(add_help=False)
    p.add_argument("folder_or_config", nargs="?", default=None, help="文件夹路径或配置文件")
    p.add_argument("-c", "--config", default=None, help="配置文件路径")
    p.add_argument("-o", "--output", default=None, help="输出路径")
    p.add_argument("--data-dir", dest="data_dir", default=None,
                   help="数据文件所在目录（配置中数据源相对路径基于此目录，默认配置文件所在目录）")
    p.add_argument("--pivot-file", "--pivot", dest="pivot_file", default=None,
                   help="透视分析结果文件路径 (.xlsx)")
    p.add_argument("--check", action="store_true", help="仅校验配置，不执行")
    p.add_argument("--ts", dest="timestamp", action="store_true",
                   help="输出文件名追加时间戳（格式: _YYYYMMDD_HHMMSS）")
    return p


def _dispatch_from_gui(raw_args):
    """从 GUI 调用的分析入口（auto 模式）"""
    parser = _make_config_parser()
    args = parser.parse_args(raw_args)
    config_path = _resolve_config_path(args)
    _validate_required_args(args, "auto")
    detected = _detect_mode(config_path)
    data_dir = getattr(args, 'data_dir', None)
    output_path = getattr(args, 'output', None)
    add_ts = getattr(args, 'timestamp', False)
    if detected == "all":
        print("[信息] 检测到综合配置（PPT + 透视分析），执行全部模式")
        # auto 模式 -o 指定输出目录，生成 PPT + 透视 + HTML 三个文件
        out_dir = output_path if os.path.isdir(output_path) else os.path.dirname(output_path)
        os.makedirs(out_dir, exist_ok=True)
        base = os.path.basename(config_path).rsplit(".", 1)[0]
        ts_suffix = f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}" if add_ts else ""
        ppt_out = os.path.join(out_dir, f"{base}_报告{ts_suffix}.pptx")
        pivot_out = os.path.join(out_dir, f"{base}_分析{ts_suffix}.xlsx")
        html_out = os.path.join(out_dir, f"{base}_报告{ts_suffix}.html")
        _pivot_ret = _run_pivot_mode(config_path, pivot_out, data_dir=data_dir)
        _pivot_tasks = _pivot_ret[1] if isinstance(_pivot_ret, tuple) else None
        print()
        _run_ppt_mode(config_path, ppt_out, pivot_data_file=pivot_out, data_dir=data_dir)
        print()
        _run_html_from_pivot(pivot_out, html_out, tasks=_pivot_tasks)
        return
    print(f"[信息] 自动检测配置类型: {detected}")
    if detected == "unknown":
        print("[错误] 未识别的配置类型，请确认 Excel 包含正确的配置 Sheet")
        sys.exit(1)
    if detected == "pivot":
        pivot_out_path = _apply_timestamp(output_path, add_ts)
        _pivot_ret = _run_pivot_mode(config_path, pivot_out_path, data_dir=data_dir)
        _pivot_tasks = _pivot_ret[1] if isinstance(_pivot_ret, tuple) else None
        print()
        _run_html_from_pivot(pivot_out_path, tasks=_pivot_tasks)
    else:
        _run_ppt_mode(config_path, _apply_timestamp(output_path, add_ts), data_dir=data_dir)


def _run_template_mode(template_path, pivot_file=None, output_path=None, pivot_dir=None, mark_missing=True):
    """基于 PPT 模板和透视结果进行数据替换"""
    from src.template_filler import fill_template

    if not os.path.exists(template_path):
        print(f"[错误] 模板文件不存在: {template_path}")
        sys.exit(1)

    template_dir = os.path.dirname(os.path.abspath(template_path))

    # 自动查找的搜索目录：优先 pivot_dir，否则模板所在目录
    search_dir = os.path.abspath(pivot_dir) if pivot_dir else template_dir

    # 自动查找透视结果文件（--pivot 未指定或为 "latest"）
    if not pivot_file or pivot_file.strip().lower() == "latest":
        latest = _auto_find_latest_pivot(search_dir)
        if latest:
            pivot_file = latest
            ctime_local = datetime.fromtimestamp(os.path.getctime(pivot_file))
            print(f"[信息] 自动找到最新数据文件（搜索目录: {search_dir}）:")
            print(f"       文件名:   {os.path.basename(pivot_file)}")
            print(f"       创建时间: {ctime_local.strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"       路径:     {pivot_file}")
        else:
            print(f"[错误] 未找到可用数据文件（.xlsx），搜索目录: {search_dir}，请用 --pivot 指定")
            sys.exit(1)
    elif not os.path.exists(pivot_file):
        # 相对路径处理
        abs_pivot = os.path.join(template_dir, pivot_file) if not os.path.isabs(pivot_file) else pivot_file
        if os.path.exists(abs_pivot):
            pivot_file = abs_pivot
        else:
            print(f"[错误] 透视结果文件不存在: {pivot_file}")
            sys.exit(1)

    if not output_path:
        base_name = os.path.splitext(os.path.basename(template_path))[0]
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        # 输出到数据文件同目录（而非模板目录），方便数据与产出统一管理
        out_dir = os.path.dirname(os.path.abspath(pivot_file)) if pivot_file else template_dir
        output_path = os.path.join(out_dir, f"{base_name}_填充_{ts}.pptx")

    # 图片搜索目录：pivot_dir > pivot_file 所在目录 > 模板所在目录
    if pivot_dir:
        abs_image_dir = os.path.abspath(pivot_dir)
    elif pivot_file and os.path.exists(pivot_file):
        abs_image_dir = os.path.dirname(os.path.abspath(pivot_file))
    else:
        abs_image_dir = None

    fill_template(template_path, pivot_file, output_path, image_dir=abs_image_dir, mark_missing=mark_missing)


def main():
    if len(sys.argv) == 1:
        # 无参数：启动 GUI 模式
        _select_folder()
        return

    # CLI 模式：安装日志 Tee，所有输出同时写入 logs/last_run.log，方便远程排查
    _install_cli_log_tee()

    raw_args = sys.argv[1:]

    # 检测子命令
    mode = "auto"
    if raw_args and raw_args[0] in ("ppt", "pivot", "template"):
        mode = raw_args[0]
        raw_args = raw_args[1:]
    elif raw_args and raw_args[0] in ("-h", "--help"):
        mode = "help"

    # ===== help 模式：打印统一帮助 =====
    if mode == "help":
        print("Excel 统一分析工具 — PPT生成 & 透视分析 & 模板填充\n")
        print("用法:")
        print("  python main.py [子命令] -c <配置文件> --data-dir <数据目录> -o <输出路径> [选项]\n")
        print("子命令:")
        print("  (无)      自动检测配置类型并执行（综合配置时先透视后PPT）")
        print("  ppt       生成PPT报告")
        print("  pivot     透视分析")
        print("  template  基于PPT模板填充数据\n")
        print("公共必填参数 (ppt/pivot/auto):")
        print("  -c, --config <路径>      配置文件路径（位置参数也接受配置文件）")
        print("  --data-dir <目录>        数据文件所在目录（数据源相对路径基于此目录）")
        print("  -o, --output <路径>      输出路径（auto模式为输出目录）\n")
        print("公共可选参数:")
        print("  --pivot-file <路径>      透视分析结果文件路径 (.xlsx)")
        print("  --check                  仅校验配置，不执行")
        print("  --ts                     输出文件名追加时间戳（格式: _YYYYMMDD_HHMMSS）\n")
        print("template 必填参数:")
        print("  <模板.pptx>              PPT模板文件路径 (位置参数)")
        print("  --image-dir <目录>       图片和透视结果文件的搜索目录")
        print("  -o, --output <路径>      输出PPT路径\n")
        print("template 可选参数:")
        print("  --pivot-file <路径>      透视分析结果文件路径（支持 --pivot 别名，填 latest 自动找最新）")
        print("  --ts                     输出文件名追加时间戳")
        print("  --no-mark                关闭缺失标注（默认未替换占位符以黄底[缺失:...]标注）\n")
        print("示例:")
        print("  python main.py pivot -c 配置.xlsx --data-dir 数据目录 -o 结果.xlsx")
        print("  python main.py pivot -c 配置.xlsx --data-dir 数据目录 -o 结果.xlsx --ts")
        print("  python main.py ppt -c 配置.xlsx --data-dir 数据目录 -o 报告.pptx")
        print("  python main.py -c 配置.xlsx --data-dir 数据目录 -o 输出目录/")
        print("  python main.py template 模板.pptx --image-dir 数据目录 -o 输出.pptx")
        print("  python main.py template 模板.pptx --pivot-file 透视结果.xlsx --image-dir 数据目录 -o 输出.pptx")
        return

    # ===== template 模式：位置参数为模板文件，单独解析 =====
    if mode == "template":
        tpl_parser = argparse.ArgumentParser(add_help=False)
        tpl_parser.add_argument("template_path", help="PPT模板文件路径 (.pptx)")
        tpl_parser.add_argument("--pivot-file", "--pivot", dest="pivot_file", default=None,
                                help="透视分析结果文件路径 (.xlsx)，填 latest 自动找最新 xlsx")
        tpl_parser.add_argument("--image-dir", "--pivot-dir", dest="image_dir", default=None,
                                help="图片和透视结果文件的搜索目录")
        tpl_parser.add_argument("-o", "--output", default=None, help="输出PPT路径")
        tpl_parser.add_argument("--ts", dest="timestamp", action="store_true",
                                help="输出文件名追加时间戳（格式: _YYYYMMDD_HHMMSS）")
        tpl_parser.add_argument("--no-mark", dest="mark_missing", action="store_false", default=True,
                                help="关闭缺失标注（默认未替换的占位符以黄底[缺失:...]标注）")
        tpl_args = tpl_parser.parse_args(raw_args)
        # template 模式必填 --image-dir 和 -o
        _require_arg(tpl_args.image_dir, "--image-dir <图片/数据搜索目录>", "template")
        _require_arg(tpl_args.output, "-o/--output <输出PPT路径>", "template")
        _run_template_mode(tpl_args.template_path, tpl_args.pivot_file,
                           _apply_timestamp(tpl_args.output, getattr(tpl_args, 'timestamp', False)),
                           tpl_args.image_dir,
                           mark_missing=getattr(tpl_args, 'mark_missing', True))
        return

    # ===== auto/ppt/pivot 模式：共享公共参数 parser =====
    parser = _make_config_parser()
    args = parser.parse_args(raw_args)
    config_path = _resolve_config_path(args)
    _validate_required_args(args, mode)
    output_path = args.output
    validate_only = getattr(args, 'check', False)
    data_dir = getattr(args, 'data_dir', None)
    add_ts = getattr(args, 'timestamp', False)

    if mode == "auto":
        detected = _detect_mode(config_path)
        if detected == "all":
            print("[信息] 检测到综合配置（PPT + 透视分析），执行全部模式")
            # auto 模式 -o 指定输出目录，生成 PPT + 透视 + HTML 三个文件
            out_dir = output_path if os.path.isdir(output_path) else os.path.dirname(output_path)
            os.makedirs(out_dir, exist_ok=True)
            base = os.path.basename(config_path).rsplit('.', 1)[0]
            ts_suffix = f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}" if add_ts else ""
            ppt_out = os.path.join(out_dir, f"{base}_报告{ts_suffix}.pptx")
            pivot_out = os.path.join(out_dir, f"{base}_分析{ts_suffix}.xlsx")
            html_out = os.path.join(out_dir, f"{base}_报告{ts_suffix}.html")
            _pivot_ret = _run_pivot_mode(config_path, pivot_out, validate_only=validate_only, data_dir=data_dir)
            _pivot_tasks = _pivot_ret[1] if isinstance(_pivot_ret, tuple) else None
            if not validate_only:
                print()
                _run_ppt_mode(config_path, ppt_out, pivot_data_file=pivot_out, data_dir=data_dir)
                print()
                _run_html_from_pivot(pivot_out, html_out, tasks=_pivot_tasks)
            return
        print(f"[信息] 自动检测配置类型: {detected}")
        if detected == "unknown":
            print("[错误] 未识别的配置类型，请确认 Excel 包含正确的配置 Sheet")
            sys.exit(1)
        mode = detected

    if mode == "pivot":
        pivot_out_path = _apply_timestamp(output_path, add_ts)
        _pivot_ret = _run_pivot_mode(config_path, pivot_out_path,
                        validate_only=validate_only, data_dir=data_dir)
        _pivot_tasks = _pivot_ret[1] if isinstance(_pivot_ret, tuple) else None
        if not validate_only:
            print()
            _run_html_from_pivot(pivot_out_path, tasks=_pivot_tasks)
    else:
        _run_ppt_mode(config_path, _apply_timestamp(output_path, add_ts),
                      pivot_data_file=getattr(args, "pivot_file", None),
                      validate_only=validate_only, data_dir=data_dir)


if __name__ == "__main__":
    main()
